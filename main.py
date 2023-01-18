import pathlib

from flask import Flask, url_for, redirect, render_template, request, session, send_file
import mysql.connector
from flask import flash
from config import DevConfig
from werkzeug.security import generate_password_hash
from werkzeug.security import check_password_hash
import userform
from datetime import datetime
import string
import random
from werkzeug.utils import secure_filename
import os
import time
#Reporte en pdf
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER
from reportlab.lib import colors
#Reporte en excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font


app = Flask(__name__)

app.config.from_object(DevConfig)


folder = os.path.abspath("./media/")
ext_p = set(["pdf"])
ext_c = set(["xls", "xlsx", "xlsm", "xlsb", "xltx"])
app.config['folder'] = folder

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="proyecto"
)
global bcurso
bcurso =0
global user_datos
user_datos=[]
global vector
vector = []
global vector2
vector2 = []
global cursos
cursos=[]
global idmaterias
idmaterias = 0
global alumnos
alumnos = []
global balumno
balumno = 0
global idcurso
idcurso = 0
global ci
ci = 0
global tel
tel = 0
global band
band = 0
global tipoins
tipoins = 0
@app.before_request
def before_request(): #antes de cargar la pag
  # Verificar si existe una sesion o no, en algun punto de acceso
  if 'username' not in session and request.endpoint in ['bienvenidoalumno', 'bienvenidoadmin', 'bienvenidoprofe',
                                                        'proceso', 'verproceso', 'vermaterias', 'inscribirdir',
                                                        'sacarmat', 'inscurmat', 'inscribirprof', 'inscmatxprof',
                                                        'misconfig', 'misdatos', 'cerrarsesion', 'procesoss',
                                                        'alumnosproceso', 'modificarproceso2', 'modproceso', 'miscursos']:

    return redirect(url_for('login'))
  if 'username' in session and request.endpoint in ['login', 'registro', 'recuperar']:
    datos = session['username']
    if datos[7] == 1:  # alumnos
      return redirect(url_for('vermaterias'))
    if datos[6] == 2:  # profesores
      return redirect(url_for('bienvenidoprofe'))
    if datos[6] == 3:  # Admin
      return redirect(url_for('bienvenidoadmin'))
  if 'username' in session:
    datos = session['username']
    if datos[7] == 1:
      if request.endpoint in ['bienvenidoadmin', 'inscribirdir','sacarmat', ' inscurmat', 'inscribirprof','proceso',
                              'procesoss', 'alumnosproceso', 'inscmatxprof', 'modproceso', 'modificarproceso2','miscursos']:
        return redirect(url_for('vermaterias'))
    if datos[6] == 2:
      if request.endpoint in ['verproceso','vermaterias ', 'bienvenidoadmin', 'inscribirdir', 'sacarmat',
                              'inscurmat', 'inscribirprof', 'inscmatxprof']:
        return redirect(url_for('bienvenidoprofe'))
    if datos[6] == 3:
      if request.endpoint in ['verproceso', 'vermaterias', 'proceso', 'procesoss', 'alumnosproceso', 'modproceso',
                              'modificarproceso2','miscursos']:
        return redirect(url_for('bienvenidoadmin'))


@app.route('/bienvenidoprofe')
def bienvenidoprofe():
  datos = session['username']
  print(datos)
  return redirect(url_for('miscursos'))

#Parte de vista de alumnos materias
@app.route('/bienvenidoalumno')
def vermaterias():
  #Alertas
  global band
  if band == 6:
    flash("","noa")
    band = 0
  datos = session['username']
  print(datos)
  #Materias el cual esta matriculado el alumno
  mycursor = mydb.cursor()
  sql = "SELECT matxalum.id_materia,des_m, des_c, sec_c, ano_m, des_e, id_profesor FROM matxalum, materias, cursos, enfasis" \
        " WHERE id_alumno = %s and cursos.id_curso =%s and matxalum.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis"
  val = [datos[0], datos[4]]
  mycursor.execute(sql, val)
  data = mycursor.fetchall()
  print(data)
  return render_template('materias.html', datos=datos, materias=data)

@app.route('/bienvenidoadmin')
def bienvenidoadmin():
  datos = session['username']
  print(datos)
  return render_template('directorview.html', datos = datos)


@app.route('/', methods = ['GET', 'POST']) #Login
def login():
  global band
  user = userform.User(request.form)
  username = user.username.data
  password = user.password.data
  t_u = 0
  if band == 1:
    band = 0
    flash("","creado")
  #print(username)
  if request.method == 'POST':
    mycursor = mydb.cursor()
    mycursor.execute('SELECT * FROM user WHERE username = %(username)s', {'username': username})
    data = mycursor.fetchall()
    if data:
      userdata = data[0]
      #print(userdata)
      if userdata:
        passcheck = userdata[3]
        #print(passcheck)
      if userdata and check_password_hash(passcheck, password):
        global tipo
        tipo = userdata[4]
        if userdata[4] == 1:
          sql = "SELECT * FROM alumnos WHERE id_user = %s"
          val = [userdata[0]]
          mycursor.execute(sql, val)
          data = mycursor.fetchall()
          datos = data[0]
          session['tipo_u'] = userdata[4]
          session['username'] = datos
          #flash("Correcto", "success")
          time.sleep(1)
          return redirect(url_for('vermaterias'))
        if userdata[4] == 2:
          sql = "SELECT * FROM profesores WHERE id_user = %s"
          val = [userdata[0]]
          mycursor.execute(sql, val)
          data = mycursor.fetchall()
          datos = data[0]
          session['username'] = datos
          #flash("Correcto", "success")
          return redirect(url_for('bienvenidoprofe'))
        if userdata[4] == 3:
          sql = "SELECT * FROM admin WHERE id_user = %s"
          val = [userdata[0]]
          mycursor.execute(sql, val)
          data = mycursor.fetchall()
          datos = data[0]
          session['username'] = datos
          #flash("Correcto", "success")
          return redirect(url_for('bienvenidoadmin'))
      else:
        flash("Contra", "error_p")
    else:
      flash("Not found", "error_n")

  return render_template('login.html', user = user)

@app.route('/registro', methods = ['GET', 'POST']) #Registro, falta ver username cambiar por cedula ❌
def registro():
  user = userform.User(request.form)
  global band
  #pasw = user.password.data
  if request.method == 'POST':
    password = createpassword(user.password.data)
    create = False
    ci = int(user.ci.data)
    print(ci)
    mycursor = mydb.cursor()
    sql = "SELECT * FROM alumnos WHERE ci_a = %s"
    val = [ci]
    mycursor.execute(sql, val)
    data = mycursor.fetchall()
    t_u = 0
    if data:
      t_u = 1 #Alumnos
      print("alumno encontrado")
      create = True
    else:
      sql = "SELECT * FROM profesores WHERE ci_p = %s"
      val = [ci]
      mycursor.execute(sql, val)
      data = mycursor.fetchall()
      if data:
        t_u = 2  #Profesores
        print("profesor encontrado")
        create = True
      else:
        sql = "SELECT * FROM admin WHERE ci_ad = %s"
        val = [ci]
        mycursor.execute(sql, val)
        data = mycursor.fetchall()
        if data:
          t_u = 3  # Admin
          print("admin encontrado")
          create = True
    if data:
      print("a")
      data = data[0]
      print(data)
      if data[6]:
        create = False
        flash("Ya tiene cuenta", "existe_ci")
        print("test")
    else:
        create = False
        flash("Error", "ci_e")
    if t_u > 0:
      sql = "SELECT * FROM user WHERE username = %s"
      val = [user.username.data]
      mycursor.execute(sql, val)
      coc = mycursor.fetchall()
      if coc:
        flash("User existente", "existe_us")
        print("a")
        create = False
      else:
        sql = "SELECT * FROM user WHERE email = %s"
        val = [user.email.data]
        mycursor.execute(sql, val)
        coe = mycursor.fetchall()
        if coe:
          flash("Correo existente", "existe_co")
          create = False
    if create == True:
      if user.password.data == user.confirmpassword.data:
        mycursor.execute('INSERT INTO user (username, email, password, tipo_u) VALUES (%s, %s, %s, %s)', (user.username.data,
                                                                                         user.email.data, password, t_u))
        mydb.commit()
        sql = "SELECT * FROM user WHERE username = %s"
        val = [user.username.data]
        mycursor.execute(sql, val)
        aux = mycursor.fetchall()
        aux = aux[0]
        print(aux)
        if t_u == 1:
          id = aux[0]
          sql = "UPDATE alumnos SET id_user= %s WHERE ci_a = %s"
          val = (id,ci)
          mycursor.execute(sql, val)
          mydb.commit()
        if t_u == 2:
          id = aux[0]
          sql = "UPDATE profesores SET id_user= %s WHERE ci_p = %s"
          val = (id, ci)
          mycursor.execute(sql, val)
          mydb.commit()
        else:
          id = aux[0]
          sql = "UPDATE admin SET id_user= %s WHERE ci_ad = %s"
          val = (id, ci)
          mycursor.execute(sql, val)
          mydb.commit()
        band = 1
        #flash("Correcto", "success")
        return redirect(url_for('login'))
      else:
        print("e")
        flash("Error", "psw")
  return render_template('register.html', user = user)

@app.route('/recuperar', methods = ['GET', 'POST']) #Recuperar cuenta ❌
def recuperar():
  user = userform.User(request.form)
  return render_template('recuperar.html', user=user)

#Parte de Proceso de Alumnos, Carga de trabajos
@app.route('/proceso', methods = ['GET', 'POST']) #Procesos ✔
def proceso():
  global bcurso
  global balumno
  datos = session['username']
  #print(datos)
  mycursor = mydb.cursor()
  sql = "SELECT * FROM matxpro WHERE id_profesor = %s"
  val = [datos[0]]
  mycursor.execute(sql, val)
  profe = mycursor.fetchall()
  #print(profe) #ID, Id Materia, Id profe, Carga horaria, Id curso
  profemat = []
  profemati = []
  cantm = 0
  global user_datos
  global tipoins
  #Mensajes flash
  global band
  if band == 1:
    flash("car","car")
    band = 0
  global cursos
  global alumnos
  for x in range(0, len(profe)):
    #print(x)
    aux = profe[x]
    sql = "SELECT * FROM materias WHERE id_materia = %s"
    val = [aux[1]]
    mycursor.execute(sql, val)
    aux = mycursor.fetchall()
    aux = aux[0]
    #print(aux)
    #print(x)
    if x == 0:
      profemat.append(aux)
      #print(profemat)
      cantm += 1
      #print(x)
      #print(cantm)
    elif profemat[x-cantm-1] != aux:
      #print(cantm)
      cantm += 1
      profemat.append(aux)
      #print(profemat)
  #print(profemat)
  task = userform.Task(request.form)
  if request.method == 'POST' and task.validate(): #ID prof = datos[0],
    #Informacion del trabajo
    print("Info: ")
    nombre_t = task.nombre.data
    print(nombre_t)
    tipo_t = request.form.get(('tipo_t'))
    print(tipo_t)
    puntaje_t = task.puntaje.data
    puntaje_t = int(puntaje_t)
    print(puntaje_t)
    etapa_t = request.form.get(('etapa_t'))
    print(etapa_t)
    global idcurso
    print(idcurso)
    # Informacion del trabajo
    print("Indicadores: ")
    indicadores = []
    puntaje_i = []
    create = True
    for x in range(1, 11):
      if x == 1:
        i1 = request.form.get(('i1'))
        p1 = request.form.get(('p1'))
        if i1 and p1:
          indicadores.append(i1)
          puntaje_i.append(p1)
      if x == 2:
        i2 = request.form.get(('i2'))
        p2 = request.form.get(('p2'))
        if i2 and p2:
          indicadores.append(i2)
          puntaje_i.append(p2)
      if x == 3:
        i3 = request.form.get(('i3'))
        p3 = request.form.get(('p3'))
        if i3 and p3:
          indicadores.append(i3)
          puntaje_i.append(p3)
      if x == 4:
        i4 = request.form.get(('i4'))
        p4 = request.form.get(('p4'))
        if i4 and p4:
          indicadores.append(i4)
          puntaje_i.append(p4)
      if x == 5:
        i5 = request.form.get(('i5'))
        p5 = request.form.get(('p5'))
        if i5 and p5:
          indicadores.append(i5)
          puntaje_i.append(p5)
      if x == 6:
        i6 = request.form.get(('i6'))
        p6 = request.form.get(('p6'))
        if i6 and p6:
          indicadores.append(i6)
          puntaje_i.append(p6)
      if x == 7:
        i7 = request.form.get(('i7'))
        p7 = request.form.get(('p7'))
        if i7 and p7:
          indicadores.append(i7)
          puntaje_i.append(p7)
      if x == 8:
        i8 = request.form.get(('i8'))
        p8 = request.form.get(('p8'))
        if i8 and p8:
          indicadores.append(i8)
          puntaje_i.append(p8)
      if x == 9:
        i9 = request.form.get(('i9'))
        p9 = request.form.get(('p9'))
        if i9 and p9:
          indicadores.append(i9)
          puntaje_i.append(p9)
      if x == 10:
        i10 = request.form.get(('i10'))
        p10 = request.form.get(('p10'))
        if i10 and p10:
          indicadores.append(i10)
          puntaje_i.append(p10)
    tt = 0
    for x in range(0, len(puntaje_i)):
      aux = int(puntaje_i[x])
      tt = tt + aux
    print(indicadores)
    print(puntaje_i)
    print(tt)
    #Comprobar si el puntaje de los indicadores es mayor que el puntaje del trabajo o si no esta completo
    if tt > puntaje_t:
      create = False
      flash("pun_m", "pun_m")
      #flash("")
    if tt != puntaje_t:
      create = False
      print("El puntaje no coincide")
      flash("pun_m", "pun_m")
    global idmaterias
    fecha = datetime.now()
    fecha = datetime.strftime(fecha, '%Y/%m/%d')
    if idmaterias != 0:
      if idcurso !=0:
        if create == True:
          clavet = crearclavet()
          #Guardamos el trabajo
          mycursor = mydb.cursor()
          mycursor.execute(
            'INSERT INTO trabajos (fec_t, des_t, pun_t, id_profesor, id_materia, tipo_t, clave_t, id_curso, etapa) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)',
            (fecha, nombre_t, puntaje_t, datos[0], idmaterias, tipo_t, clavet, idcurso, etapa_t))
          mydb.commit()
          # Saca el id del trabajo
          sql = "SELECT id_trabajo, id_curso FROM trabajos WHERE id_materia = %s and id_profesor = %s and fec_t = %s and des_t= %s and pun_t = %s and tipo_t = %s and clave_t = %s"
          val = [idmaterias, datos[0], fecha, task.nombre.data, task.puntaje.data, tipo_t, clavet]
          mycursor.execute(sql, val)
          idtra = mycursor.fetchall()
          idtra = idtra[0]
          print(idtra[0])
          #Guarda los indicadores
          for x in range(0, len(puntaje_i)):
            aux = indicadores[x]
            aux2 = puntaje_i[x]
            #print(aux)
            #print(aux2)
            mycursor.execute(
              'INSERT INTO indicadores (des_i, pun_i, id_trabajo) VALUES (%s, %s, %s)',
              (aux, aux2, idtra[0]))
            mydb.commit()
            print(idtra[0])
          return redirect(url_for('cargarpuntaje', id = idtra[0]))
      else:
        # Errorsito
        flash("Error", "cur")
    else:
      # Errorsito
      flash("Error", "mat")


  return render_template('procesotest.html', task = task, datos = datos, materias = profemat,bcurso = bcurso, cursos = cursos,
                         alumnos = alumnos, balumno = balumno )

@app.route('/procesoss', methods = ['POST']) #Parte de procesos, aca quitamos la materia ✔
def procesoss():
  if request.method == 'POST':
    id_mat = request.form.get(('idmat'))
    if id_mat:
      global bcurso
      bcurso = 1
    print(id_mat)
    datos = session['username']
    mycursor = mydb.cursor()
    print(datos[0])
    print(id_mat)
    sql = "SELECT * FROM matxpro WHERE id_profesor = %s and id_materia =%s"
    val = [datos[0], id_mat]
    global idmaterias
    idmaterias = id_mat
    mycursor.execute(sql, val)
    data = mycursor.fetchall()
    print("Cursos con esa materia:")
    print(data)
    #print(cursos[0])
    #print(len(cursos))
    global balumno
    global cursos
    cursos = []
    if cursos:
      balumno = 1
    for x in range(0, len(data)):
      print(x)
      aux = data[x]
      mycursor = mydb.cursor()
      sql = "SELECT id_curso, des_c, sec_c,des_e FROM cursos,enfasis WHERE id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
      val = [aux[3]]
      mycursor.execute(sql, val)
      aux = mycursor.fetchall()
      aux = aux[0]
      cursos.append(aux)
    print("Cursos procesados:")
    print(cursos)
  return redirect(url_for('proceso'))

@app.route('/alumnos', methods = ['POST']) #Parte de procesos, quitamos los cursos y los alumnos relacionados con la materia ✔
def alumnosproceso():
  if request.method == 'POST':
    id_curso = request.form.get(('idcurso'))
    datos = session['username']
    global balumno
    balumno = 1
    global cursos
    global idmaterias
    mycursor = mydb.cursor()
    sql = "SELECT alumnos.id_alumno,nmb_a, ape_a FROM alumnos, " \
          "matxalum WHERE matxalum.id_curso =%s and alumnos.id_alumno = matxalum.id_alumno " \
          "and matxalum.id_materia = %s"
    val = [id_curso, idmaterias]
    mycursor.execute(sql, val)
    global idcurso
    idcurso = id_curso
    data = mycursor.fetchall()
    data = sorted(data, key=lambda data: data[2])  # Ordena por orden alfabetico
    print(id_curso)
    print(idmaterias)
    print("Alumnos con esta materia:")
    print(data)
    global alumnos
    alumnos = data
  return redirect(url_for('proceso'))

@app.route('/cargarpuntaje/<int:id>', methods =  ['GET', 'POST'])# Carga de trabajo y indicadores ✔
def cargarpuntaje(id):
  datos = session['username']
  idtra = id
  print(idtra)
  # Saca el trabajo
  mycursor = mydb.cursor()
  sql = "SELECT id_trabajo, des_t, pun_t, id_materia, id_curso, id_profesor, etapa FROM trabajos WHERE id_trabajo = %s"
  val = [idtra]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  trabajos = trabajos[0]
  print(trabajos)
  # Saca los indicadores
  sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
  val = [idtra]
  mycursor.execute(sql, val)
  indicadores = mycursor.fetchall()
  print(indicadores)
  # Saca los alumnos
  mycursor = mydb.cursor()
  sql = "SELECT alumnos.id_alumno,nmb_a, ape_a FROM alumnos, " \
        "matxalum WHERE matxalum.id_curso =%s and alumnos.id_alumno = matxalum.id_alumno " \
        "and matxalum.id_materia = %s"
  val = [trabajos[4], trabajos[3]]
  mycursor.execute(sql, val)
  alumnos = mycursor.fetchall()
  alumnos = sorted(alumnos, key=lambda alumnos: alumnos[2])
  print(alumnos)
  # Saca el curso:
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [trabajos[4]]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  #Comprobar si se cargo correctamente
  create = True
  if request.method == 'POST':
    puntajes = []
    pl = 0
    # Comprobar que se hayan cargado todos los indicadores
    for x in range(0, len(alumnos)):
      for y in range(1, len(indicadores) + 1):
        aux = request.form.getlist(('{a}{i}'.format(a=x + 1, i=y)))
        if not aux:
          create = False
    # Solo si esta cargado correctamente
    if create == True:
      #Carga de indicador por alumno, puntaje logrado por indicador y acumular puntaje total logrado
      for x in range(0, len(alumnos)):
        aux1 = alumnos[x]
        #print(aux1)
        for y in range(1, len(indicadores) + 1):
          aux2 = indicadores[y - 1]
          #print(aux2)
          if y == 1:
            pl = 0
            #print("Alumno {}".format(alumnos[x]))
          aux = request.form.getlist(('{a}{i}'.format(a = x + 1, i = y)))
          #print('{a}{i}'.format(a=x + 1, i=y))
          #print(aux)
          p = int(aux[0])
          #print(p)
          pl +=int(aux[0])
          mycursor.execute(
            'INSERT INTO indxalum (id_indicador, id_trabajo, id_alumno, pun_l) VALUES (%s, %s, %s, %s)',
            (aux2[0], trabajos[0], aux1[0], p))
          mydb.commit()
          if y == len(indicadores):
            puntajes.append(pl)
      print(puntajes)
      fecha = datetime.now()
      fecha = datetime.strftime(fecha, '%Y')
      #print(fecha)
      for x in range(0, len(alumnos)):
        aux1 = alumnos[x]
        aux2 = puntajes[x]
        #Cargando el trabajo al alumno con su puntaje logrado
        mycursor.execute(
          'INSERT INTO traxalum (id_alumno, pun_l, fec_t, id_trabajo) VALUES (%s, %s, %s, %s)',
          (aux1[0], aux2, fecha, trabajos[0]))
        mydb.commit()
        #Sumando a su puntaje acumulado
        if trabajos[6] == 1: #Primera Etapa..
          sql = "UPDATE matxalum SET pun_ac = (pun_ac + %s) WHERE id_alumno = %s and id_materia = %s and id_curso = %s and id_profesor = %s"
          val = (aux2, aux1[0],  trabajos[3], trabajos[4], trabajos[5])
          mycursor.execute(sql, val)
          mydb.commit()
        if trabajos[6] == 2:
          sql = "UPDATE matxalum SET pun_ac2 = (pun_ac2 + %s) WHERE id_alumno = %s and id_materia = %s and id_curso = %s and id_profesor = %s"
          val = (aux2, aux1[0], trabajos[3], trabajos[4], trabajos[5])
          mycursor.execute(sql, val)
          mydb.commit()
      #Mandar mensaje flash..
      global band
      band = 1
      return redirect(url_for('proceso'))
    else:
      flash("error","error")
  return render_template('cargartrabajo.html', datos = datos, trabajo = trabajos, alumnos = alumnos, indicadores= indicadores,
                         cursos = cursos)


@app.route('/vermateria/<string:id>') #Ver proceso de la materia seleccionada.
def verproceso(id):
  datos = session['username']
  #print(id) #Pimer digito materia, segundo id del profe
  x = [int(a) for a in str(id)]
  #print(x)
  #Trabajos primera etapa
  mycursor = mydb.cursor()
  sql = "SELECT trabajos.id_trabajo,des_t, trabajos.fec_t, pun_t, pun_l FROM trabajos, traxalum WHERE trabajos.id_trabajo = traxalum.id_trabajo and id_materia = %s and id_alumno = %s and etapa = %s"
  val = [id, datos[0], 1]
  mycursor.execute(sql, val)
  data = mycursor.fetchall()
  print(data)
  #Trabajos segunda etapa
  sql = "SELECT trabajos.id_trabajo,des_t, trabajos.fec_t, pun_t, pun_l FROM trabajos, traxalum WHERE trabajos.id_trabajo = traxalum.id_trabajo and id_materia = %s and id_alumno = %s and etapa = %s"
  val = [id, datos[0], 2]
  mycursor.execute(sql, val)
  data2 = mycursor.fetchall()
  print(data2)
  sql = "SELECT nmb_p, ape_p, des_m FROM profesores, matxalum, materias WHERE matxalum.id_materia = %s and matxalum.id_profesor = profesores.id_profesor " \
        "and matxalum.id_materia = materias.id_materia"
  val = [id]
  mycursor.execute(sql, val)
  profesor = mycursor.fetchall()
  print(profesor)
  #Sumatoria de el puntaje del alumno
  #Segunda Etapa
  if data2:
    profesor = profesor[0]
    print(profesor)
    suml2 = 0
    sumt2 = 0
    for x in range(0, len(data2)):
      aux = data2[x]
      pl = aux[4]
      pt = aux[3]
      suml2 += pl
      # print(suml)
      sumt2 += pt
    print(suml2)
  #Primera etapa
  if data:
    #print(data)
    suml = 0
    sumt = 0
    for x in range(0, len(data)):
      aux = data[x]
      pl=aux[4]
      pt=aux[3]
      suml+= pl
      #print(suml)
      sumt+= pt
  #En caso de no tener proceso tira una alerta y redirige
  else:
    flash("no", "nom")
    return redirect(url_for('vermaterias'))
  return render_template('vermateria.html', datos=datos, procesos = data, suml = suml, sumt = sumt, data = profesor, data2 = data2, suml2 = suml2, sumt2 = sumt2)

@app.route('/vermitrabajo/<int:id>')
def vermitrabajo(id):
  datos = session['username']
  #print(datos)
  print(id)
  #Saca el trabajo
  mycursor = mydb.cursor()
  sql = "SELECT des_t, pun_t, fec_t, id_materia FROM trabajos WHERE trabajos.id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  data = mycursor.fetchall()
  data = data[0]
  print(data)
  #Saca los indicadores
  mycursor = mydb.cursor()
  sql = "SELECT indxalum.id_indicador, des_i, indxalum.id_trabajo, pun_l, pun_i FROM indxalum, indicadores WHERE indxalum.id_trabajo = %s and id_alumno = %s and " \
        "indxalum.id_indicador = indicadores.id_indicador"
  val = [id, datos[0]]
  mycursor.execute(sql, val)
  indi = mycursor.fetchall()
  print(indi)
  return render_template('vermitrabajo.html', datos=datos, data=data, indicadores = indi)

@app.route('/verasistenciaal/<int:id>', methods =  ['GET', 'POST'])
def verasistenciaal(id):
  datos = session['username']
  print(datos)
  mycursor = mydb.cursor()
  #Datos de la disciplina
  sql = "SELECT DISTINCT nmb_p, ape_p, des_m FROM profesores, matxalum, materias WHERE matxalum.id_materia = %s and matxalum.id_profesor = profesores.id_profesor " \
        "and matxalum.id_materia = materias.id_materia"
  val = [id]
  mycursor.execute(sql, val)
  profesor = mycursor.fetchall()
  print(profesor)
  #Datos de su asistencia
  sql = "SELECT id_asisalum, fecha, asistio FROM asistenciaalum WHERE id_alumno = %s and id_materia = %s and id_curso = %s ORDER BY asistenciaalum.fecha DESC LIMIT 7"
  val = [datos[0], id, datos[4]]
  mycursor.execute(sql, val)
  asistencias = mycursor.fetchall()
  print(asistencias)
  #Si no tienes asistencias tira alerta y redirige
  if not asistencias:
    global band
    band = 6
    return redirect(url_for('vermaterias'))
  if request.method == "POST":
    #Filtro de tiempo..
    filtro = int(request.form.get("idfiltro"))
    if filtro == 1:
      return redirect(url_for('verasistenciaal', id= id))
    if filtro == 2:
      sql = "SELECT id_asisalum, fecha, asistio FROM asistenciaalum WHERE id_alumno = %s and id_materia = %s and id_curso = %s ORDER BY asistenciaalum.fecha DESC LIMIT 31"
      val = [datos[0], id, datos[4]]
      mycursor.execute(sql, val)
      asistencias = mycursor.fetchall()
      print(asistencias)
      return render_template('verasistenciaal.html', datos=datos, data=profesor[0], asistencias=asistencias, filtro=filtro, id=id)
    if filtro == 3:
      sql = "SELECT id_asisalum, fecha, asistio FROM asistenciaalum WHERE id_alumno = %s and id_materia = %s and id_curso = %s ORDER BY asistenciaalum.fecha DESC"
      val = [datos[0], id, datos[4]]
      mycursor.execute(sql, val)
      asistencias = mycursor.fetchall()
      print(asistencias)
      return render_template('verasistenciaal.html', datos=datos, data=profesor[0], asistencias=asistencias, filtro=filtro, id=id)
  return render_template('verasistenciaal.html', datos=datos, data = profesor[0], asistencias=asistencias, filtro = 1, id = id)

@app.route('/inscribiral', methods = ['GET', 'POST']) #Inscripcion de alumnos. Parte de la Carga
def inscribirdir():
  alumno = userform.Alumno(request.form)
  datos = session['username']
  global band
  global tipoins
  tipoins = 1 #Alumnos
  print(band)
  if band == 1:
    flash("", "si")
    band = 0
  if request.method == "POST":
    cedu = request.files["cedu"]
    ante = request.files["ante"]
    auto = request.files["auto"]
    ficha = request.files["ficha"]
    fecha = (request.form['fec_n'])
    print(fecha)
    if fecha:
      # Parte Cedula.. Preguntar si conviene hacer funcion o q
      if "cedu" not in request.files:
        #print("No envio nada")
        pass
      elif cedu.filename == "":
        #print("No mando nada")
        pass
      elif cedu and archpermi(cedu.filename):
        filename = secure_filename(cedu.filename)
        extc = filename.split('.')
        #print(extc)
        #print(filename)
        #print(alumno.num_c.data)
        filename = "cedula_" + alumno.num_c.data + "." + extc[1]
        print(filename)
        cedu.save(os.path.join(app.config["folder"], filename))
        flash("", "") #Ya guarda el archivo
      else:
        print("archivo no permitido")
      #Parte Antecedentes
      if "ante" not in request.files:
        #print("No envio nada")
        pass
      elif ante.filename == "":
        #print("No mando nada")
        pass
      elif ante and archpermi(ante.filename):
        filename = secure_filename(ante.filename)
        extc = filename.split('.')
        print(extc)
        print(filename)
        print(alumno.num_c.data)
        filename = "antecedente_" + alumno.num_c.data + "." + extc[1]
        print(filename)
        ante.save(os.path.join(app.config["folder"], filename))
        flash("", "")  # Ya guarda el archivo
      else:
        print("archivo no permitido")
      #Parte Autorizacion de Padres
      if "auto" not in request.files:
        #print("No envio nada")
        pass
      elif auto.filename == "":
        #print("No mando nada")
        pass
      elif auto and archpermi(auto.filename):
        filename = secure_filename(auto.filename)
        extc = filename.split('.')
        print(extc)
        print(filename)
        print(alumno.num_c.data)
        filename = "autorizacion_" + alumno.num_c.data + "." + extc[1]
        print(filename)
        auto.save(os.path.join(app.config["folder"], filename))
        flash("", "")  # Ya guarda el archivo
      else:
        print("archivo no permitido")
      # Parte de ficha medica
      if "ficha" not in request.files:
        # print("No envio nada")
        pass
      elif ficha.filename == "":
        # print("No mando nada")
        pass
      elif ficha and archpermi(ficha.filename):
        filename = secure_filename(ficha.filename)
        extc = filename.split('.')
        print(extc)
        print(filename)
        print(alumno.num_c.data)
        filename = "ficha_" + alumno.num_c.data + "." + extc[1]
        print(filename)
        ficha.save(os.path.join(app.config["folder"], filename))
        flash("", "")  # Ya guarda el archivo
      else:
        print("archivo no permitido")
      mycursor = mydb.cursor()
      print(cursos)
      global cur
      global enf
      global sec
      global ci
      global tel
      cur = request.form.get(('curso'))
      enf = request.form.get(('enfasis'))
      sec = request.form.get(('seccion'))
      ci = alumno.num_c.data
      tel = alumno.num_t.data
      # Validar si la cedula y el num de telefono no son repetidos..
      sql = "SELECT ci_a FROM alumnos WHERE ci_a = %s"
      val = [ci]
      mycursor.execute(sql, val)
      comp_a = mycursor.fetchall()
      if comp_a:
        print(comp_a)
        flash("", "error")
      else:
        # Inserta en tabla alumnnos todos los datos
        mycursor.execute(
          'INSERT INTO alumnos (nmb_a, ape_a, tel_a, ci_a, edad, email, loc_a, nmb_tu, tel_tu, fec_a, bar_a) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)',
          (alumno.nombre.data, alumno.apellido.data, alumno.num_t.data, alumno.num_c.data, alumno.edad.data,
           alumno.email.data, alumno.localidad.data, alumno.nmb_pa.data, alumno.num_pa.data, fecha, alumno.barrio.data))
        mydb.commit()
        return redirect(url_for('sacarmat'))
    else:
      print("no fehca")
      flash("fecha","fecha")
  return render_template('inscribiral.html', datos = datos, alumno = alumno)

@app.route('/inscribiral/<filename>') #Parte de buscar el archivo o documento del alumno, falta testear mejor
def getfile(filename):
  return send_file(os.path.join(app.config["folder"], filename))

@app.route('/inscribiralum', methods = ['GET', 'POST']) #Parte de inscripcion, carga de materias por alumno
def inscurmat():
  bver = 0
  datos = session['username']
  global idmaterias #Contiene el nombre de las materias y sus id
  global ci
  global tel
  fecha = datetime.now()
  fecha = datetime.strftime(fecha, '%Y')
  print(fecha)
  print("a")
  if request.method == 'POST':
    print("a")
    if ci == 0 or tel == 0:
      print("a2")
      return redirect(url_for('inscribirdir'))
    idmat = request.form.getlist(('idmat')) #Aca si tengo los id de las materias en q se insc
    mycursor = mydb.cursor()
    sql = "SELECT id_alumno FROM alumnos WHERE ci_a = %s and tel_a = %s"
    val = [ci,tel]
    mycursor.execute(sql, val)
    id_a = mycursor.fetchall()
    id_a = id_a[0]
    print(cur)
    print("q pasa")
    if id_a:
      print(id_a[0])
    if idmaterias:
      print(idmat)
      bver = 1
      for x in range(0, len(idmat)):
        aux = idmat[x]
        #print(aux)
        #print(x)
        sql = "SELECT id_profesor FROM matxpro WHERE id_materia = %s and id_curso = %s"
        val = [aux, cur]
        mycursor.execute(sql, val)
        data = mycursor.fetchall()
        if data:
          data = data[0]
          print(id_a[0])
          print(data[0])
          mycursor = mydb.cursor()
          mycursor.execute(
             'INSERT INTO matxalum (id_alumno, id_materia, id_curso, ano_m, id_profesor) VALUES (%s, %s, %s, %s, %s)',
            (id_a[0], aux, cur, fecha, data[0]))
          mydb.commit()
        else:
          mycursor = mydb.cursor()
          mycursor.execute(
            'INSERT INTO matxalum (id_alumno, id_materia, id_curso, ano_m, id_profesor) VALUES (%s, %s, %s, %s, %s)',
            (id_a[0], aux, cur, fecha, 0))
          mydb.commit()
      print("termino")
      #time.sleep(3.5)
      flash("", "si")
      global band
      band = 1
      print(band)
      print("a")
      return redirect(url_for('inscribirdir'))
    else:
      print("a")
      return redirect(url_for('inscribirdir'))
  return render_template('inscribiral2.html', datos = datos, cursos = cursos, materias=idmaterias, bver = bver)

@app.route('/sacarmat', methods = ['GET', 'POST']) #Va antes que inscurmat #Parte de Inscripcion de Alumnos, saca las materias con nombre y ID.
def sacarmat():
  global cur
  global enf
  global sec
  global ci
  global tel
  global tipoins
  if tipoins == 1: # Osea Alumnos
    if enf == "Sociales":  # Dependiendo del enfasis para no consultar la bd
      enf = 2
    else:
      enf = 1
    print(cur)
    print(enf)
    mycursor = mydb.cursor()
    sql = "SELECT id_curso FROM cursos WHERE des_c = %s and sec_c = %s and id_enfasis = %s"
    val = [cur, sec, enf]
    mycursor.execute(sql, val)
    data = mycursor.fetchall()
    if data:
      print(data[0])
      data = data[0]
    #Test
    #print(ci)
    #print(tel)
    # Updatea con el curso asignado
    sql = "UPDATE alumnos SET id_curso = %s WHERE ci_a = %s and tel_a = %s"
    val = (data[0],ci, tel)
    mycursor.execute(sql, val)
    mydb.commit()
    cur = data[0]
    #Sacar los nombres de materia y sus id
    sql = "SELECT matxcur.id_materia, des_m FROM matxcur, materias WHERE matxcur.id_curso = %s and matxcur.id_enfasis = %s " \
          "and matxcur.id_materia = materias.id_materia "
    val = [data[0], enf]
    mycursor.execute(sql, val)
    mat = mycursor.fetchall()
    print(mat)
    bver = 1
    global idmaterias
    idmaterias = mat
    return redirect(url_for('inscurmat'))


@app.route('/inscribirprofe', methods = ['GET', 'POST']) #Inscripcion de Profesor, carga.
def inscribirprof():
  alumno = userform.Alumno(request.form) #Reutilizo el form
  datos = session['username']
  global band
  global idmaterias
  print(band)
  if band == 1:
    flash("", "si")
    band = 0
  if band == 3:
    band  = 0
    flash("", "no")
  if band == 4:
    band = 0
    flash("", "profe")
  global tipoins
  tipoins = 2
  if request.method == "POST":
    co_i = request.files["co_i"]
    cedu = request.files["cedu_p"]
    co_c = request.files["co_c"]
    fecha = (request.form['fec_n'])
    print(fecha)
    if fecha:
      # Parte Cedula...
      if "cedu_p" not in request.files:
        # print("No envio nada")
        pass
      elif cedu.filename == "":
        # print("No mando nada")
        pass
      elif cedu and archpermi(cedu.filename):
        filename = secure_filename(cedu.filename)
        extc = filename.split('.')
        print(extc)
        print(filename)
        print(alumno.num_c.data)
        filename = "cedulaprof_" + alumno.num_c.data + "." + extc[1]
        print(filename)
        cedu.save(os.path.join(app.config["folder"], filename))
        flash("", "")  # Ya guarda el archivo
      else:
        print("archivo no permitido")
      # Parte Constancia de Ingresos
      if "co_i" not in request.files:
        #print("No envio nada")
        pass
      elif co_i.filename == "":
        #print("No mando nada")
        pass
      elif co_i and archpermi(co_i.filename):
        filename = secure_filename(co_i.filename)
        extc = filename.split('.')
        print(extc)
        print(filename)
        print(alumno.num_c.data)
        filename = "constancia_ingresos_" + alumno.num_c.data + "." + extc[1]
        print(filename)
        co_i.save(os.path.join(app.config["folder"], filename))
        flash("", "")  # Ya guarda el archivo
      else:
        print("archivo no permitido")
      # Parte Autorizacion de Padres
      if "co_c" not in request.files:
        #print("No envio nada")
        pass
      elif co_c.filename == "":
        #print("No mando nada")
        pass
      elif co_c and archpermi(co_c.filename):
        filename = secure_filename(co_c.filename)
        extc = filename.split('.')
        print(extc)
        print(filename)
        print(alumno.num_c.data)
        filename = "constancia_cargos_" + alumno.num_c.data + "." + extc[1]
        print(filename)
        co_c.save(os.path.join(app.config["folder"], filename))
        flash("", "")  # Ya guarda el archivo
      else:
        print("archivo no permitido")
      mycursor = mydb.cursor()
      print(cursos)
      global cur
      global enf
      global sec
      global ci
      global tel
      cur = request.form.get(('curso'))
      enf = request.form.get(('enfasis'))
      sec = request.form.get(('seccion'))
      ci = alumno.num_c.data
      tel = alumno.num_t.data
      # Validar si la cedula y el num de telefono no son repetidos..
      sql = "SELECT ci_p FROM profesores WHERE ci_p = %s"
      val = [ci]
      mycursor.execute(sql, val)
      comp_p = mycursor.fetchall()
      if comp_p:
        print(comp_p)
        flash("", "error")
      else:
        # Inserta en tabla profesores todos los datos
        mycursor.execute(
          'INSERT INTO profesores (nmb_p, ape_p, tel_p, ci_p, edad, email, loc_p, fec_p, bar_p) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)',
          (alumno.nombre.data, alumno.apellido.data, alumno.num_t.data, alumno.num_c.data, alumno.edad.data,
           alumno.email.data, alumno.localidad.data, fecha, alumno.barrio.data))
        mydb.commit()
        return redirect(url_for('inscmatxprof'))
    else:
      print("no fehca")
      flash("fecha", "fecha")
  return render_template('inscribirprof.html', datos=datos, alumno=alumno, profe = idmaterias)

@app.route('/inscribirprofee', methods = ['GET', 'POST']) #Parte de incorporacion, carga de materias por  profe
def inscmatxprof():
  bver = 0
  datos = session['username']
  global idmaterias #Contiene el nombre de las materias y sus id
  global ci
  global tel
  global tipoins
  global band
  tipoins = 2 #Profesores
  fecha = datetime.now()
  fecha = datetime.strftime(fecha, '%Y')
  print(fecha)
  mycursor = mydb.cursor(buffered=True)
  for x in range(1, 7):
    print(x)
    if x <= 3:
      sql = "SELECT matxcur.id_curso, materias.id_materia, des_m FROM matxcur, materias WHERE id_curso = %s and id_enfasis = %s and " \
            "materias.id_materia = matxcur.id_materia"
      val = [x, 1]
      mycursor.execute(sql, val)
      if x == 1:
        con_p = mycursor.fetchall()
        print(con_p)
      if x == 2:
        con_s = mycursor.fetchall()
        print(con_s)
    if x == 3:
      sql = "SELECT matxcur.id_curso, materias.id_materia, des_m FROM matxcur, materias WHERE id_curso = %s and id_enfasis = %s and " \
            "materias.id_materia = matxcur.id_materia"
      val = [x, 2]
      mycursor.execute(sql, val)
      soc_p = mycursor.fetchall()
      print(soc_p)
    if x == 4:
      sql = "SELECT matxcur.id_curso, materias.id_materia, des_m FROM matxcur, materias WHERE id_curso = %s and id_enfasis = %s and " \
            "materias.id_materia = matxcur.id_materia"
      val = [x, 1]
      mycursor.execute(sql, val)
      con_t = mycursor.fetchall()
      print(con_t)
    if x > 4:
      sql = "SELECT matxcur.id_curso, materias.id_materia, des_m FROM matxcur, materias WHERE id_curso = %s and id_enfasis = %s and " \
            "materias.id_materia = matxcur.id_materia"
      val = [x, 2]
      mycursor.execute(sql, val)
      if x == 5:
        soc_s = mycursor.fetchall()
        print(soc_s)
      if x == 6:
        soc_t = mycursor.fetchall()
        print(soc_t)
  if request.method == 'POST':
    if ci == 0 or tel == 0:
      band = 3
      return redirect(url_for('inscribirprof'))
    idmat = request.form.getlist(('idmat')) #Aca si tengo los id de las materias en q se insc
    idcur = request.form.getlist(('idcur'))
    print(idmat)
    print(idcur)
    mycursor = mydb.cursor()
    sql = "SELECT id_profesor FROM profesores WHERE ci_p = %s and tel_p = %s"
    val = [ci,tel]
    mycursor.execute(sql, val)
    id_p = mycursor.fetchall()
    #print(cur)
    if id_p:
      id_p = id_p[0]
      print(id_p[0])
      if idmat:
        print(idmat)
        for x in range(0, len(idmat)):
          aux = idmat[x]
          aux2 = aux.split(',')
          aux3 = int(aux2[0])
          aux2 = int(aux2[1])
          print((aux2))
          print((aux3))
          mycursor = mydb.cursor()
          sql = "SELECT * FROM matxpro WHERE id_materia = %s and id_curso = %s"
          val = [aux2, aux3]
          mycursor.execute(sql, val)
          comp_m = mycursor.fetchall()
          if comp_m:
            comp_m = comp_m[0]
            print(comp_m)
            if comp_m[2] != 0:
              band = 4
              print("ya tiene profesor")
            if comp_m[2] == 0:
              sql = "UPDATE matxpro SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
              val = (id_p, aux2, aux3)
              mycursor.execute(sql, val)
              mydb.commit()
              # Asigna el profesor a los alumnos
              sql = "UPDATE matxalum SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
              val = (id_p, aux2, aux3)
              mycursor.execute(sql, val)
              mydb.commit()
              # Asigna el profesor al horario
              sql = "UPDATE horarios SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
              val = (id_p, aux2, aux3)
              mycursor.execute(sql, val)
              mydb.commit()
              # Comprueba si cargo un trabajo
              sql = "SELECT id_trabajo, id_materia FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s"
              val = [0, aux2, aux3]
              mycursor.execute(sql, val)
              trabajos = mycursor.fetchall()
              if trabajos:
                for x in range(0, len(trabajos)):
                  aux = trabajos[x]
                  id_t = aux[0]
                  sql = "UPDATE trabajos SET id_profesor = %s WHERE id_trabajo = %s"
                  val = (id, id_t)
                  mycursor.execute(sql, val)
                  mydb.commit()
              print("el profesor fue dado de baja")
              band = 1
          else:
            mycursor.execute(
              'INSERT INTO matxpro (id_materia, id_profesor, id_curso, fecha) VALUES (%s, %s, %s, %s)',
              (aux2, id_p, aux3, fecha))
            mydb.commit()
            band = 1
            print(band)
            print("se le asigna")
            # Asigna el profesor a los alumnos
            sql = "UPDATE matxalum SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
            val = (id_p, aux2, aux3)
            mycursor.execute(sql, val)
            mydb.commit()
            # Asigna el profesor al horario
            sql = "UPDATE horarios SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
            val = (id_p, aux2, aux3)
            mycursor.execute(sql, val)
            mydb.commit()
        return redirect(url_for('inscribirprof'))
      else:
        band = 3
        return redirect(url_for('inscribirprof'))
    else:
      band == 3
      return redirect(url_for('inscribirprof'))
  return render_template('inscribirprof2.html', datos = datos, cursos = cursos, materias=idmaterias, bver = bver,
                         soc_p = soc_p, soc_s = soc_s, soc_t = soc_t, con_p=con_p, con_s = con_s, con_t = con_t)


#@app.route('/cargar', methods = ['GET', 'POST']) #ABM Cargar Materias, desabilitado.
def cargar():
  alumno = userform.Alumno(request.form)
  if request.method == 'POST':
    cur = request.form.getlist(('idalum'))
    enf = request.form.get(('enfasis'))
    sec = request.form.get(('seccion'))
    ch1 = request.form.get(('car_1'))
    ch2 = request.form.get(('car_2'))
    ch3 = request.form.get(('car_3'))
    #print(alumno.nombre.data)
    #print(cur)
    #print(ch1)
    nmb = (alumno.nombre.data ,)
    sql = "SELECT id_materia FROM materias WHERE des_m = %s"
    mycursor = mydb.cursor()
    val = [alumno.nombre.data]
    mycursor.execute(sql, val)
    id_m = mycursor.fetchall()
    #print(enf)
    #print(sec)
    mycursor = mydb.cursor()
    if enf == "Sociales": #Dependiendo del enfasis para no consultar la bd
      enf = 2
    else:
      enf = 1
    #print(enf)
    if id_m:
      pass
    else:
      mycursor.execute('INSERT INTO materias (des_m) VALUES (%s)', (nmb))
      mydb.commit()
    cursos = []
    for x in range(0, len(cur)): #Sacar los cursos seleccionados
      #print(x)
      aux = cur[x]
      mycursor = mydb.cursor()
      sql = "SELECT id_curso, id_enfasis FROM cursos WHERE des_c = %s and sec_c = %s and id_enfasis = %s"
      val = [aux, sec, enf]
      mycursor.execute(sql, val)
      aux = mycursor.fetchall()
      #print(aux)
      aux = aux[0]
      #print(aux)
      cursos.append(aux)
    print(cursos)
    sql = "SELECT id_materia FROM materias WHERE des_m = %s"
    val = [alumno.nombre.data]
    mycursor.execute(sql, val)
    id_m = mycursor.fetchall()
    id_m = id_m[0]
    print(id_m)
    for x in range(0, len(cursos)):  # Sacar los cursos seleccionados
      # print(x)
      aux = cursos[x]
      print(aux)
      print(cur[x])
      if cur[x] == "Primer Curso":
        ch = ch1
      if cur[x] == "Segundo Curso":
        ch = ch2
      if cur[x] == "Tercer Curso":
        ch = ch3
      mycursor = mydb.cursor()
      mycursor.execute(
        'INSERT INTO matxcur (id_curso, id_materia, id_enfasis, car_h) VALUES (%s, %s, %s ,%s)',
        (aux[0], id_m[0], aux[1], ch))
      mydb.commit()
  return render_template('cargarm.html', alumno = alumno)
@app.route('/misdatos') #Nav Bar, Boton "Mis Datos"
def misdatos():
  datos = session['username']
  print(datos)
  if datos[7] == 1: #alumnos
    return render_template('verdatos.html', datos = datos)
  if datos[6] == 2: #profesores
    return render_template('verdatosprof.html', datos=datos)
  if datos[6] ==3: #Admin
    return render_template('verdatosad.html', datos=datos)

@app.route('/miconfig', methods = ['GET', 'POST']) #Nav Bar, Boton "Configuracion de Cuenta"
def misconfig():
  global band
  datos = session['username']
  user = userform.User(request.form)
  if band == 1:
    band = 0
    flash("","c_u")
  if band == 2:
    band = 0
    flash("","co_u")
  if band == 3:
    band = 0
    flash("", "p_u")
  mycursor = mydb.cursor()
  if datos[7] == 1:  # alumnos
    sql = "SELECT * FROM user WHERE id_user = %s"
    val = [datos[6]]
    mycursor.execute(sql, val)
    data = mycursor.fetchall()
    print(data[0])
  else:
    sql = "SELECT * FROM user WHERE id_user = %s"
    val = [datos[5]]
    mycursor.execute(sql, val)
    data = mycursor.fetchall()
    print(data[0])

  if request.method == 'POST':
    us = user.username.data
    pasw = user.password.data
    em = user.email.data
    cpasw = user.confirmpassword.data
    cpasw2 = user.confirmpassword2.data
    if us and pasw:
      sql = "SELECT id_user FROM user WHERE username = %s"
      val = [us]
      mycursor.execute(sql, val)
      comp_u = mycursor.fetchall()
      if comp_u:
        flash("", "us_e")
      else:
        if data:
          userdata = data[0]
          # print(userdata)
          if userdata:
            passcheck = userdata[3]
            # print(passcheck)
            print(us)
          if userdata and check_password_hash(passcheck, pasw):
            sql = "UPDATE user SET username =  %s WHERE id_user = %s"
            val = (us, datos[6])
            mycursor.execute(sql, val)
            mydb.commit()
            band = 1
            return redirect(url_for('misconfig'))
          else:
            flash("", "psw")
    if em and pasw:
      if data:
        userdata = data[0]
        # print(userdata)
        if userdata:
          passcheck = userdata[3]
        if userdata and check_password_hash(passcheck, pasw):
          sql = "UPDATE user SET email = %s WHERE id_user = %s"
          val = (em, datos[6])
          mycursor.execute(sql, val)
          mydb.commit()
          band = 2
          return redirect(url_for('misconfig'))
        else:
          flash("", "psw")
    if pasw and cpasw and cpasw2:
      if data:
        userdata = data[0]
        if userdata:
          passcheck = userdata[3]
        if userdata and check_password_hash(passcheck, pasw):
          if cpasw == cpasw2:
            password = createpassword(cpasw)
            sql = "UPDATE user SET password = %s WHERE id_user = %s"
            val = (password, datos[6])
            mycursor.execute(sql, val)
            mydb.commit()
            band = 3
            return redirect(url_for('misconfig'))
          else:
            flash("", "psw2")
        else:
          flash("", "psw")
  if datos[7] == 1:  # alumnos
    return render_template('miconfig.html', datos=datos, data=data[0], user=user)
  if datos[6] == 2:  # profesores
    return render_template('miconfigpro.html', datos=datos, data=data[0], user=user)
  if datos[6] == 3:  # Admin
    return render_template('miconfigad.html', datos=datos, data=data[0], user=user)


@app.route('/cerrarsesion') #Nav Bar, Boton "Mis Datos"
def cerrarsesion():
  if 'username' in session:
    session.pop('username')
  return redirect(url_for('login'))
@app.route('/modificarproceso/<int:id>', methods = ['GET', 'POST']) #Modificar trabajos del docente
def modproceso(id):
  global band
  #Mensaje flash
  if band == 1:
    band = 0
    flash("", "si_c")
  if band == 3:
    band = 0
    flash("", "eliminado")
  datos = session['username']
  mycursor = mydb.cursor()
  print(datos)
  #Obtiene los trabajos relacionados a la materia
  sql = "SELECT id_trabajo, fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and trabajos.id_materia = %s and " \
        "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  #print(trabajos)
  if request.method == 'POST':
    filtro = int(request.form.get(('idfiltro')))
    #print(filtro)
    #Dependiendo del filtro, busca por etapa
    if filtro == 1:
      sql = "SELECT id_trabajo, fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and etapa = %s and trabajos.id_materia = %s and " \
            "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
      val = [datos[0], filtro, id]
      mycursor.execute(sql, val)
      trabajos = mycursor.fetchall()
    if filtro == 2:
      sql = "SELECT id_trabajo, fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and etapa = %s and trabajos.id_materia = %s and " \
            "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
      val = [datos[0], filtro, id]
      mycursor.execute(sql, val)
      trabajos = mycursor.fetchall()
    if filtro == 3:
      return redirect(url_for('modproceso', id = id))
    #print(trabajos)
    return render_template('modproceso.html', datos=datos, trabajos=trabajos, filtro = filtro, id= id)
  return render_template('modproceso.html', datos=datos, trabajos = trabajos, filtro = 0, id = id)

@app.route('/exportarexcelproceso/<int:id>', methods = ['GET', 'POST'])
def generarexcel(id):
  #print(id)
  datos = session['username']
  celdas = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'AA', 'AB', 'AC', 'AE', 'AF'
          , 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC',
            'BD', 'BE']
  if request.method == 'POST':
    mycursor = mydb.cursor()
    filtro = int(request.form.get(('idfiltro')))
    print(filtro)
    #Trabajos de la materia
    sql = "SELECT id_trabajo, fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and etapa = %s and trabajos.id_materia = %s and " \
          "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
    val = [datos[0], filtro, id]
    mycursor.execute(sql, val)
    trabajos = mycursor.fetchall()
    a = trabajos[0]
    if not a:
      return redirect(url_for('modproceso', id=id))
    print(trabajos)
    inf = datetime.now()
    # Extraemos la fecha
    fecha = datetime.strftime(inf, '%Y/%m/%d')
    # Cargamos el archivo
    wb = load_workbook('templateprocesos.xlsx')
    ws = wb["Hoja1"]
    ws['C4'] = datos[1] + ", " + datos[2]
    ws['V4'] = a[5]
    ws['AP4'] = a[8]
    ws['C5'] = filtro
    if a[7] == "Primer Curso":
      ws['G5'] = 1
    if a[7] == "Segundo Curso":
      ws['G5'] = 2
    if a[7] == "Tercer Curso":
      ws['G5'] = 3
    ws['T5'] = "T.N"
    # Insertamos los indicadores
    cont = 0
    # Para saber que celdas combinar
    ci = ""
    cic = 0
    cf = ""
    cfc = 0
    #Combinar celdas
    sheet = wb.active

    for i in range(0, len(trabajos)):
      inicio = 12
      aux = trabajos[i]
      # Saca los indicadores
      sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
      val = [aux[0]]
      mycursor.execute(sql, val)
      indicadores = mycursor.fetchall()
      if i == 0:
        # Saca los alumnos relacionados con el trabajo
        sql = "SELECT nmb_a, ape_a, id_txa, traxalum.id_alumno, pun_l, fec_t, id_trabajo FROM traxalum, alumnos WHERE id_trabajo = %s and traxalum.id_alumno = alumnos.id_alumno"
        val = [aux[0]]
        mycursor.execute(sql, val)
        datas = mycursor.fetchall()
        datas = sorted(datas, key=lambda datas: datas[1])  # Ordena por orden alfabetic
      for x in range(0, len(indicadores)):
        aux2 = indicadores[x]
        if x == 0:
          ci = celdas[cont]
          cic = cont
        ws['{a}11'.format(a=celdas[cont])] = aux2[1]
        ws['{a}41'.format(a=celdas[cont])] = aux2[1]
        #Puntaje de cada indicador
        ws['{a}12'.format(a=celdas[cont])] = aux2[2]
        ws['{a}42'.format(a=celdas[cont])] = aux2[2]
        if x == len(indicadores) - 1:
          cf = celdas[cont + 1]
          cfc = cont + 1
          ws['{a}11'.format(a=celdas[cont + 1])] = "TOTAL INDICADORES LOGRADOS"
          ws['{a}41'.format(a=celdas[cont + 1])] = "TOTAL INDICADORES LOGRADOS"
          color1 = PatternFill(start_color='cc9cfc',
                                end_color='cc9cfc',
                                fill_type='solid')
          ws['{a}11'.format(a=celdas[cont + 1])].fill = color1
          ws['{a}41'.format(a=celdas[cont + 1])].fill = color1
          cont += 1
        cont += 1
      print(ci, cic, cf, cfc)
      #Fecha del trabajo
      ws['{a}6'.format(a=ci)] = aux[1]
      #Tema o contenido
      ws['{a}9'.format(a=ci)] = aux[2]
      #Total de puntos
      ws['{a}12'.format(a=cf)] = aux[3]
      ws['{a}42'.format(a=cf)] = aux[3]
      a = ws['{a}42'.format(a=cf)]
      b = ws['{a}12'.format(a=cf)]
      a.font = Font(bold=True)
      b.font = Font(bold=True)
      #Une las celdas
      for y in range(6,11):
        sheet.merge_cells('{a}{c}:{b}{c}'.format(a=ci, b= cf, c=y))
      total = 0
      for i in range(0, len(datas)):
        alumno = datas[i]
        # Saca los indicadores del alumno
        sql = "SELECT id_ixa, pun_l FROM indxalum WHERE id_trabajo = %s and id_alumno = %s"
        val = [aux[0], alumno[3]]
        mycursor.execute(sql, val)
        indicadores = mycursor.fetchall()
        print(indicadores)
        fc = 0
        for x in range(cic, cfc + 1):
          #print(x)
          indicador = indicadores[fc]
          pl = indicador[1]
          print(indicador)
          if not x == cfc - 1:
            fc += 1
            #print(fc)
          if inicio != 32 and x != cfc:
            ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = pl
            #print('{b}{a}'.format(a=str(inicio + 1), b=celdas[x]))
            total += pl
          if inicio != 32 and x == cfc:
            ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = total
            pn = ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])]
            pn.font = Font(bold=True)
          #Salto a laa planilla de abajo
          if inicio == 32 and x != cfc:
            inicio = 42
            ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = pl
            inicio += 1
          if inicio == 32 and x == cfc:
            ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = total
            pn = ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])]
            pn.font = Font(bold=True)

        til = ws['BP{a}'.format(a=str(inicio + 1))]
        if not til.value:
          ws['BP{a}'.format(a=str(inicio + 1))] = 0
          til = ws['BP{a}'.format(a=str(inicio + 1))]
        print(til.value)
        ws['BP{a}'.format(a=str(inicio + 1))] = til.value + total
        print("endfor")
        total = 0
        inicio += 1
    # Insertamos alumnos
    inicio = 12
    for x in range(0, len(datas)):
      alumno = datas[x]
      print(alumno)
      if inicio != 32:
        ws['B{a}'.format(a=str(inicio + 1))] = alumno[1] + ", " + alumno[0]
        inicio += 1
      else:
        inicio = 42
        ws['B{a}'.format(a=str(inicio + 1))] = alumnos[x]
        inicio += 1
    # Guarda el archivo
    ws = wb["Hoja1"]
    ws.title ="Planilla de Proceso"
    wb.save('./media/Planilla de Proceso.xlsx')
    ruta = pathlib.Path('./media/')
    filename = "Planilla de Proceso.xlsx"
    archivo = ruta / filename  # Si existe un docx con su nombre
    print(archivo)
    if archivo.exists():
      print("El arhivo existe")
      return send_file(archivo, as_attachment=True)
    else:
      print("No existe")
  return redirect(url_for('modproceso', id = id))

@app.route('/modificarproceso2/<int:id>', methods = ['GET', 'POST'])
def modificarproceso2(id):
  print(id)
  datos = session['username']
  #Alumnos
  global user_datos
  mycursor = mydb.cursor()
  #Saca los alumnos relacionados con el trabajo
  sql = "SELECT nmb_a, ape_a, id_txa, traxalum.id_alumno, pun_l, fec_t, id_trabajo FROM traxalum, alumnos WHERE id_trabajo = %s and traxalum.id_alumno = alumnos.id_alumno"
  val = [id]
  mycursor.execute(sql, val)
  datas = mycursor.fetchall()
  datas = sorted(datas, key=lambda datas: datas[1]) #Ordena por orden alfabetico
  print(datas)
  #Saca los datos relacionados con el trabajo
  sql = "SELECT des_t, pun_t, tipo_t, id_materia, id_trabajo FROM trabajos WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  print(datas)
  punt_t = trabajos[0]
  print(punt_t)
  alumnos = []
  #Mensajes
  global band
  if band == 1:
    band = 0
    flash("si_c", "si_c")
  if band == 2:
    band = 0
    flash("ocurrio", "ocurrio")
  if request.method == 'POST':
    #Saca los ids de los alumnos
    idalum = request.form.getlist(('idalum'))
    print(idalum)
    #Saca los alumnos seleccionados por el docente
    for x in range(0, len(idalum)):
      aux = int(idalum[x])
      print(aux)
      mycursor = mydb.cursor()
      sql = "SELECT id_alumno,nmb_a, ape_a FROM alumnos WHERE id_alumno = %s"
      val = [aux]
      mycursor.execute(sql, val)
      aux2 = mycursor.fetchall()
      aux2 = aux2[0]
      #print(aux2)
      alumnos.append(aux2)
    #print(alumnos)
    #Guardamos en una global para la funcion de cargar puntaje
    user_datos = alumnos
    return redirect(url_for('modificarpuntaje', id = id))
  return render_template('modproceso2.html', datos=datos, datas = datas, trabajos = trabajos[0])

@app.route('/modificarpuntaje/<int:id>', methods = ['GET', 'POST']) #Modificar puntaje obtenido por el alumno
def modificarpuntaje(id):
  datos = session['username']
  #Alumnos
  global user_datos
  # Saca el trabajo
  mycursor = mydb.cursor()
  sql = "SELECT id_trabajo, des_t, pun_t, id_materia, id_curso, id_profesor, etapa FROM trabajos WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  trabajos = trabajos[0]
  print(trabajos)
  # Saca los indicadores
  sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  indicadores = mycursor.fetchall()
  print(indicadores)
  #Saca los alumnos
  global user_datos
  alumnos = user_datos
  print(alumnos)
  #Comprobar que este bien cargado
  create = True
  #Para mandar mensajes
  global band
  # Comprueba que la variable global este funcionando correctamente
  if not alumnos:
    band = 2
    return redirect(url_for('modificarproceso2', id=id))
  if request.method == 'POST':
    puntajes = []
    pl = 0
    #Comprobar que se hayan cargado todos los indicadores
    for x in range(0, len(alumnos)):
      for y in range(1, len(indicadores) + 1):
        aux = request.form.getlist(('{a}{i}'.format(a=x + 1, i=y)))
        if not aux:
          create = False
    #Solo si esta cargado correctamente
    if create == True:
      #Carga de indicador por alumno, actualizancion de puntaje logrado por indicador y puntaje total logrado
      for x in range(0, len(alumnos)):
        aux1 = alumnos[x]
        # print(aux1)
        for y in range(1, len(indicadores) + 1):
          aux2 = indicadores[y - 1]
          print(aux2)
          if y == 1:
            pl = 0
            # print("Alumno {}".format(alumnos[x]))
          aux = request.form.getlist(('{a}{i}'.format(a=x + 1, i=y)))
          # print('{a}{i}'.format(a=x + 1, i=y))
          print(aux)
          p = int(aux[0])
          # print(p)
          pl += int(aux[0])
          sql = "UPDATE indxalum SET pun_l =  %s WHERE id_alumno = %s and id_trabajo = %s and id_indicador = %s"
          val = (p,aux1[0],id, aux2[0])
          mycursor.execute(sql, val)
          mydb.commit()
          if y == len(indicadores):
            puntajes.append(pl)
      print(puntajes)
      #Actualizando el puntaje total logrado por el alumno
      for x in range(0, len(alumnos)):
        aux1 = alumnos[x]
        aux2 = puntajes[x]
        #Sacando primer puntaje obtenido
        sql = "SELECT id_alumno, pun_l FROM traxalum WHERE id_alumno = %s and id_trabajo = %s"
        val = [aux1[0], id]
        mycursor.execute(sql, val)
        pp = mycursor.fetchall()
        pp = pp[0]
        print(pp)
        # Actualizando el trabajo al alumno con su puntaje logrado
        sql = "UPDATE traxalum SET pun_l =  %s WHERE id_alumno = %s and id_trabajo = %s"
        val = (aux2, aux1[0], id)
        mycursor.execute(sql, val)
        #Actualizando su puntaje acumulado
        ac = aux2 - int(pp[1])
        print(ac)
        if trabajos[6] == 1:  # Primera Etapa..
          if ac > 0:
            print('mas')
            sql = "UPDATE matxalum SET pun_ac = (pun_ac + %s) WHERE id_alumno = %s and id_materia = %s and id_profesor = %s"  # Saco mas puntaje
            val = (ac, aux1[0], trabajos[3], trabajos[5])
            mycursor.execute(sql, val)
            mydb.commit()
          else:
            print('menos')
            ac *= -1 #sino no  resta bien el acumulado
            sql = "UPDATE matxalum SET pun_ac = (pun_ac - %s) WHERE id_alumno = %s and id_materia = %s and id_profesor = %s"  # Saco mas puntaje
            val = (ac, aux1[0], trabajos[3], trabajos[5])
            mycursor.execute(sql, val)
            mydb.commit()
        if trabajos[6] == 2: #Segunda
          if ac > 0:
            print('mas')
            sql = "UPDATE matxalum SET pun_ac2 = (pun_ac2 + %s) WHERE id_alumno = %s and id_materia = %s and id_profesor = %s"  # Saco mas puntaje
            val = (ac, aux1[0], trabajos[3], trabajos[5])
            mycursor.execute(sql, val)
            mydb.commit()
          else:
            print('menos')
            ac *= -1 #sino no  resta bien el acumulado
            sql = "UPDATE matxalum SET pun_ac2 = (pun_ac2 - %s) WHERE id_alumno = %s and id_materia = %s and id_profesor = %s"  # Saco mas puntaje
            val = (ac, aux1[0], trabajos[3], trabajos[5])
            mycursor.execute(sql, val)
            mydb.commit()
            band = 1
        return redirect(url_for('modificarproceso2', id=id))
    else:
      flash("error", "error")
  return render_template('modproceso3.html', datos=datos, trabajos =trabajos, indicadores = indicadores, alumnos=user_datos)

@app.route('/eliminartrabajo/<int:id>', methods = ['GET', 'POST']) #Eliminar el trabajo seleccionado
def eliminartrabajo(id):
  print(id)
  datos = session['username']
  # Saca el trabajo
  mycursor = mydb.cursor()
  sql = "SELECT id_trabajo, des_t, pun_t, id_materia, id_curso, id_profesor, etapa FROM trabajos WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  trabajos = trabajos[0]
  print(trabajos)
  # Saca los indicadores
  sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  indicadores = mycursor.fetchall()
  print(indicadores)
  # Saca los alumnos relacionados con el trabajo
  sql = "SELECT nmb_a, ape_a, id_txa, traxalum.id_alumno, pun_l, fec_t, id_trabajo FROM traxalum, alumnos WHERE id_trabajo = %s and traxalum.id_alumno = alumnos.id_alumno"
  val = [id]
  mycursor.execute(sql, val)
  datas = mycursor.fetchall()
  datas = sorted(datas, key=lambda datas: datas[1])  # Ordena por orden alfabetico
  print(datas)
  if request.method == 'POST':
    #Borra los indicadores por alumno
    mycursor = mydb.cursor()
    sql = "DELETE FROM indxalum WHERE id_trabajo = %s"
    val = [id]
    mycursor.execute(sql, val)
    mydb.commit()
    # Borra el trabajo del sistema
    mycursor = mydb.cursor()
    sql = "DELETE FROM trabajos WHERE id_trabajo = %s"
    val = [id]
    mycursor.execute(sql, val)
    mydb.commit()
    # Borra el trabajo al alumno
    mycursor = mydb.cursor()
    sql = "DELETE FROM traxalum WHERE id_trabajo = %s"
    val = [id]
    mycursor.execute(sql, val)
    mydb.commit()
    # Updatea al acumulado del alumno en la materia
    for x in range(0, len(datas)):
      print(x)
      aux = datas[x]
      print(aux)
      aux2 = aux[4]
      print(aux2)
      aux3 = aux[3]
      print(aux3)
      if trabajos[6] == 1: #Primera Etapa:
        sql = "UPDATE matxalum SET pun_ac = (pun_ac - %s) WHERE id_materia = %s and id_profesor = %s and id_curso = %s and id_alumno = %s"
        val = (aux2, trabajos[3], trabajos[5], trabajos[4], aux3)
        mycursor.execute(sql, val)
        mydb.commit()
      if trabajos[6] == 2:# Segunda Etapa:
        sql = "UPDATE matxalum SET pun_ac2 = (pun_ac2 - %s) WHERE id_materia = %s and id_profesor = %s and id_curso = %s and id_alumno = %s"
        val = (aux2, trabajos[3], trabajos[5], trabajos[4])
        mycursor.execute(sql, val)
        mydb.commit()
    #Mensaje flash y redirigir
    global band
    band = 3
    return redirect(url_for('modproceso'))
  return render_template('modproceso4.html', datos=datos, trabajos=trabajos, indicadores=indicadores,
                         datas=datas, id = id)
@app.route('/miscursos') #Listar las materias que tiene por curso el prof
def miscursos():
  datos = session['username']
  print(datos)
  mycursor = mydb.cursor()
  sql = "SELECT id_mxp, matxpro.id_materia, matxpro.id_profesor, matxpro.id_curso, des_m, des_c, des_e, fecha" \
        " FROM matxpro, materias, cursos, enfasis WHERE id_profesor = %s and matxpro.id_materia = materias.id_materia " \
        "and matxpro.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis "
  val = [datos[0]]
  mycursor.execute(sql, val)
  mat = mycursor.fetchall()
  print(mat)
  global band
  if band == 7:
    flash("","noclases")
    band = 0
  if band == 6:
    flash("","si")
    band = 0
  if band == 5:
    flash("", "ya")
    band = 0
  return render_template('miscursos.html', datos = datos, materias = mat)

@app.route('/veralum/<int:id>', methods = ['GET', 'POST']) #Ver los alumnos matriculados con sus cal y puntajes
def misalumnos(id):
  datos = session['username']
  mycursor = mydb.cursor()
  #Saca los datos del alumno, los nombres y puntajes obtenidos en la materia
  sql = "SELECT id_mxa, matxalum.id_alumno, matxalum.id_materia, cal, matxalum.id_curso, pun_ac, des_m, des_c, des_e , nmb_a, ape_a, pun_ac2, cal2" \
        " FROM matxalum, materias, cursos, enfasis, alumnos WHERE id_profesor = %s and matxalum.id_materia = %s and matxalum.id_materia = materias.id_materia " \
        "and matxalum.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis and matxalum.id_alumno = alumnos.id_alumno"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  mat = mycursor.fetchall()
  #Saco los datos de la materia del vector...
  for x in range(0, 1):
    aux = mat[x]
    data = aux
    #print(data)
  #Ordena alfabetaicamente los alumnos del vector...
  mat = sorted(mat, key=lambda mat: mat[10])
  #print(mat)
  #Saca el total de puntos en la primera etapa
  sql = "SELECT id_trabajo, pun_t FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s and etapa = %s"
  val = [datos[0], id, data[4], 1]
  mycursor.execute(sql, val)
  trabajosp = mycursor.fetchall()
  total_p = 0
  #print(trabajosp)
  for x in range(0, len(trabajosp)):
    aux = trabajosp[x]
    aux = aux[1]
    total_p += aux
  #print(total_p)
  # Saca el total de puntos en la segunda etapa
  sql = "SELECT id_trabajo, pun_t FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s and etapa = %s"
  val = [datos[0], id, data[4], 2]
  mycursor.execute(sql, val)
  trabajoss = mycursor.fetchall()
  total_s = 0
  #print(trabajoss)
  for x in range(0, len(trabajoss)):
    aux = trabajoss[x]
    aux = aux[1]
    total_s += aux
  #print(total_s)
  #Alumnos criticos y buenos en primera etapa
  criticos_p = []
  bien_p = []
  estado = (70 * total_p) / 100  # Calcula el 70% del total
  #print(estado)
  for x in range(0, len(mat)):
    aux = mat[x]
    aux2 = aux[5] #Puntaje primera etapa
    if aux2 >= estado:
      bien_p.append(aux)
    else:
      criticos_p.append(aux)
  #print(criticos_p)
  #print(bien_p)
  # Alumnos criticos y buenos en segunda etapa
  criticos_s = []
  bien_s = []
  estado2 = (70 * total_s) / 100  # Calcula el 70% del total
  #print(estado2)
  for x in range(0, len(mat)):
    aux3 = mat[x]
    aux4 = aux3[11]  # Puntaje primera etapa
    if aux4 >= estado2:
      bien_s.append(aux3)
    else:
      criticos_s.append(aux3)
  #print(criticos_s)
  #print(bien_s)
  if request.method == 'POST':
    filtro = int(request.form.get(('idfiltro')))
    #print(filtro)
    if filtro == 1:
      return render_template('misalumnos.html', datos=datos, materias=mat, data=data, filtro=1, id = id, criticos_p = criticos_p, bien_p = bien_p,
                             criticos_s = criticos_s, bien_s = bien_s, total_s = total_s, total_p = total_p)
    if filtro == 2:
      return render_template('misalumnos.html', datos=datos, materias=mat, data=data, filtro=2, id = id, criticos_p = criticos_p, bien_p = bien_p,
                             criticos_s = criticos_s, bien_s = bien_s, total_s = total_s, total_p = total_p )
  return render_template('misalumnos.html', datos = datos, materias = mat, data = data, filtro = 0, id = id, criticos_p = criticos_p, bien_p = bien_p,
                             criticos_s = criticos_s, bien_s = bien_s, total_s = total_s, total_p = total_p)


@app.route('/exportaralumnos/<int:id>', methods = ['GET', 'POST'])
def exportaralumnos(id):
  datos = session['username']
  #print(datos)
  mycursor = mydb.cursor()
  # Saca los datos del alumno, los nombres y puntajes obtenidos en la materia
  sql = "SELECT id_mxa, matxalum.id_alumno, matxalum.id_materia, cal, matxalum.id_curso, pun_ac, des_m, des_c, des_e , nmb_a, ape_a, pun_ac2, cal2" \
        " FROM matxalum, materias, cursos, enfasis, alumnos WHERE id_profesor = %s and matxalum.id_materia = %s and matxalum.id_materia = materias.id_materia " \
        "and matxalum.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis and matxalum.id_alumno = alumnos.id_alumno"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  mat = mycursor.fetchall()
  print(mat)
  # Saco los datos de la materia del vector...
  for x in range(0, 1):
    aux = mat[x]
    data = aux
    print(data)
  #Ordena alfabetaicamente los alumnos del vector...
  mat = sorted(mat, key=lambda mat: mat[10])
  # Saca el total de puntos en la primera etapa
  sql = "SELECT id_trabajo, pun_t FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s and etapa = %s"
  val = [datos[0], id, data[4], 1]
  mycursor.execute(sql, val)
  trabajosp = mycursor.fetchall()
  total_p = 0
  # print(trabajosp)
  for x in range(0, len(trabajosp)):
    aux = trabajosp[x]
    aux = aux[1]
    total_p += aux
  # print(total_p)
  # Saca el total de puntos en la segunda etapa
  sql = "SELECT id_trabajo, pun_t FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s and etapa = %s"
  val = [datos[0], id, data[4], 2]
  mycursor.execute(sql, val)
  trabajoss = mycursor.fetchall()
  total_s = 0
  # print(trabajoss)
  for x in range(0, len(trabajoss)):
    aux = trabajoss[x]
    aux = aux[1]
    total_s += aux

  if request.method == 'POST':
    filtro = int(request.form.get(('idfiltro')))
    print(filtro)
    #Dando formato al pdf
    pdf = SimpleDocTemplate(
      "{a}_alumnos.pdf".format(a = datos[2]),
      pagesize=A4,
      rightMargin=inch,
      leftMargin=inch,
      topMargin=inch,
      bottomMargin=inch / 2
    )
    Story = []
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
    # Cabecera
    text = ''' <strong><font size=14>Alumnos matriculados en {a}.</font></strong>
    '''.format(a=data[6])
    text2 = '''
        <strong><font size=10>Docente: {a}, {b}</font></strong>
    '''.format(a=datos[1], b= datos[2])
    text3 = '''
        <strong><font size=10>Curso: {a}, {b}</font></strong>
    '''.format(a= data[7], b= data[8])
    # Adjuntamos los titulos declarados mas arriba, osea las cabeceras
    Story.append(Paragraph(text2))
    Story.append(Paragraph(text3))
    Story.append(Paragraph(text, styles['Center']))
    Story.append(Spacer(1, 20))
    # Declaramos los datos de la cabecera de la tabla, los titulos y tambien sus estilos
    if filtro == 1: #Primera etapa
      data = [(
        Paragraph('<strong><font size=6>#</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Nombre y Apellido</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Primera Etapa PT:{a}</font></strong>'.format(a = total_p), styles['Center']),
        Paragraph('<strong><font size=6>Calificación</font></strong>', styles['Center'])
      )]
      # Aqui acomplamos los registros o datos a nuestra tabla data, estos seran los datos mostrados de bajo de los headers
      for x in range(0, len(mat)):
        aux = mat[x]
        count = str(x + 1)
        print(count)
        alumno = aux[9] + "," + aux[10]  # nombre y apellido
        pp = aux[5]  # puntaje primera etapa
        if aux[3]:  # Calificacion primera etapa
          cal = aux[3]
        else:
          cal = "Sin calificación"
        data.append((
          Paragraph('<font size=6>%s</font>' % count, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % alumno, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % pp, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % cal, styles['Normal'])
        ))
    if filtro == 2: #Segunda
      data = [(
        Paragraph('<strong><font size=6>#</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Nombre y Apellido</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Segunda Etapa. PT:{a}</font></strong>'.format(a = total_s), styles['Center']),
        Paragraph('<strong><font size=6>Calificación</font></strong>', styles['Center'])
      )]
      # Aqui acomplamos los registros o datos a nuestra tabla data, estos seran los datos mostrados de bajo de los headers
      for x in range(0, len(mat)):
        aux = mat[x]
        count = str(x + 1)
        print(count)
        alumno = aux[9] + "," + aux[10]  # nombre y apellido
        ps = aux[11]  # puntaje segunda etapa
        if aux[12]:
          cal2 = aux[12]
        else:
          cal2 = "Sin calificación"
        data.append((
          Paragraph('<font size=6>%s</font>' % count, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % alumno, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % ps, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % cal2, styles['Normal'])
        ))
    if filtro == 3: #Todas las etapas
      data = [(
        Paragraph('<strong><font size=6>#</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Nombre y Apellido</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Primera Etapa. PT:{a}</font></strong>'.format(a = total_p), styles['Center']),
        Paragraph('<strong><font size=6>Calificación</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Segunda Etapa. PT:{a}</font></strong>'.format(a = total_s), styles['Center']),
        Paragraph('<strong><font size=6>Calificación</font></strong>', styles['Center'])
      )]
      # Aqui acomplamos los registros o datos a nuestra tabla data, estos seran los datos mostrados de bajo de los headers
      for x in range(0, len(mat)):
        aux = mat[x]
        count = str(x + 1)
        print(count)
        alumno = aux[9] + "," + aux[10]  # nombre y apellido
        pp = aux[5]  # puntaje primera etapa
        if aux[3]:  # Calificacion primera etapa
          cal = aux[3]
        else:
          cal = "Sin calificación"
        ps = aux[11]  # puntaje segunda etapa
        if aux[12]:
          cal2 = aux[12]
        else:
          cal2 = "Sin calificación"
        data.append((
          Paragraph('<font size=6>%s</font>' % count, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % alumno, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % pp, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % cal, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % ps, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % cal2, styles['Normal'])
        ))
    # Declaramamos que la tabla recibira como dato los datos anteriores y le damos la dimensiones a cada uno de nuestros campos
    table = Table(
      data,
      colWidths=[20, 140, 80, 60, 80, 60]
    )
    table.setStyle(
      TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
      ])
    )
    Story.append(table)
    pdf.build(Story)
    #Comprueba si existe el archivo y lo descarga para el usuario
    ruta = pathlib.Path('.')
    filename = "{a}_alumnos.pdf".format(a=datos[2])
    archivo = ruta / filename  # Si existe un pdf con su nombre
    print(archivo)
    if archivo.exists():
      return send_file(archivo, as_attachment=True)
  return redirect(url_for('misalumnos', id=id))
@app.route('/crearseman') #Creacion de Semana de Clases
def crearsemana():
  global band
  if band == 6:
    flash("","si")
    band =0
  if band == 3:
    flash("","eliminado")
    band = 0
  if band == 4:
    flash("", "noh")
    band = 0
  if band == 5:
    flash("", "yae")
    band = 0
  if band != 0:
    band = 0
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT cursos.id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_enfasis = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [1]
  mycursor.execute(sql, val)
  cursos_c = mycursor.fetchall()
  sql = "SELECT cursos.id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_enfasis = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [2]
  mycursor.execute(sql, val)
  cursos_s = mycursor.fetchall()
  return render_template('crear_sem.html', datos=datos, cursos_c = cursos_c, cursos_s = cursos_s)

@app.route('/cargar/<int:id>' , methods = ['GET', 'POST']) #Sin utilizar
def asignarpof(id):
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  sql = "SELECT id_matxcur, matxcur.id_curso, matxcur.id_materia, des_m, car_h FROM matxcur, materias WHERE matxcur.id_curso = %s and matxcur.id_materia = materias.id_materia "
  val = [id]
  mycursor.execute(sql, val)
  materias = mycursor.fetchall()
  materias = sorted(materias, key=lambda materias: materias[3])
  sql = "SELECT id_materia, id_profesor FROM matxpro WHERE id_curso = %s"
  val = [id]
  mycursor.execute(sql, val)
  profesores = mycursor.fetchall()
  print(cursos)
  print(materias)
  print(profesores)
  create = True
  global band
  if request.method == 'POST':
    dias = request.form.getlist(('dias'))
    hora_i = request.form.getlist(('hora_i'))
    hora_f = request.form.getlist(('hora_f'))
    #Cantidad de dias
    clu = 0
    cma = 0
    cmi = 0
    cju = 0
    cvi = 0
    csa = 0
    for x in range(0, len(dias)):
      aux1 = dias[x]
      aux2 = hora_i[x]
      aux3 = hora_f[x]
      if aux1 and aux2 and aux3:
        hora1 = aux2.split(':')
        hora2 = aux3.split(':')
        print(hora2)
        if len(hora1) == 1:
          aux2 = aux2 + ":00"
        if len(hora2) == 1:
          aux3 = aux3 + ":00"
        print(aux2)
        print(aux3)
        if aux1 == "Lunes":
          clu += 1
        if aux1 == "Martes":
          cma += 1
        if aux1 == "Miercoles":
          cmi += 1
        if aux1 == "Jueves":
          cju += 1
        if aux1 == "Viernes":
          cvi += 1
        if aux1 == "Sabado":
          csa += 1
        if aux2 == aux3:
          flash("","misma")
          print("misma hora")
          create = False
    if clu > 5 or cma > 5 or cmi > 5 or cju > 5 or cvi > 5 or csa > 5: # Comprueba que no se asignen muchas materias a ese dia
      print("dias")
      create = False
    if create == True:
      band += 1
      for x in range(0, len(dias)):
        aux1 = dias[x]
        aux2 = hora_i[x]
        aux3 = hora_f[x]
        if aux1 and aux2 and aux3:
          hora1 = aux2.split(':')
          hora2 = aux3.split(':')
          aux4 = materias[x]
          if len(hora1) == 1:
            aux2 = aux2 + ":00"
          if len(hora2) == 1:
            aux3 = aux3 + ":00"
          if aux1 == "Lunes":
            aux1 = 1
          if aux1 == "Martes":
            aux1 = 2
          if aux1 == "Miercoles":
            aux1 = 3
          if aux1 == "Jueves":
            aux1 = 4
          if aux1 == "Viernes":
            aux1 = 5
          if aux1 == "Sabado":
            aux1 = 6
          print('Datos:')
          print(aux1)
          print(aux2)
          print(aux3)
          print(aux4)
          mycursor = mydb.cursor()
          mycursor.execute(
            'INSERT INTO horarios (id_curso, id_materia, id_dia, hora_i, hora_f) VALUES (%s, %s ,%s, %s, %s)',
            (id, aux4[2], aux1, aux2, aux3))
          mydb.commit()
    print(band)
    print(dias)
    print(hora_i)
    print(hora_f)
    if band == 6:
      return redirect(url_for('crearsemana'))
  return render_template('crear_dias_mantenimiento.html', datos=datos, cursos = cursos[0], materias = materias, band = band)

@app.route('/cargar2/<int:id>' , methods = ['GET', 'POST'])
def cargarhora(id):
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  sql = "SELECT id_matxcur, matxcur.id_curso, matxcur.id_materia, des_m, car_h FROM matxcur, materias WHERE matxcur.id_curso = %s and matxcur.id_materia = materias.id_materia "
  val = [id]
  mycursor.execute(sql, val)
  materias = mycursor.fetchall()
  materias = sorted(materias, key=lambda materias: materias[3])
  global band
  create = True
  print(band)
  #Comprueba que exista un horario cargado
  for x in range(1, 7):
    sql = "SELECT id_horario, id_dia, hora_i, hora_f, des_m FROM horarios, materias WHERE id_curso = %s and id_dia = %s and horarios.id_materia = materias.id_materia ORDER BY horarios.hora_f ASC"
    val = [id, x]
    mycursor.execute(sql, val)
    if x == 1: #Lunes
      h_lu = mycursor.fetchall()
      print(h_lu)
    if x == 2: #Martes
      h_ma = mycursor.fetchall()
      print(h_ma)
    if x == 3:  # Miercoles
      h_mi = mycursor.fetchall()
      print(h_mi)
    if x == 4:  # Jueves
      h_ju = mycursor.fetchall()
      print(h_ju)
    if x == 5:  # Viernes
      h_vi = mycursor.fetchall()
      print(h_vi)
    if x == 6:  # Sabado
      h_sa = mycursor.fetchall()
      print(h_sa)
  #En caso que exista no permite crear otro horario
  if h_lu or h_ma or h_mi or h_ju or h_vi or h_sa:
    band = 5
    return redirect(url_for('crearsemana'))
  if request.method == 'POST':
    idmat = request.form.getlist(('idmat'))
    hora_i = request.form.getlist(('hora_i'))
    hora_f = request.form.getlist(('hora_f'))
    print(idmat)
    print(hora_i)
    print(hora_f)
    for x in range(0, len(idmat)):
      aux = idmat[x]
      if not aux:
        create = False
    if create == True:
      for x in range(0, len(idmat)):
        aux = idmat[x]
        aux2 = hora_i[x]
        aux3 = hora_f[x]
        sql = "SELECT id_mxp, id_profesor FROM matxpro WHERE id_materia = %s and id_curso = %s"
        val = [aux, id]
        mycursor.execute(sql, val)
        aux4 = mycursor.fetchall()
        if aux4:
          aux4 = aux4[0]
          aux4 = aux4[1]
          print(aux4)
        if aux4 and aux4 != 0: #Osea que tiene profesor
          mycursor = mydb.cursor()
          mycursor.execute(
            'INSERT INTO horarios (id_curso, id_materia, id_dia, hora_i, hora_f, id_profesor) VALUES (%s, %s ,%s, %s, %s, %s)',
            (id, aux, band + 1, aux2, aux3, aux4))
          mydb.commit()
        else:
          mycursor = mydb.cursor()
          mycursor.execute(
            'INSERT INTO horarios (id_curso, id_materia, id_dia, hora_i, hora_f, id_profesor) VALUES (%s, %s ,%s, %s, %s, %s)',
            (id, aux, band + 1, aux2, aux3, 0))
          mydb.commit()
      band += 1
      if band == 6:
        return redirect(url_for('crearsemana'))
  return render_template('crear_dias.html', datos=datos, cursos=cursos[0], materias=materias, band=band)

@app.route('/eliminarhorario/<int:id>' , methods = ['GET', 'POST'])
def eliminarhora(id):
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  mycursor = mydb.cursor()
  global band
  for x in range(1, 7):
    sql = "SELECT id_horario, id_dia, hora_i, hora_f, des_m FROM horarios, materias WHERE id_curso = %s and id_dia = %s and horarios.id_materia = materias.id_materia ORDER BY horarios.hora_f ASC"
    val = [id, x]
    mycursor.execute(sql, val)
    if x == 1:  # Lunes
      h_lu = mycursor.fetchall()
      print(h_lu)
    if x == 2:  # Martes
      h_ma = mycursor.fetchall()
      print(h_ma)
    if x == 3:  # Miercoles
      h_mi = mycursor.fetchall()
      print(h_mi)
    if x == 4:  # Jueves
      h_ju = mycursor.fetchall()
      print(h_ju)
    if x == 5:  # Viernes
      h_vi = mycursor.fetchall()
      print(h_vi)
    if x == 6:  # Sabado
      h_sa = mycursor.fetchall()
      print(h_sa)
  if request.method == 'POST':
    mycursor = mydb.cursor()
    sql = "DELETE FROM horarios WHERE id_curso = %s"
    val = [id]
    mycursor.execute(sql, val)
    mydb.commit()
    band = 3
    return redirect(url_for('crearsemana'))
  return render_template('eliminarhorario.html', cursos=cursos[0], datos=datos, h_lu=h_lu, h_ma=h_ma, h_mi=h_mi, h_ju=h_ju, h_vi=h_vi, h_sa=h_sa)
@app.route('/verhorario/<int:id>') #Ver Horarios
def verhorarioadmin(id):
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  mycursor = mydb.cursor()
  global band
  for x in range(1, 7):
    sql = "SELECT id_horario, id_dia, hora_i, hora_f, des_m FROM horarios, materias WHERE id_curso = %s and id_dia = %s and horarios.id_materia = materias.id_materia ORDER BY horarios.hora_f ASC"
    val = [id, x]
    mycursor.execute(sql, val)
    if x == 1: #Lunes
      h_lu = mycursor.fetchall()
      print(h_lu)
    if x == 2: #Martes
      h_ma = mycursor.fetchall()
      print(h_ma)
    if x == 3:  # Miercoles
      h_mi = mycursor.fetchall()
      print(h_mi)
    if x == 4:  # Jueves
      h_ju = mycursor.fetchall()
      print(h_ju)
    if x == 5:  # Viernes
      h_vi = mycursor.fetchall()
      print(h_vi)
    if x == 6:  # Sabado
      h_sa = mycursor.fetchall()
      print(h_sa)
  if datos[7] == 1:  # alumnos
    return render_template('verhorariosal.html', cursos=cursos[0], datos=datos, h_lu=h_lu, h_ma=h_ma, h_mi=h_mi,
                           h_ju=h_ju, h_vi=h_vi, h_sa=h_sa)
  if datos[6] == 3:  # Admin
    if h_lu and h_ma and h_mi and h_ju and h_vi and h_sa:
      return render_template('verhorariosad.html', cursos = cursos[0], datos = datos, h_lu = h_lu, h_ma = h_ma, h_mi = h_mi, h_ju = h_ju, h_vi = h_vi, h_sa= h_sa)
    else: #Si no existe un horario totalmente cargado
      band = 4
      return redirect(url_for('crearsemana'))



@app.route('/editarhorario/<string:id>' , methods = ['GET', 'POST'])
def editarhorario(id):
  datos = session['username']
  aux = id.split('D')
  id_c = int(aux[0])
  id_d = int(aux[1])
  #print(id_c)
  #print(id_d)
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id_c]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  sql = "SELECT id_horario, des_m FROM horarios, materias WHERE id_curso = %s and id_dia = %s and horarios.id_materia = materias.id_materia ORDER BY horarios.hora_f ASC"
  val = [id_c, id_d]
  mycursor.execute(sql, val)
  horarios = mycursor.fetchall()
  #print(horarios)
  sql = "SELECT id_matxcur, matxcur.id_curso, matxcur.id_materia, des_m, car_h FROM matxcur, materias WHERE matxcur.id_curso = %s and matxcur.id_materia = materias.id_materia "
  val = [id_c]
  mycursor.execute(sql, val)
  materias = mycursor.fetchall()
  materias = sorted(materias, key=lambda materias: materias[3])
  create = True
  if request.method == 'POST':
    idmat = request.form.getlist(('idmat'))
    print(idmat)
    for x in range(0, len(idmat)):
      aux = idmat[x]
      if not aux:
        create = False
    if create == True:
      for x in range(0, len(idmat)):
        aux = idmat[x]
        print(aux)
        sql = "SELECT id_mxp, id_profesor FROM matxpro WHERE id_materia = %s and id_curso = %s"
        val = [aux, id_c]
        mycursor.execute(sql, val)
        aux4 = mycursor.fetchall()
        print(aux4)
        aux5 = horarios[x]
        print(aux5)
        if aux4:
          aux4 = aux4[0]
          aux4 = aux4[1]
          print(aux4)
        if aux4 and aux4 != 0: #Osea que tiene profesor
          mycursor = mydb.cursor()
          aux5 = aux5[0]
          print(aux5)
          print(aux)
          sql = "UPDATE horarios SET id_materia = %s, id_profesor = %s WHERE id_horario = %s"
          val = (aux, aux4, aux5)
          mycursor.execute(sql, val)
          mydb.commit()
        else:
          mycursor = mydb.cursor()
          aux5 = aux5[0]
          print(aux5)
          print(aux)
          sql = "UPDATE horarios SET id_materia = %s, id_profesor = %s WHERE id_horario = %s"
          val = (aux, 0, aux5)
          mycursor.execute(sql, val)
          mydb.commit()
      return redirect(url_for('verhorarioadmin', id = id_c))
  return render_template('editarhorario.html', datos = datos, cursos=cursos[0], horarios = horarios, materias = materias, band = id_d - 1)
@app.route('/asistencia',  methods = ['GET', 'POST']) #Marcar Asistencia por Docente
def asistenciaprof():
  global user_datos #datos del profe para alerta
  global vector # hora de entrada
  global band
  global vector2 # hora de salida
  print(band)
  if band == 1:
    profesor = user_datos
    print(profesor)
    flash("", "entrada")
    band = 0
  if band == 2:
    profesor = user_datos
    print(profesor)
    print(vector)
    flash("", "salida")
    band = 0
  if band == 3: #ya marco
    profesor = user_datos
    print(profesor)
    print(vector2)
    print(vector)
    flash("", "no")
    band = 0
  if band == 4:
    profesor = user_datos
    print(profesor)
    flash("", "no_dia")
    band = 0
  if band == 5:
    band = 0
  user = userform.User(request.form)
  inf = datetime.now()
  #Extraemos la fecha
  fecha = datetime.strftime(inf, '%Y/%m/%d')
  #Extraemos la hora
  hora = datetime.strftime(inf, '%H:%M:%S')
  dia = datetime.today().weekday()
  print(fecha)
  print(hora)
  dia += 1  #Por que yo realize que Lunes = 1
  #dia += 1
  print(dia)
  profesor = user_datos
  mycursor = mydb.cursor()
  #Crea un nuevo dia para asistencia de los profes
  sql = "SELECT id_asisprof, fec_a FROM asistenciaprof WHERE fec_a = %s"
  val = [fecha]
  mycursor.execute(sql, val)
  compf = mycursor.fetchall()
  print(compf)
  compc = []
  if not compf:
    #Saca los ids de los profes que deben marcar hoy
    sql = "SELECT id_horario, id_profesor FROM horarios WHERE id_dia =%s"
    val = [dia]
    mycursor.execute(sql, val)
    idsp = mycursor.fetchall()
    print(idsp)
    cont = 0
    for x in range(0, len(idsp)):
      aux = idsp[x]
      if aux[1] != 0:
        if cont == 0:
          mycursor = mydb.cursor()
          mycursor.execute(
            'INSERT INTO asistenciaprof (id_profesor, fec_a) VALUES (%s, %s)',
            (aux[1], fecha))
          mydb.commit()
          cont = 1
          compc.append(aux[1])
        if aux[1] not in compc:
          print(aux[1])
          mycursor = mydb.cursor()
          mycursor.execute(
            'INSERT INTO asistenciaprof (id_profesor, fec_a) VALUES (%s, %s)',
            (aux[1], fecha))
          mydb.commit()
          compc.append(aux[1])
          print(compc)
  if request.method == 'POST':
    #Codigo del input que sale del escaneo de la tarjeta
    codigo = user.username.data
    codigo = codigo.split('P')
    print(codigo)
    #Buscamos por cedula del proesor
    sql = "SELECT * FROM profesores WHERE ci_p = %s"
    val = [codigo[1]]
    mycursor.execute(sql, val)
    profesor = mycursor.fetchall()
    if profesor:
      profesor = profesor[0]
      user_datos = profesor #guardar para las alertas
      print(profesor)
      #Comprueba que el profesor tenga clases hoy
      sql = "SELECT id_horario, id_curso, id_materia, id_dia, hora_i, hora_f FROM horarios WHERE id_profesor = %s and id_dia =%s  ORDER BY horarios.hora_i ASC"
      val = [profesor[0], dia]
      mycursor.execute(sql, val)
      ent = mycursor.fetchall()
      if ent:
        ent = ent[0]
        ent = ent[4]
      print("Entrada:")
      print(ent)
      sql = "SELECT id_horario, id_curso, id_materia, id_dia, hora_i, hora_f FROM horarios WHERE id_profesor = %s and id_dia =%s  ORDER BY horarios.hora_f DESC"
      val = [profesor[0], dia]
      mycursor.execute(sql, val)
      sal = mycursor.fetchall()
      if sal:
        sal = sal[0]
        sal = sal[5]
      print("Salida:")
      print(sal)
      if ent and sal:
        sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s and fec_a = %s"
        val = [profesor[0], fecha]
        mycursor.execute(sql, val)
        comp = mycursor.fetchall()
        print(comp)
        if comp:
          comp = comp[0]
          if comp[0] and not comp[1]:
            print("Ya Marco entrada, marca salida")
            vector = comp[0]  # entrada
            sql = "UPDATE asistenciaprof SET hora_s = %s WHERE id_profesor = %s and fec_a = %s"
            val = (hora, profesor[0], fecha)
            mycursor.execute(sql, val)
            mydb.commit()
            band = 2
            return redirect(url_for('asistenciaprof'))
          if comp[1]:
            vector = comp[0]
            vector2 = comp[1]  # salida
            print("Marco Salida, ya no puede marcar hoy")
            band = 3
            return redirect(url_for('asistenciaprof'))
          if not comp[0] and not comp[1]:
            print("No marco hoy, marca entrada")
            mycursor = mydb.cursor()
            sql = "UPDATE asistenciaprof SET hora_e = %s WHERE id_profesor = %s and fec_a = %s"
            val = (hora, profesor[0], fecha)
            mycursor.execute(sql, val)
            mydb.commit()
            band = 1
            return redirect(url_for('asistenciaprof'))
        else:
          print("No es su dia, no puede marcar")
          print(user_datos)
          band = 4
          return redirect(url_for('asistenciaprof'))
      else:
        band = 4
        return redirect(url_for('asistenciaprof'))
    else:
      print("Error no esta en el sistema")
      flash("","noesta")

  return render_template('asistenciaprof.html', user = user, profesor = profesor, vector= vector, vector2= vector2, hora = hora)

@app.route('/listadocursos') #Listado de Alumnos admin
def listadocursos():
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT cursos.id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_enfasis = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [1]
  mycursor.execute(sql, val)
  cursos_c = mycursor.fetchall()
  sql = "SELECT cursos.id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_enfasis = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [2]
  mycursor.execute(sql, val)
  cursos_s = mycursor.fetchall()
  print(cursos_s)
  return render_template('listadocursos.html', datos=datos, cursos_c = cursos_c, cursos_s= cursos_s)


@app.route('/listadoalumnos/<int:id>', methods = ['POST', 'GET']) #Ver todos los alumnos del curso
def listadoalumnos(id):
  global band
  global tipoins
  tipoins = 1
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  print(cursos)
  sql = "SELECT id_matxcur, matxcur.id_curso, matxcur.id_materia, des_m, car_h FROM matxcur, materias WHERE matxcur.id_curso = %s and matxcur.id_materia = materias.id_materia "
  val = [id]
  mycursor.execute(sql, val)
  materias = mycursor.fetchall()
  materias = sorted(materias, key=lambda materias: materias[3])
  print(materias)
  sql = "SELECT DISTINCT matxalum.id_alumno, nmb_a, ape_a FROM matxalum, alumnos WHERE matxalum.id_curso = %s and matxalum.id_alumno = alumnos.id_alumno"
  val = [id]
  mycursor.execute(sql, val)
  alum_t = mycursor.fetchall()
  alum_t = sorted(alum_t, key=lambda alum_t: alum_t[2])
  print(alum_t)
  band = id
  filtro = 0
  materia = []
  if request.method == 'POST':
    filtro = 1
    idmat = request.form.get("idmat")
    print(idmat)
    idmat = int(idmat)
    if idmat > 6: #Comprueba que no marque todos los alumnos
      sql = "SELECT DISTINCT matxalum.id_alumno, nmb_a, ape_a FROM matxalum, alumnos WHERE matxalum.id_curso = %s and matxalum.id_materia = %s and matxalum.id_alumno = alumnos.id_alumno "
      val = [id, idmat]
      mycursor.execute(sql, val)
      alum_t = mycursor.fetchall()
      alum_t = sorted(alum_t, key=lambda alum_t: alum_t[2])
      print(alum_t)
      sql = "SELECT des_m FROM materias WHERE id_materia = %s"
      val = [idmat]
      mycursor.execute(sql, val)
      materia = mycursor.fetchall()
      materia = materia[0]
    else:
      sql = "SELECT DISTINCT matxalum.id_alumno, nmb_a, ape_a FROM matxalum, alumnos WHERE matxalum.id_curso = %s and matxalum.id_alumno = alumnos.id_alumno"
      val = [id]
      mycursor.execute(sql, val)
      alum_t = mycursor.fetchall()
      alum_t = sorted(alum_t, key=lambda alum_t: alum_t[2])
      print(alum_t)
      filtro = 0
  return render_template('listadoalumnosad.html', datos=datos, cursos = cursos[0], materias = materias, alum_t = alum_t, filtro = filtro, materia = materia)


@app.route('/exportaralumnosad/<int:id>', methods = ['GET', 'POST'])
def exportaralumnosad(id):
  datos = session['username']
  # Datos del curso
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  #Saca los alumnos
  sql = "SELECT DISTINCT matxalum.id_alumno, nmb_a, ape_a, ci_a, edad FROM matxalum, alumnos WHERE matxalum.id_curso = %s and matxalum.id_alumno = alumnos.id_alumno"
  val = [id]
  mycursor.execute(sql, val)
  alum_t = mycursor.fetchall()
  alum_t = sorted(alum_t, key=lambda alum_t: alum_t[2])
  if request.method == 'POST':
    pdf = SimpleDocTemplate(
      "{a}{b}_alumnos.pdf".format(a=cursos[1], b=cursos[3]),
      pagesize=A4,
      rightMargin=inch,
      leftMargin=inch,
      topMargin=inch,
      bottomMargin=inch / 2
    )
    Story = []
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
    # Cabecera
    text = ''' <strong><font size=14>Alumnos matriculados</font></strong>
          '''
    text2 = '''
              <strong><font size=10>Diretor: {a}, {b}</font></strong>
          '''.format(a=datos[1], b=datos[2])
    text3 = '''
              <strong><font size=10>Curso: {a}, {b}</font></strong>
          '''.format(a=cursos[1], b=cursos[3])
    # Adjuntamos los titulos declarados mas arriba, osea las cabeceras
    Story.append(Paragraph(text2))
    Story.append(Paragraph(text3))
    Story.append(Paragraph(text, styles['Center']))
    Story.append(Spacer(1, 20))
    data = [(
      Paragraph('<strong><font size=6>#</font></strong>', styles['Center']),
      Paragraph('<strong><font size=6>Nombre y Apellido</font></strong>', styles['Center']),
      Paragraph('<strong><font size=6>Numero de Cedula</font></strong>', styles['Center']),
      Paragraph('<strong><font size=6>Edad</font></strong>', styles['Center'])
    )]
    # Aqui acomplamos los registros o datos a nuestra tabla data, estos seran los datos mostrados de bajo de los headers
    for x in range(0, len(alum_t)):
      aux = alum_t[x]
      count = str(x + 1)
      alumno = aux[1] + ", " + aux[2]
      ci = aux[3]
      edad = aux[4]
      data.append((
        Paragraph('<font size=6>%s</font>' % count, styles['Normal']),
        Paragraph('<font size=6>%s</font>' % alumno, styles['Normal']),
        Paragraph('<font size=6>%s</font>' % ci, styles['Normal']),
        Paragraph('<font size=6>%s</font>' % edad, styles['Normal']),
      ))
    # Declaramamos que la tabla recibira como dato los datos anteriores y le damos la dimensiones a cada uno de nuestros campos
    table = Table(
      data,
      colWidths=[20, 140, 80, 80]
    )
    table.setStyle(
      TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
      ])
    )
    Story.append(table)
    pdf.build(Story)
    # Comprueba si existe el archivo y lo descarga para el usuario
    ruta = pathlib.Path('.')
    filename = "{a}{b}_alumnos.pdf".format(a=cursos[1], b=cursos[3])
    archivo = ruta / filename  # Si existe un pdf con su nombre
    print(archivo)
    if archivo.exists():
      return send_file(archivo, as_attachment=True)
  return redirect(url_for('listadoalumnos', id=id))

@app.route('/moddatos/<int:id>', methods = ['POST', 'GET']) #Ver o modificar datos del alumno, listo
def moddatos(id):
  global band
  global tipoins
  if not tipoins:
    return redirect(url_for('listadocursos'))
  global idcurso
  if band == 3:
    flash("", "ok")
    band = 0
  if band == 8:
    flash("", "nod")
    band = 0
  user = userform.User(request.form)
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT * FROM alumnos WHERE id_alumno = %s"
  val = [id]
  mycursor.execute(sql, val)
  data = mycursor.fetchall()
  #print(tipoins)
  if data:
    data = data[0]
    print(data)
    idcurso = 0
  if not data:
    sql = "SELECT * FROM profesores WHERE id_profesor = %s"
    val = [id]
    mycursor.execute(sql, val)
    data = mycursor.fetchall()
    data = data[0]
    print(data)
    if not data:
      return redirect(url_for('listadocursos'))
  if request.method == 'POST':
    edad = request.form.get("edad")
    loca = request.form.get("loc")
    barr = request.form.get("bar")
    nume = request.form.get("num")
    emai = request.form.get("ema")
    nomt = request.form.get("nmt")
    numt = request.form.get("nut")
    print(edad)
    print(loca)
    print(barr)
    print(nume)
    print(emai)
    print(nomt)
    print(numt)
    if nomt and numt:
      print("mod alumno")
      sql = "UPDATE alumnos SET edad = %s, loc_a = %s, bar_a = %s, tel_a = %s, email = %s, nmb_tu = %s, tel_tu =%s" \
            " WHERE id_alumno = %s"
      val = (edad,loca,barr,nume,emai,nomt, numt, id)
      mycursor.execute(sql, val)
      mydb.commit()
      band = 3
      return redirect(url_for('moddatos', id = id))
    else:
      print("mod profe")
      sql = "UPDATE profesores SET edad = %s, loc_p = %s, bar_p = %s, tel_p = %s, email = %s WHERE id_profesor = %s"
      val = (edad, loca, barr, nume, emai, id)
      mycursor.execute(sql, val)
      mydb.commit()
      band = 3
      return redirect(url_for('moddatos', id=id))
  return render_template('moddatosal.html', datos=datos, data = data, band=band, user = user, tipo = tipoins, id = idcurso, id2 = id)


@app.route('/verdocumento/<int:id>', methods=['GET', 'POST'])
def verdocumento(id):
  datos = session['username']
  ruta = pathlib.Path('./media/')
  global band
  if request.method == "POST":
    filtro = int(request.form.get(('idfiltro')))
    if filtro <= 4:
      mycursor = mydb.cursor()
      sql = "SELECT id_alumno, ci_a FROM alumnos WHERE id_alumno = %s"
      val = [id]
      mycursor.execute(sql, val)
      data = mycursor.fetchall()
      print(data)
      if data:
        data = data[0]
        data = str(data[1])
        if filtro == 1:
          filename = "cedula_" + data + ".pdf"
        if filtro == 2:
          filename = "antecedente_" + data + ".pdf"
        if filtro == 3:
          filename = "autorizacion_" + data + ".pdf"
        if filtro == 4:
          filename = "ficha_" + data + ".pdf"
        archivo = ruta / filename
        print(archivo)
        if archivo.exists():
          print("El arhivo existe")
          return send_file(archivo, as_attachment=True)
        else:
          band = 8
        if datos[7] == 1:  # alumnos
          return redirect(url_for('misdatos'))
        if datos[6] == 3:  # Admin
          return redirect(url_for('moddatos', id=id))

      if datos[7] == 1:  # alumnos
        redirect(url_for('misdatos'))
      if datos[6] == 3:  # Admin
        return redirect(url_for('moddatos', id=id))
    else:
      mycursor = mydb.cursor()
      sql = "SELECT id_profesor, ci_p FROM profesores WHERE id_profesor = %s"
      val = [id]
      mycursor.execute(sql, val)
      data = mycursor.fetchall()
      print(data)
      print("profe")
      if data:
        data = data[0]
        data = str(data[1])
        if filtro == 5:
          filename = "cedulaprof_" + data + ".pdf"
        if filtro == 6:
          filename = "constancia_ingresos_" + data + ".pdf"
        if filtro == 7:
          filename = "constancia_cargos_" + data + ".pdf"
        archivo = ruta / filename
        print(archivo)
        if archivo.exists():
          print("El arhivo existe")
          return send_file(archivo, as_attachment=True)
        else:
          band = 8
        if datos[6] == 2:  # profesores
          return redirect(url_for('misdatos'))
        if datos[6] == 3:  # Admin
          return redirect(url_for('moddatos', id=id))

      if datos[6] == 2:  # profesores
        redirect(url_for('misdatos'))
      if datos[6] == 3:  # Admin
        return redirect(url_for('moddatos', id=id))

@app.route('/estado/<int:id>', methods = ['POST', 'GET'])
def verestado(id):
  datos = session['username']
  print(id)
  #Saca el idcurso
  mycursor = mydb.cursor()
  sql = "SELECT id_alumno, id_curso, nmb_a, ape_a FROM alumnos WHERE id_alumno = %s"
  val = [id]
  mycursor.execute(sql, val)
  idcurso = mycursor.fetchall()
  idcurso = idcurso[0]
  print(idcurso[1])
  #Puntaje total en primera etapa
  sql = "SELECT id_alumno, matxalum.id_materia, des_m, cal, id_curso, pun_ac, ano_m, id_profesor FROM matxalum, materias WHERE id_alumno = %s and id_curso = %s and matxalum.id_materia = materias.id_materia"
  val = [id, idcurso[1]]
  mycursor.execute(sql, val)
  materias = mycursor.fetchall()
  print(materias)
  puntos_t = []
  estado_a = []
  total = 0
  for x in range(0, len(materias)): #Busca trabajos asociados a la materia en segunda etapa
    aux = materias[x]
    #print(aux)
    sql = "SELECT id_materia, pun_t FROM trabajos WHERE id_materia = %s and id_curso = %s and etapa = %s"
    val = [aux[1], idcurso[1], 1]
    mycursor.execute(sql, val)
    trabajos = mycursor.fetchall()
    if trabajos:
      #print(trabajos)
      for x in range(0, len(trabajos)): #puntaje total por materia
        aux2 = trabajos[x]
        aux3 = int(aux2[1])
        total = total + aux3
      puntos_t.append(total)
      total = 0
    else:
      puntos_t.append(trabajos)
  print(puntos_t)
  for x in range(0, len(materias)): #Calcula el estado en la materia
    aux = materias[x]
    aux2 = puntos_t[x]
    if aux2:
      estado = (70 * aux2) / 100 #Calcula el 70% del total
      if aux[5] >= estado:
        estado = "bien"
        #print(estado)
        estado_a.append(estado)
      else:
        estado = "mal"
        #print(estado)
        estado_a.append(estado)
    else:
      estado = "bien"
      estado_a.append(estado)
  print(estado_a)
  # Puntaje total en segunda etapa
  sql = "SELECT id_alumno, matxalum.id_materia, des_m, cal2, id_curso, pun_ac2, ano_m, id_profesor FROM matxalum, materias WHERE id_alumno = %s and id_curso = %s and matxalum.id_materia = materias.id_materia"
  val = [id, idcurso[1]]
  mycursor.execute(sql, val)
  materias2 = mycursor.fetchall()
  print(materias2)
  puntos_t2 = []
  estado_a2 = []
  total2 = 0
  pre = [] #presentes
  aus = [] #ausentes
  for x in range(0, len(materias2)): #Busca trabajos asociados a la materia en segunda etapa
    aux = materias2[x]
    #print(aux)
    sql = "SELECT id_materia, pun_t FROM trabajos WHERE id_materia = %s and id_curso = %s and etapa = %s"
    val = [aux[1], idcurso[1], 2]
    mycursor.execute(sql, val)
    trabajos2 = mycursor.fetchall()
    # Asistencias
    sql = "SELECT id_alumno, asistio FROM asistenciaalum WHERE id_alumno = %s and id_curso = %s and id_materia = %s"
    val = [id, idcurso[1], aux[1]]
    mycursor.execute(sql, val)
    compa = mycursor.fetchall()
    #print(compa)
    countp = 0
    counta = 0
    if compa:
      for x in range(0, len(compa)):
        a = compa[x]
        print(compa)
        if a[1] == 'P':
          countp += 1
        else:
          counta += 1
        if x == len(compa) - 1:
          pre.append(countp)
          aus.append(counta)
          print(countp, counta)
    else:
      pre.append(countp)
      aus.append(counta)
    if trabajos2:
      #print(trabajos)
      for x in range(0, len(trabajos2)): #puntaje total por materia
        aux2 = trabajos2[x]
        aux3 = int(aux2[1])
        total2 = total2 + aux3
      puntos_t2.append(total2)
      total2 = 0
    else:
      puntos_t2.append(trabajos2)
  print(puntos_t2)
  print(pre, aus)
  for x in range(0, len(materias2)): #Calcula el estado en la materia
    aux = materias2[x]
    aux2 = puntos_t2[x]
    if aux2:
      estado = (70 * aux2) / 100 #Calcula el 70% del total

      if aux[5] >= estado:
        estado = "bien"
        #print(estado)
        estado_a2.append(estado)
      else:
        estado = "mal"
        #print(estado)
        estado_a2.append(estado)
    else:
      estado = "bien"
      estado_a2.append(estado)
  print(estado_a2)
  estado_a3 = []
  #Calcula el estado del alumno en su asistencia por materia
  for x in range(0, len(materias2)):
    aux = aus[x]
    aux2 = pre[x]
    total = aux + aux2
    if total > 0:
      print(total)
      estado = (60 * total) / 100
      print(estado)
      if aux2 >= estado:
        estado = "bien"
        estado_a3.append(estado)
      else:
        estado = "mal"
        estado_a3.append(estado)
    else:
      estado = "bien"
      estado_a3.append(estado)
  print(estado_a3)
  if request.method == 'POST':
    filtro = int(request.form.get("idfiltro"))
    print(filtro)
    if filtro == 1:
      return render_template('verestadoadmin.html', datos=datos, materias=materias, puntos_t=puntos_t,
                             estado_a=estado_a, materias2=materias2, puntos_t2=puntos_t2, estado_a2=estado_a2, presentes=pre,
                             ausencias=aus, estado_a3=estado_a3, data=idcurso, filtro=1)
    if filtro == 2:
      return render_template('verestadoadmin.html', datos=datos, materias=materias, puntos_t=puntos_t,
                             estado_a=estado_a, materias2=materias2, puntos_t2=puntos_t2, estado_a2=estado_a2,
                             presentes=pre,
                             ausencias=aus, estado_a3=estado_a3, data=idcurso, filtro=2)
    if filtro == 3:
      return render_template('verestadoadmin.html', datos=datos, materias=materias, puntos_t=puntos_t,
                             estado_a=estado_a, materias2=materias2, puntos_t2=puntos_t2, estado_a2=estado_a2,
                             presentes=pre,
                             ausencias=aus, estado_a3=estado_a3, data=idcurso, filtro=3)
    if filtro == 4:
      return redirect(url_for('verestado', id=id))
  return render_template('verestadoadmin.html', datos=datos, materias = materias, puntos_t = puntos_t, estado_a=estado_a,
                         materias2 = materias2, puntos_t2 = puntos_t2, estado_a2 = estado_a2, presentes = pre, ausencias = aus,
                         estado_a3 = estado_a3, data = idcurso, filtro = 4)
@app.route('/listadodocentes/<int:id>', methods = ['POST', 'GET']) #Ver todos los docentes del curso
def listadodocentes(id):
  global tipoins
  tipoins = 2
  datos = session['username']
  global idcurso
  idcurso = id
  #Datos del curso
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  #Saca las materias del curso
  sql = "SELECT id_matxcur, matxcur.id_curso, matxcur.id_materia, des_m, car_h FROM matxcur, materias WHERE matxcur.id_curso = %s and matxcur.id_materia = materias.id_materia "
  val = [id]
  mycursor.execute(sql, val)
  materias = mycursor.fetchall()
  materias = sorted(materias, key=lambda materias: materias[3])
  print(materias)
  #Saca los profesores
  sql = "SELECT DISTINCT matxpro.id_profesor, nmb_p, ape_p, des_m FROM matxpro, profesores, materias WHERE matxpro.id_curso = %s " \
        "and matxpro.id_profesor = profesores.id_profesor and matxpro.id_materia = materias.id_materia"
  val = [id]
  mycursor.execute(sql, val)
  profe_t = mycursor.fetchall()
  profe_t = sorted(profe_t, key=lambda profe_t: profe_t[2])
  print(profe_t)
  materia = []
  filtro = 0
  if request.method == 'POST':
    idmat = request.form.get("idmat")
    print(idmat)
    idmat = int(idmat)
    filtro = 1
    if idmat > 6: #Comprueba que no marque todos los docentes
      sql = "SELECT DISTINCT matxpro.id_profesor, nmb_p, ape_p, des_m FROM matxpro, profesores, materias WHERE matxpro.id_curso = %s and matxpro.id_materia = %s " \
            "and matxpro.id_profesor = profesores.id_profesor and matxpro.id_materia = materias.id_materia "
      val = [id, idmat]
      mycursor.execute(sql, val)
      profe_t = mycursor.fetchall()
      profe_t = sorted(profe_t, key=lambda profe_t: profe_t[2])
      print(profe_t)
      if profe_t:
        materia = profe_t[0]
      else:
        flash("","nop")
        return redirect(url_for('listadodocentes', id=id))
  return render_template('listadodocentesad.html', datos=datos, cursos = cursos, materias = materias, profesores = profe_t, materia = materia, filtro = filtro)


@app.route('/exportardocentesad/<int:id>', methods = ['GET', 'POST'])
def exportardocentesad(id):
  datos = session['username']
  #Datos del curso
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  # Saca los profesores
  sql = "SELECT DISTINCT matxpro.id_profesor, nmb_p, ape_p, des_m FROM matxpro, profesores, materias WHERE matxpro.id_curso = %s " \
        "and matxpro.id_profesor = profesores.id_profesor and matxpro.id_materia = materias.id_materia"
  val = [id]
  mycursor.execute(sql, val)
  profe_t = mycursor.fetchall()
  profe_t = sorted(profe_t, key=lambda profe_t: profe_t[2])
  print(profe_t)
  if request.method == 'POST':
    pdf = SimpleDocTemplate(
      "{a}{b}_docentes.pdf".format(a=cursos[1], b = cursos[3]),
      pagesize=A4,
      rightMargin=inch,
      leftMargin=inch,
      topMargin=inch,
      bottomMargin=inch / 2
    )
    Story = []
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
    # Cabecera
    text = ''' <strong><font size=14>Docentes </font></strong>
       '''
    text2 = '''
           <strong><font size=10>Diretora: {a}, {b}</font></strong>
       '''.format(a=datos[1], b=datos[2])
    text3 = '''
           <strong><font size=10>Curso: {a}, {b}</font></strong>
       '''.format(a=cursos[1], b=cursos[3])
    # Adjuntamos los titulos declarados mas arriba, osea las cabeceras
    Story.append(Paragraph(text2))
    Story.append(Paragraph(text3))
    Story.append(Paragraph(text, styles['Center']))
    Story.append(Spacer(1, 20))
    data = [(
      Paragraph('<strong><font size=6>#</font></strong>', styles['Center']),
      Paragraph('<strong><font size=6>Nombre y Apellido</font></strong>', styles['Center']),
      Paragraph('<strong><font size=6>Disciplina a su cargo</font></strong>', styles['Center'])
    )]
    # Aqui acomplamos los registros o datos a nuestra tabla data, estos seran los datos mostrados de bajo de los headers
    for x in range(0, len(profe_t)):
      aux = profe_t[x]
      count = str(x + 1)
      docente = aux[1] + "," + aux[2]
      materia = aux[3]
      data.append((
        Paragraph('<font size=6>%s</font>' % count, styles['Normal']),
        Paragraph('<font size=6>%s</font>' % docente, styles['Normal']),
        Paragraph('<font size=6>%s</font>' % materia, styles['Normal'])
      ))
    # Declaramamos que la tabla recibira como dato los datos anteriores y le damos la dimensiones a cada uno de nuestros campos
    table = Table(
       data,
      colWidths=[20, 140, 140]
    )
    table.setStyle(
      TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
      ])
    )
    Story.append(table)
    pdf.build(Story)
    # Comprueba si existe el archivo y lo descarga para el usuario
    ruta = pathlib.Path('.')
    filename = "{a}{b}_docentes.pdf".format(a=cursos[1], b = cursos[3])
    archivo = ruta / filename  # Si existe un pdf con su nombre
    print(archivo)
    if archivo.exists():
      return send_file(archivo, as_attachment=True)
  return redirect(url_for('listadodocentes', id=id))
@app.route('/asignarprofe', methods = ['POST', 'GET']) #Listado de Alumnos admin  #falta swa
def asignarprofe():
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT id_profesor, nmb_p, ape_p, ci_p FROM profesores"
  val = []
  mycursor.execute(sql, val)
  data = mycursor.fetchall()
  data = sorted(data, key=lambda data: data[2])
  print(data)
  estado = 3
  if request.method == 'POST':
    estado = request.form.get("estado")
    estado = int(estado)
    print(estado)
    if estado == 1 or estado == 0:
      mycursor = mydb.cursor()
      sql = "SELECT id_profesor, nmb_p, ape_p, ci_p FROM profesores WHERE estado = %s"
      val = [estado]
      mycursor.execute(sql, val)
      data = mycursor.fetchall()
      data = sorted(data, key=lambda data: data[2])
      print(data)
      return render_template('asignarprofe.html', datos=datos, data=data, estado = estado)
    else:
      return redirect(url_for('asignarprofe'))

  return render_template('asignarprofe.html', datos=datos, data = data, estado = estado)

@app.route('/asignarmaterias/<int:id>', methods = ['POST', 'GET']) #falta swa
def asignarmaterias(id):
  global user_datos
  user_datos = id
  global band
  datos = session['username']
  if band == 1:
    flash("","bien")
    band = 0
  if band == 4:
    flash("","mal")
    band = 0
  mycursor = mydb.cursor()
  sql = "SELECT id_profesor, nmb_p, ape_p, ci_p FROM profesores WHERE id_profesor = %s"
  val = [id]
  mycursor.execute(sql, val)
  profesor = mycursor.fetchall()
  profesor = profesor[0]
  print(profesor)
  sql = "SELECT matxpro.id_materia, des_m, des_c, des_e, matxpro.id_curso FROM matxpro, materias, cursos, enfasis WHERE id_profesor = %s " \
        "and matxpro.id_materia = materias.id_materia and matxpro.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  materias = mycursor.fetchall()
  print(materias)
  fecha = datetime.now()
  fecha = datetime.strftime(fecha, '%Y')
  print(fecha)
  mycursor = mydb.cursor(buffered=True)
  for x in range(1, 7):
    #print(x)
    if x <= 3:
      sql = "SELECT matxcur.id_curso, materias.id_materia, des_m FROM matxcur, materias WHERE id_curso = %s and id_enfasis = %s and " \
            "materias.id_materia = matxcur.id_materia"
      val = [x, 1]
      mycursor.execute(sql, val)
      if x == 1:
        con_p = mycursor.fetchall()
        #print(con_p)
      if x == 2:
        con_s = mycursor.fetchall()
        #print(con_s)
    if x == 3:
      sql = "SELECT matxcur.id_curso, materias.id_materia, des_m FROM matxcur, materias WHERE id_curso = %s and id_enfasis = %s and " \
            "materias.id_materia = matxcur.id_materia"
      val = [x, 2]
      mycursor.execute(sql, val)
      soc_p = mycursor.fetchall()
      #print(soc_p)
    if x == 4:
      sql = "SELECT matxcur.id_curso, materias.id_materia, des_m FROM matxcur, materias WHERE id_curso = %s and id_enfasis = %s and " \
            "materias.id_materia = matxcur.id_materia"
      val = [x, 1]
      mycursor.execute(sql, val)
      con_t = mycursor.fetchall()
      #print(con_t)
    if x > 4:
      sql = "SELECT matxcur.id_curso, materias.id_materia, des_m FROM matxcur, materias WHERE id_curso = %s and id_enfasis = %s and " \
            "materias.id_materia = matxcur.id_materia"
      val = [x, 2]
      mycursor.execute(sql, val)
      if x == 5:
        soc_s = mycursor.fetchall()
        #print(soc_s)
      if x == 6:
        soc_t = mycursor.fetchall()
        #print(soc_t)
  if request.method == "POST":
    idmat = request.form.getlist(('idmat')) #Aca si tengo los id de las materias en q se insc
    print(idmat)
    if idmat:
      for x in range(0, len(idmat)):
        aux = idmat[x]
        aux2 = aux.split(',')
        aux3 = int(aux2[0])
        aux2 = int(aux2[1])
        print((aux2))
        print((aux3))
        mycursor = mydb.cursor()
        sql = "SELECT * FROM matxpro WHERE id_materia = %s and id_curso = %s"
        val = [aux2, aux3]
        mycursor.execute(sql, val)
        comp_m = mycursor.fetchall()
        print(comp_m)
        if comp_m:
            comp_m = comp_m[0]
            print(comp_m)
            if comp_m[2] != 0:
              band = 4
              print("ya tiene profesor")
            if comp_m[2] == 0:
              sql = "UPDATE matxpro SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
              val = (id,aux2, aux3)
              mycursor.execute(sql, val)
              mydb.commit()
              # Asigna el profesor a los alumnos
              sql = "UPDATE matxalum SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
              val = (id, aux2, aux3)
              mycursor.execute(sql, val)
              mydb.commit()
              # Asigna el profesor al horario
              sql = "UPDATE horarios SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
              val = (id, aux2, aux3)
              mycursor.execute(sql, val)
              mydb.commit()
              #Comprueba si cargo un trabajo
              sql = "SELECT id_trabajo, id_materia FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s"
              val = [0, aux2,aux3]
              mycursor.execute(sql, val)
              trabajos = mycursor.fetchall()
              print(trabajos)
              if trabajos:
                for x in range(0, len(trabajos)):
                  aux = trabajos[x]
                  id_t = aux[0]
                  sql = "UPDATE trabajos SET id_profesor = %s WHERE id_trabajo = %s"
                  val = (id, id_t)
                  mycursor.execute(sql, val)
                  mydb.commit()
              print("se asigno, el profesor anterior fue dado de baja")
        else:
          mycursor.execute(
            'INSERT INTO matxpro (id_materia, id_profesor, id_curso, fecha) VALUES (%s, %s, %s, %s)',
            (aux2, id, aux3, fecha))
          mydb.commit()
          band = 1
          print(band)
          print("se le asigna")
          #Asigna el profesor a los alumnos
          sql = "UPDATE matxalum SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
          val = (id, aux2, aux3)
          mycursor.execute(sql, val)
          mydb.commit()
          # Asigna el profesor al horario
          sql = "UPDATE horarios SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
          val = (id, aux2, aux3)
          mycursor.execute(sql, val)
          mydb.commit()
    #En caso que el profe este de baja, lo pone activo
    sql = "UPDATE profesores SET estado = %s WHERE id_profesor = %s"
    val = (1, id)
    mycursor.execute(sql, val)
    mydb.commit()
    return redirect(url_for('asignarmaterias', id=id))
  return render_template('asginarmaterias.html', datos=datos, profesor = profesor,
                         soc_p = soc_p, soc_s = soc_s, soc_t = soc_t, con_p=con_p, con_s = con_s, con_t = con_t, materias = materias)

@app.route('/dardebajaprofe/<int:id>', methods = ['POST', 'GET']) #Se debe guardar en el log,  #falta swa
def dardebajaprofe(id):
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT id_profesor, nmb_p, ape_p, ci_p FROM profesores WHERE id_profesor = %s"
  val = [id]
  mycursor.execute(sql, val)
  profesor = mycursor.fetchall()
  profesor = profesor[0]
  print(profesor)
  sql = "SELECT matxpro.id_materia, des_m, des_c, des_e, matxpro.id_curso FROM matxpro, materias, cursos, enfasis WHERE id_profesor = %s " \
        "and matxpro.id_materia = materias.id_materia and matxpro.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  materias = mycursor.fetchall()
  print(materias)
  fecha = datetime.now()
  fecha = datetime.strftime(fecha, '%Y')
  print(fecha)
  if request.method == "POST":
    sql = "SELECT id_trabajo, id_materia FROM trabajos WHERE id_profesor = %s"
    val = [id]
    mycursor.execute(sql, val)
    trabajos = mycursor.fetchall()
    if materias:
      for x in range(0, len(materias)):
        aux = materias[x]
        id_m = aux[0]
        id_c = aux[4]
        #Da de baja en materia x profesor
        sql = "UPDATE matxpro SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
        val = (0, id_m, id_c)
        mycursor.execute(sql, val)
        mydb.commit()
        # Da de baja en profesor a los alumnos
        sql = "UPDATE matxalum SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
        val = (0, id_m, id_c)
        mycursor.execute(sql, val)
        mydb.commit()
        # Da de baja al profesor en horario
        sql = "UPDATE horarios SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
        val = (0, id_m, id_c)
        mycursor.execute(sql, val)
        mydb.commit()
    if trabajos:
      print(trabajos)
      for x in range(0, len(trabajos)):
        aux = trabajos[x]
        id_t = aux[0]
        sql = "UPDATE trabajos SET id_profesor = %s WHERE id_trabajo = %s"
        val = (0, id_t)
        mycursor.execute(sql, val)
        mydb.commit()
    sql = "UPDATE profesores SET estado = %s WHERE id_profesor = %s"
    val = (0, id)
    mycursor.execute(sql, val)
    mydb.commit()
    return redirect(url_for('asignarprofe'))
  return render_template('dardebajaprofe.html', datos=datos, profesor = profesor, materias = materias)

@app.route('/eliminarmatxpro/<string:id>') #falta swa
def eliminarmatxpro(id):
  print(id)
  global user_datos
  print(user_datos)
  aux = id.split('C')
  id_m = aux[0]
  id_c = aux[1]
  mycursor = mydb.cursor()
  # Da de baja en materia x profesor
  sql = "UPDATE matxpro SET id_profesor = %s WHERE id_materia = %s and id_curso = %s and id_profesor = %s"
  val = (0, id_m, id_c, user_datos)
  mycursor.execute(sql, val)
  mydb.commit()
  # Da de baja en profesor a los alumnos
  sql = "UPDATE matxalum SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
  val = (0, id_m, id_c)
  mycursor.execute(sql, val)
  mydb.commit()
  # Da de baja al profesor en horario
  sql = "UPDATE horarios SET id_profesor = %s WHERE id_materia = %s and id_curso = %s"
  val = (0, id_m, id_c)
  mycursor.execute(sql, val)
  mydb.commit()
  #Si tiene trabajos..
  sql = "SELECT id_trabajo, id_materia FROM trabajos WHERE id_materia = %s and id_curso = %s and id_profesor = %s"
  val = [id_m, id_c, user_datos]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  print(trabajos)
  if trabajos:
    for x in range(0, len(trabajos)):
      aux = trabajos[x]
      id_t = aux[0]
      sql = "UPDATE trabajos SET id_profesor = %s WHERE id_trabajo = %s"
      val = (0, id_t)
      mycursor.execute(sql, val)
      mydb.commit()
  return redirect(url_for('asignarmaterias', id = user_datos))


@app.route('/vermiasistencia', methods = ['POST', 'GET'])
def verasistenciaprofe():
  datos = session['username']
  print(datos)
  #Sacamos las veces que marco asistencia
  mycursor = mydb.cursor()
  sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC LIMIT 7"
  val = [datos[0]]
  mycursor.execute(sql, val)
  asistencias = mycursor.fetchall()
  print(asistencias)
  dt = 0  # Dias trabajados
  dp = 0  # Dias no trabajados
  dr = 0  # Dias que no marco salida
  for x in range(0, len(asistencias)):
    aux = asistencias[x]
    if aux[0] or aux[1]:
      dt += 1
    if not aux[0] and not aux[1]:
      dp += 1
    if aux[0] and not aux[1]:
      dr += 1
  print(dt, dp, dr)
  if request.method == "POST":
    dt = 0
    dp = 0
    dr = 0
    filtro = int(request.form.get("idfiltro"))
    print(filtro)
    if filtro == 2:
      sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC LIMIT 31"
      val = [datos[0]]
      mycursor.execute(sql, val)
      asistencias = mycursor.fetchall()
      for x in range(0 , len(asistencias)):
        aux = asistencias[x]
        if aux[0] or aux[1]:
          dt += 1
        if not aux[0] and not aux[1]:
          dp += 1
        if aux[0] and not aux[1]:
          dr += 1
      return render_template('verasistenciaprofe.html', datos=datos, asistencias=asistencias, filtro=filtro, dt = dt, dp = dp, dr = dr)
    if filtro == 3:
      sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC"
      val = [datos[0]]
      mycursor.execute(sql, val)
      asistencias = mycursor.fetchall()
      for x in range(0 , len(asistencias)):
        aux = asistencias[x]
        if aux[0] or aux[1]:
          dt += 1
        if not aux[0] and not aux[1]:
          dp += 1
        if aux[0] and not aux[1]:
          dr += 1
      return render_template('verasistenciaprofe.html', datos=datos, asistencias=asistencias, filtro=filtro, dt = dt, dp = dp, dr = dr)
    if filtro == 1:
      return redirect(url_for('verasistenciaprofe'))
  return render_template('verasistenciaprofe.html', datos=datos, asistencias=asistencias, filtro = 1, dt = dt, dp = dp, dr = dr)

@app.route('/verasisad/<int:id>', methods = ['POST', 'GET'])
def verasisad(id):
  print(id)
  datos = session['username']
  #datos del profe
  mycursor = mydb.cursor()
  sql = "SELECT id_profesor, nmb_p, ape_p FROM profesores WHERE id_profesor = %s"
  val = [id]
  mycursor.execute(sql, val)
  profe = mycursor.fetchall()
  profe = profe[0]
  print(profe)
  # Sacamos las veces que marco asistencia
  mycursor = mydb.cursor()
  sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC LIMIT 7"
  val = [id]
  mycursor.execute(sql, val)
  asistencias = mycursor.fetchall()
  print(asistencias)

  dt = 0  # Dias trabajados
  dp = 0  # Dias no trabajados
  dr = 0  # Dias que no marco salida
  for x in range(0, len(asistencias)):
    aux = asistencias[x]
    if aux[0] or aux[1]:
      dt += 1
    if not aux[0] and not aux[1]:
      dp += 1
    if aux[0] and not aux[1]:
      dr += 1
  print(dt, dp, dr)
  if request.method == "POST":
    dt = 0
    dp = 0
    dr = 0
    filtro = int(request.form.get("idfiltro"))
    print(filtro)
    if filtro == 2:
      sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC LIMIT 31"
      val = [datos[0]]
      mycursor.execute(sql, val)
      asistencias = mycursor.fetchall()
      for x in range(0, len(asistencias)):
        aux = asistencias[x]
        if aux[0] or aux[1]:
          dt += 1
        if not aux[0] and not aux[1]:
          dp += 1
        if aux[0] and not aux[1]:
          dr += 1
      return render_template('verasisad.html', datos=datos, asistencias=asistencias, profe=profe, filtro=filtro, dt=dt, dp=dp, dr=dr)
    if filtro == 3:
      sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC"
      val = [datos[0]]
      mycursor.execute(sql, val)
      asistencias = mycursor.fetchall()
      for x in range(0 , len(asistencias)):
        aux = asistencias[x]
        if aux[0] or aux[1]:
          dt += 1
        if not aux[0] and not aux[1]:
          dp += 1
        if aux[0] and not aux[1]:
          dr += 1
      return render_template('verasisad.html', datos=datos,asistencias=asistencias, profe=profe, filtro=filtro, dt=dt, dp=dp, dr=dr)
    if filtro == 1:
      return redirect(url_for('verasisad', id=id))
  return render_template('verasisad.html', datos=datos, asistencias=asistencias, profe = profe, filtro = 1, dt = dt, dp = dp, dr = dr)

@app.route('/llamarlista/<int:id>', methods = ['POST', 'GET'])
def asistenciaalumnos(id):
  datos = session['username']
  #print(id)
  #Saca el curso relacionado
  mycursor = mydb.cursor()
  sql = "SELECT des_m, des_c, des_e, cursos.id_curso FROM matxpro, materias, cursos, enfasis WHERE id_profesor = %s and matxpro.id_materia = %s and matxpro.id_materia = materias.id_materia " \
        "and matxpro.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  #Saca los alumnos matriculados
  sql = "SELECT matxalum.id_alumno, matxalum.id_materia, matxalum.id_curso, nmb_a, ape_a FROM matxalum, alumnos WHERE id_profesor = %s and matxalum.id_materia = %s and matxalum.id_alumno = alumnos.id_alumno"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  mat = mycursor.fetchall()
  print(mat)
  # Ordena alfabetaicamente los alumnos del vector...
  mat = sorted(mat, key=lambda mat: mat[4])
  #Estados posibles
  lista = ["P", "A"]
  # Extraemos la fecha
  inf = datetime.now()
  fecha = datetime.strftime(inf, '%Y/%m/%d')
  #Dia
  dia = datetime.today().weekday()
  dia += 1
  print(dia)
  # Comprueba que el profesor tenga clases hoy
  sql = "SELECT id_horario, id_curso, id_materia, id_dia, hora_i, hora_f FROM horarios WHERE id_profesor = %s and id_dia =%s and id_materia = %s and id_curso = %s ORDER BY horarios.hora_i ASC"
  val = [datos[0], dia, id, cursos[3]]
  mycursor.execute(sql, val)
  comp = mycursor.fetchall()
  print(comp)
  # Comprueba que aun no halla llamado lista hoy
  sql = "SELECT DISTINCT fecha FROM asistenciaalum WHERE fecha = %s"
  val = [fecha]
  mycursor.execute(sql, val)
  comp2 = mycursor.fetchall()
  print(comp2)
  #Alertas
  global band
  #Si no tiene clases hoy redirige
  if not comp:
    band = 7
    return redirect(url_for('miscursos'))
  #Si ya llamo clases redirige
  if comp2:
    band = 5
    return redirect(url_for('miscursos'))
  #Comprobar que cargo bien
  create = True
  if request.method == "POST":
    for x in range(0, len(mat)):
      aux = request.form.getlist(('{a}'.format(a=x + 1)))
      if not aux:
        create = False
    if create == True:
      for x in range(0, len(mat)):
        aux = request.form.getlist(('{a}'.format(a=x + 1)))
        aux2 = mat[x]
        #Guarda por alumno si estuvo presente o ausente
        mycursor.execute(
          'INSERT INTO asistenciaalum (id_alumno, id_materia, fecha, asistio, id_curso) VALUES (%s, %s, %s, %s, %s)',
          (aux2[0], id, fecha, aux[0], aux2[2]))
        mydb.commit()
      band = 6
      return redirect(url_for('miscursos'))
    else:
      flash("","ec")
  return render_template('llamarlista.html', datos=datos, cursos = cursos, alumnos = mat, listas = lista, fecha = fecha, id = id)

@app.route('/verasistencias/<int:id>', methods = ['POST', 'GET'])
def verasistencias(id):
  datos = session['username']
  mycursor = mydb.cursor()
  #Saca el curso y el nombre de la materia
  sql = "SELECT des_m, des_c, des_e, cursos.id_curso FROM matxpro, materias, cursos, enfasis WHERE id_profesor = %s and matxpro.id_materia = %s and matxpro.id_materia = materias.id_materia " \
        "and matxpro.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  #Saca los datos de la asistencia
  sql = "SELECT DISTINCT fecha FROM asistenciaalum WHERE id_materia = %s and id_curso = %s  ORDER BY asistenciaalum.fecha DESC LIMIT 7"
  val = [id, cursos[3]]
  mycursor.execute(sql, val)
  asis = mycursor.fetchall()
  print(asis)
  #ID MATERIA
  global user_datos
  user_datos = id
  global vector
  vector = cursos[3]
  if request.method == "POST":
    filtro = int(request.form.get("idfiltro"))
    print(filtro)
    if filtro == 1:
      return redirect(url_for('verasistencias', id = id))
    if filtro == 2:
      sql = "SELECT DISTINCT fecha FROM asistenciaalum WHERE id_materia = %s and id_curso = %s  ORDER BY asistenciaalum.fecha DESC LIMIT 31"
      val = [id, cursos[3]]
      mycursor.execute(sql, val)
      asis = mycursor.fetchall()
      print(asis)
      return render_template('verlista.html', datos=datos, cursos=cursos, listas=asis, filtro=filtro, id=id)
    if filtro == 3:
      sql = "SELECT DISTINCT fecha FROM asistenciaalum WHERE id_materia = %s and id_curso = %s  ORDER BY asistenciaalum.fecha DESC"
      val = [id, cursos[3]]
      mycursor.execute(sql, val)
      asis = mycursor.fetchall()
      print(asis)
      return render_template('verlista.html', datos=datos, cursos=cursos, listas=asis, filtro=filtro, id=id)
  return render_template('verlista.html', datos=datos, cursos = cursos, listas = asis, filtro = 1, id = id)

@app.route('/verificarlista/<string:id>', methods = ['POST', 'GET'])
def verificarlista(id):
  print(id)
  #idmateria
  global user_datos
  #idcurso
  global vector
  print(vector)
  print(user_datos)
  if not user_datos or not vector:
    return redirect(url_for('miscursos'))
  datos = session['username']
  mycursor = mydb.cursor()
  # Saca los datos de la asistencia
  sql = "SELECT id_asisalum, asistenciaalum.id_alumno, asistio, nmb_a, ape_a FROM asistenciaalum, alumnos WHERE id_materia = %s and asistenciaalum.id_curso = %s and fecha = %s " \
        "and asistenciaalum.id_alumno = alumnos.id_alumno"
  val = [user_datos, vector, id]
  mycursor.execute(sql, val)
  asis = mycursor.fetchall()
  print(asis)
  # Ordena alfabetaicamente los alumnos del vector...
  asis = sorted(asis, key=lambda asis: asis[4])
  return render_template('verificarlista.html', datos=datos, listas = asis, id = id, vector = user_datos)

@app.route('/modificarlista/<string:id>', methods = ['POST', 'GET'])
def modlista(id):
  print(id)
  # idmateria
  global user_datos
  # idcurso
  global vector
  if not user_datos or not vector:
    return redirect(url_for('miscursos'))
  #alertas
  global band
  if band == 6:
    flash("","si")
    band = 0
  datos = session['username']
  mycursor = mydb.cursor()
  # Saca los datos de la asistencia
  sql = "SELECT id_asisalum, asistenciaalum.id_alumno, asistio, nmb_a, ape_a FROM asistenciaalum, alumnos WHERE id_materia = %s and asistenciaalum.id_curso = %s and fecha = %s " \
        "and asistenciaalum.id_alumno = alumnos.id_alumno"
  val = [user_datos, vector, id]
  mycursor.execute(sql, val)
  asis = mycursor.fetchall()
  # Ordena alfabetaicamente los alumnos del vector...
  asis = sorted(asis, key=lambda asis: asis[4])
  print(asis)
  alumnos = []
  if request.method == 'POST':
    # Saca los ids de los alumnos
    idalum = request.form.getlist(('idalum'))
    print(idalum)
    # Saca los alumnos seleccionados por el docente
    for x in range(0, len(idalum)):
      aux = int(idalum[x])
      print(aux)
      mycursor = mydb.cursor()
      sql = "SELECT id_alumno,nmb_a, ape_a FROM alumnos WHERE id_alumno = %s"
      val = [aux]
      mycursor.execute(sql, val)
      aux2 = mycursor.fetchall()
      aux2 = aux2[0]
      # print(aux2)
      alumnos.append(aux2)
    print(alumnos)
    # Ordena alfabetaicamente los alumnos del vector...
    alumnos = sorted(alumnos, key=lambda alumnos: alumnos[2])
    global vector2
    vector2 = alumnos
    return redirect(url_for('carmodlista', id=id))
  return render_template('modlista.html', datos=datos, listas=asis, id=id, vector = user_datos)

@app.route('/cargarmodificacion/<string:id>', methods = ['POST', 'GET'])
def carmodlista(id):
  datos = session['username']
  print(id)
  # idmateria
  global user_datos
  # idcurso
  global vector
  #alumnos
  global vector2
  print(vector2)
  if not user_datos or not vector or not vector2:
    return redirect(url_for('miscursos'))
  mycursor = mydb.cursor()
  # Estados posibles
  lista = ["P", "A"]
  # Comprobar que cargo bien
  create = True
  #alertas
  global band
  if request.method == "POST":
    for x in range(0, len(vector2)):
      aux = request.form.getlist(('{a}'.format(a=x + 1)))
      if not aux:
        create = False
    if create == True:
      for x in range(0, len(vector2)):
        aux = request.form.getlist(('{a}'.format(a=x + 1)))
        aux2 = vector2[x]
        # Actualiza por alumno si estuvo presente o ausente
        sql = "UPDATE asistenciaalum SET asistio = %s WHERE id_alumno = %s and id_materia = %s and id_curso = %s and fecha = %s"
        val = (aux[0], aux2[0], user_datos, vector, id)
        mycursor.execute(sql, val)
        mydb.commit()
        mydb.commit()
      band = 6
      return redirect(url_for('modlista', id=id))
    else:
      flash("", "ec")
  return render_template('modlista2.html', datos=datos, alumnos = vector2, listas = lista, id = id, vector = id)

@app.route('/enviarplanilla', methods = ['POST', 'GET'])
def enviarplanilla():
  datos = session['username']
  #Si el profesor mando planillas
  mycursor = mydb.cursor()
  sql = "SELECT estado, fecha_m, fecha_r, des_p, cp FROM planillas WHERE id_profesor = %s LIMIT 45"
  val = [datos[0]]
  mycursor.execute(sql, val)
  planillas = mycursor.fetchall()
  print(planillas)
  if planillas:
    print("ya tiene")
    cp = planillas[len(planillas) - 1]
    cp = int(cp[4]) + 1
  else:
    cp = 1
  if request.method == "POST":
    planilla = request.files["planilla"]
    desc = request.form.get("desc")
    if "planilla" not in request.files:
      print("No envio nada")
      pass
    elif planilla.filename == "":
      print("No mando nada")
      pass
    elif planilla and archpermi2(planilla.filename):
      filename = planilla.filename.split('.')
      ext = filename[len(filename) - 1]
      print(ext)
      print(filename)
      filename = "planillapendiente_" + str(datos[0]) + str(cp) + "." + ext
      print(filename)
      print(desc)
      planilla.save(os.path.join(app.config["folder"], filename))
      flash("", "")  # Ya guarda el archivo
      estado = 0
      inf = datetime.now()
      # Extraemos la fecha
      fecha = datetime.strftime(inf, '%Y/%m/%d')
      mycursor = mydb.cursor()
      if planillas:
        print("ya tiene")
        mycursor.execute(
          'INSERT INTO planillas (filename, estado, id_profesor, fecha_m, des_p, cp) VALUES (%s, %s, %s, %s, %s, %s)',
          (filename, estado, datos[0], fecha, desc, cp))
        mydb.commit()
      else:
        print("no tiene")
        mycursor.execute(
          'INSERT INTO planillas (filename, estado, id_profesor, fecha_m, des_p, cp) VALUES (%s, %s, %s, %s, %s, %s)',
          (filename, estado, datos[0], fecha, desc, 1))
        mydb.commit()
      print("mandado")
      return redirect(url_for('enviarplanilla'))
    else:
      print("archivo no permitido")
  return render_template('enviarplanillas.html', datos=datos, planillas = planillas)

@app.route('/controlplanillas', methods = ['POST', 'GET'])
def controldeplanilla():
  datos = session['username']
  #Planillas pendientes
  mycursor = mydb.cursor()
  sql = "SELECT * FROM planillas WHERE estado = %s LIMIT 45"
  val = [0]
  mycursor.execute(sql, val)
  pp = mycursor.fetchall()
  print(pp)
  # Planillas aprobadas
  mycursor = mydb.cursor()
  sql = "SELECT * FROM planillas WHERE estado = %s LIMIT 45"
  val = [1]
  mycursor.execute(sql, val)
  pa = mycursor.fetchall()
  # Planillas desaprobadas
  mycursor = mydb.cursor()
  sql = "SELECT * FROM planillas WHERE estado = %s LIMIT 45"
  val = [2]
  mycursor.execute(sql, val)
  pd = mycursor.fetchall()
  if request.method == "POST":
    filtro = int(request.form.get("idfiltro"))
    print(filtro)
    if filtro == 1:
      return redirect(url_for('controldeplanilla'))
    else:
      return render_template('controlplanillas.html', datos=datos, planillas_a=pa, planillas=pp, planillas_d=pd,
                             filtro=filtro)
  return render_template('controlplanillas.html', datos=datos, planillas_a = pa, planillas = pp, planillas_d = pd, filtro = 1)

@app.route('/descargarplanilla/<int:id>')
def descargarplanilla(id):
  datos = session['username']
  print(id)
  ruta = pathlib.Path('./media/')
  mycursor = mydb.cursor()
  sql = "SELECT id_planillas, filename FROM planillas WHERE id_planillas = %s"
  val = [id]
  mycursor.execute(sql, val)
  p = mycursor.fetchall()
  planilla = p[0]
  print(planilla)
  filename = planilla[1]
  archivo = ruta / filename  # Si existe un docx con su nombre
  print(archivo)
  if archivo.exists():
    print("El arhivo existe")
    return send_file(archivo, as_attachment=True)
  else:
    print("No existe")
  return redirect(url_for('controldeplanilla'))

@app.route('/aprobarplanilla/<int:id>')
def aprobarplanilla(id):
  datos = session['username']
  inf = datetime.now()
  # Extraemos la fecha
  fecha = datetime.strftime(inf, '%Y/%m/%d')
  mycursor = mydb.cursor()
  sql = "UPDATE planillas SET estado = %s, fecha_r = %s WHERE id_planillas = %s"
  val = (1,fecha, id)
  mycursor.execute(sql, val)
  mydb.commit()
  print("aprobado")
  return redirect(url_for('controldeplanilla'))

@app.route('/desaprobarplanilla/<int:id>')
def desaprobarplanilla(id):
  datos = session['username']
  inf = datetime.now()
  # Extraemos la fecha
  fecha = datetime.strftime(inf, '%Y/%m/%d')
  mycursor = mydb.cursor()
  sql = "UPDATE planillas SET estado = %s, fecha_r = %s WHERE id_planillas = %s"
  val = (2,fecha, id)
  mycursor.execute(sql, val)
  mydb.commit()
  print("desaprobado")
  return redirect(url_for('controldeplanilla'))

@app.route('/generarcuotas', methods = ['POST', 'GET'])
def generarcuotas():
  #Alertas
  global band
  if band == 2:
    flash("","nod")
    band = 0
  if band == 3:
    flash("","ya")
    band = 0
  if band == 4:
    flash("", "ya2")
    band = 0
  if band == 1:
    flash("","ok")
    band = 0
  if band == 7:
    flash("", "del")
    band = 0
  datos = session['username']
  #El mes
  mes = datetime.today().month
  print(mes)
  inf = datetime.now()
  # Extraemos la fecha
  fecha = datetime.strftime(inf, '%Y/%m/%d')
  #Cuotas existentes
  mycursor = mydb.cursor()
  sql = "SELECT DISTINCT fecha, id_tipoc, mes, des_c, monto FROM cuotas ORDER BY cuotas.id_cuota DESC"
  val = []
  mycursor.execute(sql, val)
  cuotas = mycursor.fetchall()
  print(cuotas)
  if request.method == "POST":
    idcuota = int(request.form.get("idcuota"))
    monto = request.form.get("monto")
    print(monto)
    if idcuota == 2:
      desc = request.form.get("desc")
      # Comprueba que no halla generado una cuota de instituto aun
      sql = "SELECT DISTINCT fecha, id_tipoc, mes, des_c, monto FROM cuotas WHERE mes = %s and id_tipoc = %s and des_c = %s"
      val = [mes, idcuota,desc]
      mycursor.execute(sql, val)
      compc = mycursor.fetchall()
      if not desc:
        band = 2
        return redirect(url_for('generarcuotas'))
      if compc:
        band = 4
        return redirect(url_for('generarcuotas'))
    else:
      desc = ""
      # Comprueba que no halla generado una cuota de instituto aun
      sql = "SELECT DISTINCT fecha, id_tipoc, mes, des_c, monto FROM cuotas WHERE mes = %s and id_tipoc = %s"
      val = [mes, idcuota]
      mycursor.execute(sql, val)
      compc = mycursor.fetchall()
      if compc:
        band = 3
        print("ya generaste una cuota de instituto este mes")
        return redirect(url_for('generarcuotas'))
    print(idcuota, desc)
    #Quitamos los cursos existentes
    mycursor = mydb.cursor()
    sql = "SELECT id_curso, id_enfasis FROM cursos"
    val = []
    mycursor.execute(sql, val)
    cursos = mycursor.fetchall()
    for x in range(0, len(cursos)):
      aux = cursos[x]
      print(aux)
      mycursor = mydb.cursor()
      sql = "SELECT DISTINCT id_alumno, id_curso FROM matxalum WHERE id_curso = %s"
      val = [aux[0]]
      mycursor.execute(sql, val)
      alumnos = mycursor.fetchall()
      print(alumnos)
      if alumnos:
        for i in range(0, len(alumnos)):
          alumno = alumnos[i]
          print(alumno)
          mycursor.execute(
            'INSERT INTO cuotas (estado, fecha, id_tipoc, id_alumno, mes, des_c, monto) VALUES (%s, %s, %s, %s, %s, %s, %s)',
            (0, fecha, idcuota, alumno[0], mes, desc, monto))
          mydb.commit()
    band = 1
    return redirect(url_for('generarcuotas'))
  return render_template('generarcuotas.html', datos=datos, cuotas = cuotas)

@app.route('/vercuotasad/<int:id>', methods = ['POST', 'GET'])
def vercuotasad(id):
  datos = session['username']
  # Datos del curso
  mycursor = mydb.cursor()
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  print(cursos)
  #Saca los alumnos
  pendiente = []
  sql = "SELECT DISTINCT matxalum.id_alumno, nmb_a, ape_a FROM matxalum, alumnos WHERE matxalum.id_curso = %s and matxalum.id_alumno = alumnos.id_alumno"
  val = [id]
  mycursor.execute(sql, val)
  alum_t = mycursor.fetchall()
  alum_t = sorted(alum_t, key=lambda alum_t: alum_t[2])
  print(alum_t)
  #Cuotas pendientes de cada alumno
  for x in range(0, len(alum_t)):
    alumno = alum_t[x]
    sql = "SELECT id_cuota, mes, id_tipoc FROM cuotas WHERE id_alumno = %s and estado = %s"
    val = [alumno[0], 0]
    mycursor.execute(sql, val)
    cuotas = mycursor.fetchall()
    if cuotas:
      print(cuotas)
      pendiente.append(len(cuotas))
    else:
      p = 0
      pendiente.append(p)
  print(pendiente)
  return render_template('vercuotas.html', datos=datos, cursos= cursos[0], alumnos = alum_t, pendientes = pendiente)

#Eliminar cuotas
@app.route('/eliminarcuota/<string:id>', methods = ['POST', 'GET'])
def eliminarcuotas(id):
  global band
  print(id)
  x = id.split("I")
  print(x)
  datos = session['username']
  mycursor = mydb.cursor()
  sql = "SELECT * FROM cuotas WHERE fecha = %s and des_c = %s LIMIT 1"
  val = [x[0], x[1]]
  mycursor.execute(sql, val)
  cuota = mycursor.fetchall()
  cuota = cuota[0]
  print(cuota)
  if request.method == "POST":
    print("eliminar")
    mycursor = mydb.cursor()
    sql = "DELETE FROM cuotas WHERE fecha = %s and des_c = %s and id_tipoc = %s"
    val = [x[0], x[1], cuota[3]]
    mycursor.execute(sql, val)
    mydb.commit()
    band = 7
    return redirect(url_for('generarcuotas'))
  return render_template('eliminarcuota.html', datos=datos, cuota = cuota)

#Ver cuotas de cada alumno
@app.route('/vercuotasdelalumno/<int:id>', methods = ['POST', 'GET'])
def vercuotasdelalumno(id):
  datos = session['username']
  #alertas
  global band
  if band == 1:
    flash("ok","ok")
    band = 0
  #Datos del alumno
  mycursor = mydb.cursor()
  sql = "SELECT id_alumno, id_curso, ape_a, nmb_a FROM alumnos WHERE id_alumno = %s"
  val = [id]
  mycursor.execute(sql, val)
  alumno = mycursor.fetchall()
  alumno = alumno[0]
  #Cuotas del alumno
  sql = "SELECT DISTINCT id_cuota,fecha, id_tipoc, mes, des_c, monto FROM cuotas WHERE id_alumno = %s and estado = %s ORDER BY cuotas.id_cuota DESC"
  val = [id, 0]
  mycursor.execute(sql, val)
  cuotas = mycursor.fetchall()
  print(cuotas)
  total = 0
  for x in range(0, len(cuotas)):
    aux = cuotas[x]
    total = total + float(aux[5])
    print(total)
  if request.method == "POST":
    #Cuotas pagadas
    sql = "SELECT DISTINCT id_cuota,fecha, id_tipoc, mes, des_c, monto FROM cuotas WHERE id_alumno = %s and estado = %s ORDER BY cuotas.id_cuota DESC"
    val = [id, 1]
    mycursor.execute(sql, val)
    cuotas = mycursor.fetchall()
    total = 0
    for x in range(0, len(cuotas)):
      aux = cuotas[x]
      total = total + float(aux[5])
      print(total)
    return render_template('vercuotasdelalumno.html', datos=datos, data=alumno, cuotas=cuotas, filtro = 2, total = total)
  return render_template('vercuotasdelalumno.html', datos=datos, data = alumno, cuotas = cuotas, filtro = 1, total = total)
#Mis cuotas pestaña de alumno
@app.route('/miscuotas', methods = ['POST', 'GET'])
def miscuotasal():
  datos = session['username']
  mycursor = mydb.cursor()
  # Cuotas del alumno
  sql = "SELECT DISTINCT id_cuota,fecha, id_tipoc, mes, des_c, monto FROM cuotas WHERE id_alumno = %s and estado = %s ORDER BY cuotas.id_cuota DESC"
  val = [datos[0], 0]
  mycursor.execute(sql, val)
  cuotas = mycursor.fetchall()
  print(cuotas)
  total = 0
  for x in range(0, len(cuotas)):
    aux = cuotas[x]
    total = total + float(aux[5])
    print(total)
  if request.method == "POST":
    filtro = int(request.form.get("idfiltro"))
    if filtro == 1:
      return redirect(url_for('miscuotasal'))
    if filtro == 2:
      sql = "SELECT DISTINCT id_cuota,fecha, id_tipoc, mes, des_c, monto FROM cuotas WHERE id_alumno = %s and estado = %s ORDER BY cuotas.id_cuota DESC"
      val = [datos[0], 1]
      mycursor.execute(sql, val)
      cuotas = mycursor.fetchall()
      print(cuotas)
      total = 0
      for x in range(0, len(cuotas)):
        aux = cuotas[x]
        total = total + float(aux[5])
        print(total)
      return render_template('miscuotas.html', datos=datos, cuotas=cuotas, filtro = filtro, total = total)
  return render_template('miscuotas.html', datos=datos, cuotas = cuotas, filtro = 1, total = total)

@app.route('/cuotapagada/<int:id>', methods = ['POST', 'GET'])
def marcarpagado(id):
  datos = session['username']
  print(id)
  global band
  mycursor = mydb.cursor()
  sql = "SELECT id_cuota, id_alumno FROM cuotas WHERE id_cuota = %s"
  val = [id]
  mycursor.execute(sql, val)
  cuotas = mycursor.fetchall()
  cuotas = cuotas[0]
  print(cuotas)
  #Marca como pagado
  mycursor = mydb.cursor()
  sql = "UPDATE cuotas SET estado = %s WHERE id_cuota = %s"
  val = (1,id)
  mycursor.execute(sql, val)
  mydb.commit()
  print("pagaado")
  band = 1
  return redirect(url_for('vercuotasdelalumno', id=cuotas[1]))
def createpassword(password):
  return generate_password_hash(password)
def crearclavet(): #Una clave random para el trabajo.
  length = 5
  letters = string.ascii_letters + string.digits
  clave = ''.join(random.choice(letters) for i in range(length))
  return clave
def conv_b(foto):
  with open(foto, 'rb') as f:
    blob= f.read()
  return blob
def guar_f(id_u,foto):
  with open('media/foto_{}.png'.format(id_u), 'wb') as f:
    f.write(foto)
#Para documentos de inscripcion de alumnos o docentes
def archpermi(filename):
  filename = filename.split('.')
  if filename[len(filename) - 1] in ext_p:
    return True
  return False
#Para planillas
def archpermi2(filename):
  filename = filename.split('.')
  if filename[len(filename) - 1] in ext_c:
      return True
  return False
def guardararchivo():

  return
if __name__=='__main__':
    app.run(debug = True, port= 8000)
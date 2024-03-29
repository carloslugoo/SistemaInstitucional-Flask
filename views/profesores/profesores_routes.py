#Blueprint libreria para vistas basandome en modulos
from flask import Blueprint
#Flask imports
from flask import url_for, redirect, render_template, request, session, flash, current_app, send_file
#SQL
import mysql.connector
#Date time
from datetime import datetime
#Exportar a pdf
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table, TableStyle, SimpleDocTemplate
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER
from reportlab.lib import colors
#System
import os
import pathlib
#Reporte en excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
#Import del formulario
import userform
#Para creacion random
import string
import random

profesores_views = Blueprint('profesores', __name__, template_folder='templates')

#Conexion a SQL
mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="",
  database="proyecto"
)

#views
@profesores_views.before_request
def before_request():
    #Si no estas autenticado, fuera
    if 'username' not in session:
        return redirect(url_for('auth.login'))
    #Si no es tu url, fuera
    if request.endpoint in ['verproceso','vermaterias ', 'bienvenidoadmin', 'inscribirdir', 'sacarmat',
                            'inscurmat', 'inscribirprof', 'inscmatxprof',
                            'listadocursos', 'controldeplanilla', 'crearsemana', 'inscribirprof',
                            'inscribirdir', 'asignarpof', 'generarcuotas', 'verlog', 'miscuotasal', 'vermaterias'
                            ]:
      return redirect(url_for('docentes.miscursos'))
    # Acceder a la variable global desde la aplicación
    band = current_app.config['band']
    if request.endpoint in ['profesores.modificarproceso2']:
      if band == 1:
        band = current_app.config['band']
        flash("si_c", "si_c")
      if band == 2:
        band = current_app.config['band']
        flash("ocurrio", "ocurrio")
    if request.endpoint in ['profesores.modlista']:
      if band == 6:
        flash("","si")
        current_app.config['band'] = 0
    if request.endpoint in ['profesores.proceso']:
      if band == 1:
        flash("car","car")
        current_app.config['band'] = 0
    if band == 7:
        flash("","noclases")
        current_app.config['band'] = 0
    if band == 6:
        flash("","si")
        current_app.config['band'] = 0
    if band == 5:
        flash("", "ya")
        current_app.config['band'] = 0
    if band == 1:
        current_app.config['band'] = 0
        flash("", "si_c")
    if band == 3:
        current_app.config['band'] = 0
        flash("", "eliminado")

@profesores_views.route('/miscursos') #Listar las materias que tiene por curso el prof
def miscursos():
  datos = session['username']
  print(datos)
  mycursor = mydb.cursor()
  sql = "SELECT id_mxp, matxpro.id_materia, matxpro.id_profesor, matxpro.id_curso, des_m, des_c, des_e, fecha" \
        " FROM matxpro, materias, cursos, enfasis WHERE id_profesor = %s and matxpro.id_materia = materias.id_materia " \
        "and matxpro.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis "
  val = [datos[0]]
  mycursor.execute(sql, val)
  mat = mycursor.fetchall()
  print(mat)
  return render_template('miscursos.html', datos = datos, materias = mat)

@profesores_views.route('/veralum/<int:id>', methods = ['GET', 'POST']) #Ver los alumnos matriculados con sus cal y puntajes
def misalumnos(id):
  datos = session['username']
  mycursor = mydb.cursor()
  #Saca los datos del alumno, los nombres y puntajes obtenidos en la materia
  sql = "SELECT id_mxa, matxalum.id_alumno, matxalum.id_materia, cal, matxalum.id_curso, pun_ac, des_m, des_c, des_e , nmb_a, ape_a, pun_ac2, cal2" \
        " FROM matxalum, materias, cursos, enfasis, alumnos WHERE id_profesor = %s and matxalum.id_materia = %s and matxalum.id_materia = materias.id_materia " \
        "and matxalum.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis and matxalum.id_alumno = alumnos.id_alumno"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  mat = mycursor.fetchall()
  #Saco los datos de la materia del vector...
  for x in range(0, 1):
    aux = mat[x]
    data = aux
    #print(data)
  #Ordena alfabetaicamente los alumnos del vector...
  mat = sorted(mat, key=lambda mat: mat[10])
  #print(mat)
  #Saca el total de puntos en la primera etapa
  sql = "SELECT id_trabajo, pun_t FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s and etapa = %s"
  val = [datos[0], id, data[4], 1]
  mycursor.execute(sql, val)
  trabajosp = mycursor.fetchall()
  total_p = 0
  #print(trabajosp)
  for x in range(0, len(trabajosp)):
    aux = trabajosp[x]
    aux = aux[1]
    total_p += aux
  #print(total_p)
  # Saca el total de puntos en la segunda etapa
  sql = "SELECT id_trabajo, pun_t FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s and etapa = %s"
  val = [datos[0], id, data[4], 2]
  mycursor.execute(sql, val)
  trabajoss = mycursor.fetchall()
  total_s = 0
  #print(trabajoss)
  for x in range(0, len(trabajoss)):
    aux = trabajoss[x]
    aux = aux[1]
    total_s += aux
  #print(total_s)
  #Alumnos criticos y buenos en primera etapa
  criticos_p = []
  bien_p = []
  estado = (70 * total_p) / 100  # Calcula el 70% del total
  #print(estado)
  for x in range(0, len(mat)):
    aux = mat[x]
    aux2 = aux[5] #Puntaje primera etapa
    if aux2 >= estado:
      bien_p.append(aux)
    else:
      criticos_p.append(aux)
  #print(criticos_p)
  #print(bien_p)
  # Alumnos criticos y buenos en segunda etapa
  criticos_s = []
  bien_s = []
  estado2 = (70 * total_s) / 100  # Calcula el 70% del total
  #print(estado2)
  for x in range(0, len(mat)):
    aux3 = mat[x]
    aux4 = aux3[11]  # Puntaje primera etapa
    if aux4 >= estado2:
      bien_s.append(aux3)
    else:
      criticos_s.append(aux3)
  #print(criticos_s)
  #print(bien_s)
  if request.method == 'POST':
    filtro = int(request.form.get(('idfiltro')))
    #print(filtro)
    if filtro == 1:
      return render_template('misalumnos.html', datos=datos, materias=mat, data=data, filtro=1, id = id, criticos_p = criticos_p, bien_p = bien_p,
                             criticos_s = criticos_s, bien_s = bien_s, total_s = total_s, total_p = total_p)
    if filtro == 2:
      return render_template('misalumnos.html', datos=datos, materias=mat, data=data, filtro=2, id = id, criticos_p = criticos_p, bien_p = bien_p,
                             criticos_s = criticos_s, bien_s = bien_s, total_s = total_s, total_p = total_p )
  return render_template('misalumnos.html', datos = datos, materias = mat, data = data, filtro = 0, id = id, criticos_p = criticos_p, bien_p = bien_p,
                             criticos_s = criticos_s, bien_s = bien_s, total_s = total_s, total_p = total_p)

@profesores_views.route('/exportaralumnos/<int:id>', methods = ['GET', 'POST'])
def exportaralumnos(id):
  datos = session['username']
  #print(datos)
  mycursor = mydb.cursor()
  # Saca los datos del alumno, los nombres y puntajes obtenidos en la materia
  sql = "SELECT id_mxa, matxalum.id_alumno, matxalum.id_materia, cal, matxalum.id_curso, pun_ac, des_m, des_c, des_e , nmb_a, ape_a, pun_ac2, cal2" \
        " FROM matxalum, materias, cursos, enfasis, alumnos WHERE id_profesor = %s and matxalum.id_materia = %s and matxalum.id_materia = materias.id_materia " \
        "and matxalum.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis and matxalum.id_alumno = alumnos.id_alumno"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  mat = mycursor.fetchall()
  print(mat)
  # Saco los datos de la materia del vector...
  for x in range(0, 1):
    aux = mat[x]
    data = aux
    print(data)
  #Ordena alfabetaicamente los alumnos del vector...
  mat = sorted(mat, key=lambda mat: mat[10])
  # Saca el total de puntos en la primera etapa
  sql = "SELECT id_trabajo, pun_t FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s and etapa = %s"
  val = [datos[0], id, data[4], 1]
  mycursor.execute(sql, val)
  trabajosp = mycursor.fetchall()
  total_p = 0
  # print(trabajosp)
  for x in range(0, len(trabajosp)):
    aux = trabajosp[x]
    aux = aux[1]
    total_p += aux
  # print(total_p)
  # Saca el total de puntos en la segunda etapa
  sql = "SELECT id_trabajo, pun_t FROM trabajos WHERE id_profesor = %s and id_materia = %s and id_curso = %s and etapa = %s"
  val = [datos[0], id, data[4], 2]
  mycursor.execute(sql, val)
  trabajoss = mycursor.fetchall()
  total_s = 0
  # print(trabajoss)
  for x in range(0, len(trabajoss)):
    aux = trabajoss[x]
    aux = aux[1]
    total_s += aux

  if request.method == 'POST':
    filtro = int(request.form.get(('idfiltro')))
    print(filtro)
    #Dando formato al pdf
    pdf = SimpleDocTemplate(
      "./static/resources/pdfs_creados/{a}_alumnos.pdf".format(a = datos[2]),
      pagesize=A4,
      rightMargin=inch,
      leftMargin=inch,
      topMargin=inch,
      bottomMargin=inch / 2
    )
    Story = []
    # Estilos
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
    # Cabecera
    text = ''' <strong><font size=14>Alumnos matriculados en {a}.</font></strong>
    '''.format(a=data[6])
    text2 = '''
        <strong><font size=10>Docente: {a}, {b}</font></strong>
    '''.format(a=datos[1], b= datos[2])
    text3 = '''
        <strong><font size=10>Curso: {a}, {b}</font></strong>
    '''.format(a= data[7], b= data[8])
    # Adjuntamos los titulos declarados mas arriba, osea las cabeceras
    Story.append(Paragraph(text2))
    Story.append(Paragraph(text3))
    Story.append(Paragraph(text, styles['Center']))
    Story.append(Spacer(1, 20))
    # Declaramos los datos de la cabecera de la tabla, los titulos y tambien sus estilos
    if filtro == 1: #Primera etapa
      data = [(
        Paragraph('<strong><font size=6>#</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Nombre y Apellido</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Primera Etapa PT:{a}</font></strong>'.format(a = total_p), styles['Center']),
        Paragraph('<strong><font size=6>Calificación</font></strong>', styles['Center'])
      )]
      # Aqui acomplamos los registros o datos a nuestra tabla data, estos seran los datos mostrados de bajo de los headers
      for x in range(0, len(mat)):
        aux = mat[x]
        count = str(x + 1)
        print(count)
        alumno = aux[9] + "," + aux[10]  # nombre y apellido
        pp = aux[5]  # puntaje primera etapa
        if aux[3]:  # Calificacion primera etapa
          cal = aux[3]
        else:
          cal = "Sin calificación"
        data.append((
          Paragraph('<font size=6>%s</font>' % count, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % alumno, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % pp, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % cal, styles['Normal'])
        ))
    if filtro == 2: #Segunda
      data = [(
        Paragraph('<strong><font size=6>#</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Nombre y Apellido</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Segunda Etapa. PT:{a}</font></strong>'.format(a = total_s), styles['Center']),
        Paragraph('<strong><font size=6>Calificación</font></strong>', styles['Center'])
      )]
      # Aqui acomplamos los registros o datos a nuestra tabla data, estos seran los datos mostrados de bajo de los headers
      for x in range(0, len(mat)):
        aux = mat[x]
        count = str(x + 1)
        print(count)
        alumno = aux[9] + "," + aux[10]  # nombre y apellido
        ps = aux[11]  # puntaje segunda etapa
        if aux[12]:
          cal2 = aux[12]
        else:
          cal2 = "Sin calificación"
        data.append((
          Paragraph('<font size=6>%s</font>' % count, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % alumno, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % ps, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % cal2, styles['Normal'])
        ))
    if filtro == 3: #Todas las etapas
      data = [(
        Paragraph('<strong><font size=6>#</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Nombre y Apellido</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Primera Etapa. PT:{a}</font></strong>'.format(a = total_p), styles['Center']),
        Paragraph('<strong><font size=6>Calificación</font></strong>', styles['Center']),
        Paragraph('<strong><font size=6>Segunda Etapa. PT:{a}</font></strong>'.format(a = total_s), styles['Center']),
        Paragraph('<strong><font size=6>Calificación</font></strong>', styles['Center'])
      )]
      # Aqui acomplamos los registros o datos a nuestra tabla data, estos seran los datos mostrados de bajo de los headers
      for x in range(0, len(mat)):
        aux = mat[x]
        count = str(x + 1)
        print(count)
        alumno = aux[9] + "," + aux[10]  # nombre y apellido
        pp = aux[5]  # puntaje primera etapa
        if aux[3]:  # Calificacion primera etapa
          cal = aux[3]
        else:
          cal = "Sin calificación"
        ps = aux[11]  # puntaje segunda etapa
        if aux[12]:
          cal2 = aux[12]
        else:
          cal2 = "Sin calificación"
        data.append((
          Paragraph('<font size=6>%s</font>' % count, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % alumno, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % pp, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % cal, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % ps, styles['Normal']),
          Paragraph('<font size=6>%s</font>' % cal2, styles['Normal'])
        ))
    # Declaramamos que la tabla recibira como dato los datos anteriores y le damos la dimensiones a cada uno de nuestros campos
    table = Table(
      data,
      colWidths=[20, 140, 80, 60, 80, 60]
    )
    table.setStyle(
      TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
      ])
    )
    Story.append(table)
    pdf.build(Story)
    #Comprueba si existe el archivo y lo descarga para el usuario
    ruta = pathlib.Path('./static/resources/pdfs_creados')
    filename = "{a}_alumnos.pdf".format(a=datos[2])
    archivo = ruta / filename  # Si existe un pdf con su nombre
    print(archivo)
    if archivo.exists():
      return send_file(archivo, as_attachment=True)
  return redirect(url_for('profesores.misalumnos', id=id))


@profesores_views.route('/modificarproceso/<int:id>', methods = ['GET', 'POST']) #Modificar trabajos del docente
def modproceso(id):
  datos = session['username']
  mycursor = mydb.cursor()
  print(datos)
  #Obtiene los trabajos relacionados a la materia
  sql = "SELECT id_trabajo, fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and trabajos.id_materia = %s and " \
        "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  #print(trabajos)
  if request.method == 'POST':
    filtro = int(request.form.get(('idfiltro')))
    #print(filtro)
    #Dependiendo del filtro, busca por etapa
    if filtro == 1:
      sql = "SELECT id_trabajo, fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and etapa = %s and trabajos.id_materia = %s and " \
            "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
      val = [datos[0], filtro, id]
      mycursor.execute(sql, val)
      trabajos = mycursor.fetchall()
    if filtro == 2:
      sql = "SELECT id_trabajo, fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and etapa = %s and trabajos.id_materia = %s and " \
            "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
      val = [datos[0], filtro, id]
      mycursor.execute(sql, val)
      trabajos = mycursor.fetchall()
    if filtro == 3:
      return redirect(url_for('profesores.modproceso', id = id))
    #print(trabajos)
    return render_template('modproceso.html', datos=datos, trabajos=trabajos, filtro = filtro, id= id)
  return render_template('modproceso.html', datos=datos, trabajos = trabajos, filtro = 0, id = id)

#Generar planilla de procesos o trabajos
@profesores_views.route('/exportarexcelproceso/<int:id>', methods = ['GET', 'POST'])
def generarexcel(id):
  #print(id)
  datos = session['username']
  celdas = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'AA', 'AB', 'AC', 'AE', 'AF'
          , 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC',
            'BD', 'BE']
  if request.method == 'POST':
    mycursor = mydb.cursor()
    filtro = int(request.form.get(('idfiltro')))
    #print(filtro)
    if filtro < 3:
      #Trabajos de la materia
      sql = "SELECT id_trabajo, fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and etapa = %s and trabajos.id_materia = %s and " \
            "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
      val = [datos[0], filtro, id]
      mycursor.execute(sql, val)
      trabajos = mycursor.fetchall()
      a = trabajos[0]
      if not a:
        return redirect(url_for('profesores.modproceso', id=id))
      print(trabajos)
      inf = datetime.now()
      # Extraemos la fecha
      fecha = datetime.strftime(inf, '%Y/%m/%d')
      # Cargamos el archivo
      wb = load_workbook('./static/resources/planillas_templates/templateprocesos.xlsx')
      ws = wb["Hoja1"]
      ws['C4'] = datos[1] + ", " + datos[2]
      ws['V4'] = a[5]
      ws['AP4'] = a[8]
      ws['C5'] = filtro
      if a[7] == "Primer Curso":
        ws['G5'] = 1
      if a[7] == "Segundo Curso":
        ws['G5'] = 2
      if a[7] == "Tercer Curso":
        ws['G5'] = 3
      ws['T5'] = "T.N"
      # Insertamos los indicadores
      cont = 0
      # Para saber que celdas combinar
      ci = ""
      cic = 0
      cf = ""
      cfc = 0
      #Combinar celdas
      sheet = wb.active
      for i in range(0, len(trabajos)):
        inicio = 12
        aux = trabajos[i]
        # Saca los indicadores
        sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
        val = [aux[0]]
        mycursor.execute(sql, val)
        indicadores = mycursor.fetchall()
        if i == 0:
          # Saca los alumnos relacionados con el trabajo
          sql = "SELECT nmb_a, ape_a, id_txa, traxalum.id_alumno, pun_l, fec_t, id_trabajo FROM traxalum, alumnos WHERE id_trabajo = %s and traxalum.id_alumno = alumnos.id_alumno"
          val = [aux[0]]
          mycursor.execute(sql, val)
          datas = mycursor.fetchall()
          datas = sorted(datas, key=lambda datas: datas[1])  # Ordena por orden alfabetic
        for x in range(0, len(indicadores)):
          aux2 = indicadores[x]
          if x == 0:
            ci = celdas[cont]
            cic = cont
          ws['{a}11'.format(a=celdas[cont])] = aux2[1]
          ws['{a}41'.format(a=celdas[cont])] = aux2[1]
          #Puntaje de cada indicador
          ws['{a}12'.format(a=celdas[cont])] = aux2[2]
          ws['{a}42'.format(a=celdas[cont])] = aux2[2]
          if x == len(indicadores) - 1:
            cf = celdas[cont + 1]
            cfc = cont + 1
            ws['{a}11'.format(a=celdas[cont + 1])] = "TOTAL INDICADORES LOGRADOS"
            ws['{a}41'.format(a=celdas[cont + 1])] = "TOTAL INDICADORES LOGRADOS"
            color1 = PatternFill(start_color='cc9cfc',
                                  end_color='cc9cfc',
                                  fill_type='solid')
            ws['{a}11'.format(a=celdas[cont + 1])].fill = color1
            ws['{a}41'.format(a=celdas[cont + 1])].fill = color1
            cont += 1
          cont += 1
        print(ci, cic, cf, cfc)
        #Fecha del trabajo
        ws['{a}6'.format(a=ci)] = aux[1]
        #Tema o contenido
        ws['{a}9'.format(a=ci)] = aux[2]
        #Total de puntos
        ws['{a}12'.format(a=cf)] = aux[3]
        ws['{a}42'.format(a=cf)] = aux[3]
        a = ws['{a}42'.format(a=cf)]
        b = ws['{a}12'.format(a=cf)]
        a.font = Font(bold=True)
        b.font = Font(bold=True)
        #Une las celdas
        for y in range(6,11):
          sheet.merge_cells('{a}{c}:{b}{c}'.format(a=ci, b= cf, c=y))
        total = 0
        for i in range(0, len(datas)):
          alumno = datas[i]
          # Saca los indicadores del alumno
          sql = "SELECT id_ixa, pun_l FROM indxalum WHERE id_trabajo = %s and id_alumno = %s"
          val = [aux[0], alumno[3]]
          mycursor.execute(sql, val)
          indicadores = mycursor.fetchall()
          print(indicadores)
          fc = 0
          for x in range(cic, cfc + 1):
            #print(x)
            indicador = indicadores[fc]
            pl = indicador[1]
            print(indicador)
            if not x == cfc - 1:
              fc += 1
              #print(fc)
            if inicio != 32 and x != cfc:
              ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = pl
              #print('{b}{a}'.format(a=str(inicio + 1), b=celdas[x]))
              total += pl
            if inicio != 32 and x == cfc:
              ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = total
              pn = ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])]
              pn.font = Font(bold=True)
            #Salto a laa planilla de abajo
            if inicio == 32 and x != cfc:
              inicio = 42
              ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = pl
              inicio += 1
            if inicio == 32 and x == cfc:
              ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])] = total
              pn = ws['{b}{a}'.format(a=str(inicio + 1), b=celdas[x])]
              pn.font = Font(bold=True)

          til = ws['BP{a}'.format(a=str(inicio + 1))]
          if not til.value:
            ws['BP{a}'.format(a=str(inicio + 1))] = 0
            til = ws['BP{a}'.format(a=str(inicio + 1))]
          print(til.value)
          ws['BP{a}'.format(a=str(inicio + 1))] = til.value + total
          print("endfor")
          total = 0
          inicio += 1
      # Insertamos alumnos
      inicio = 12
      for x in range(0, len(datas)):
        alumno = datas[x]
        print(alumno)
        if inicio != 32:
          ws['B{a}'.format(a=str(inicio + 1))] = alumno[1] + ", " + alumno[0]
          inicio += 1
        else:
          inicio = 42
          ws['B{a}'.format(a=str(inicio + 1))] = alumnos[x]
          inicio += 1
      # Guarda el archivo
      ws = wb["Hoja1"]
      ws.title ="Planilla de Proceso"
      wb.save('./static/resources/planillas_creadas/Planilla de Proceso.xlsx')
      ruta = pathlib.Path('./static/resources/planillas_creadas/')
      filename = "Planilla de Proceso.xlsx"
      archivo = ruta / filename  # Si existe un docx con su nombre
      print(archivo)
      if archivo.exists():
        print("El arhivo existe")
        return send_file(archivo, as_attachment=True)
      else:
        print("No existe")
    if filtro >= 3: #Calificaciones primera etapa o segunda
      # Trabajos de la materia
      sql = "SELECT id_trabajo, materias.id_materia fec_t, des_t, pun_t, id_profesor, materias.des_m, tipo_t, cursos.des_c, des_e, etapa, trabajos.id_curso FROM trabajos, materias, cursos, enfasis WHERE id_profesor = %s and etapa = %s and trabajos.id_materia = %s and " \
            "trabajos.id_curso = cursos.id_curso and trabajos.id_materia = materias.id_materia and cursos.id_enfasis = enfasis.id_enfasis  ORDER BY trabajos.id_trabajo DESC"
      if filtro == 3:
        val = [datos[0], 1, id]
        etapa = "1ra. Etapa"
      else:
        val = [datos[0], 2, id]
        etapa = "2da. Etapa"
      mycursor.execute(sql, val)
      trabajos = mycursor.fetchall()
      a = trabajos[0]
      if not a:
        return redirect(url_for('profesores.modproceso', id=id))
      print(trabajos)
      inf = datetime.now()
      # Extraemos la fecha
      fecha = datetime.strftime(inf, '%Y/%m/%d')
      # Cargamos el archivo
      wb = load_workbook('./static/resources/planillas_templates/templateetapa.xlsx')
      ws = wb["Hoja1"]
      #Cabecera
      ws['C9'] = "Escuela Nacional de Comercio Héroes del Chaco"
      ws['C42'] = "Escuela Nacional de Comercio Héroes del Chaco"
      ws['C10'] = etapa
      ws['C43'] = etapa
      ws['A11'] = "Curso: " + a[7] + " " + a[8]
      ws['A44'] = "Curso: " + a[7] + " " + a[8]
      ws['D11'] = "T.N"
      ws['D44'] = "T.N"
      ws['A12'] = a[5]  # Materia
      ws['E12'] = datos[1] + ", " + datos[2]
      ws['D45'] = datos[1] + ", " + datos[2]
      totalp = 0
      pla = []
      #Sacamos el total de puntos acumulado
      for i in range(0, len(trabajos)):
        aux = trabajos[i]
        pt = int(aux[3])
        totalp = pt + totalp
        # Saca los indicadores
        sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
        val = [aux[0]]
        mycursor.execute(sql, val)
        indicadores = mycursor.fetchall()
        if i == 0:
          # Saca los alumnos relacionados con el trabajo
          sql = "SELECT nmb_a, ape_a, traxalum.id_alumno FROM traxalum, alumnos WHERE id_trabajo = %s and traxalum.id_alumno = alumnos.id_alumno"
          val = [aux[0]]
          mycursor.execute(sql, val)
          datas = mycursor.fetchall()
          datas = sorted(datas, key=lambda datas: datas[1])  # Ordena por orden alfabetico
          print(datas)
      #Puntaje logrado de cada alumno
      for i in range(0, len(datas)):
        alumno = datas[i]
        if filtro == 3:
          sql = "SELECT id_alumno ,pun_ac FROM matxalum WHERE id_alumno = %s and id_materia = %s and id_curso = %s"
        else:
          sql = "SELECT id_alumno ,pun_ac2 FROM matxalum WHERE id_alumno = %s and id_materia = %s and id_curso = %s"
        val = [alumno[2], a[1], a[10]]
        mycursor.execute(sql, val)
        puntaje = mycursor.fetchall()
        puntaje = puntaje[0]
        pla.append(puntaje)
      print(pla)
      #Insertamos en el excel
      # Insertamos alumnos
      print(totalp)
      inicio = 15
      for x in range(0, len(datas)):
        alumno = datas[x]
        puntaje = pla[x]
        aux3 = puntaje[1]
        print(aux3)
        calculo = (aux3 / totalp) * 100
        print(calculo)
        porcentaje1 = (calculo * 0.7)
        print(porcentaje1)
        pt = porcentaje1 + 20 + 10
        print(pt)
        cal = 0
        if pt < 70:
          cal = 1
        elif pt < 78:
          cal = 2
        elif pt < 86:
          cal = 3
        elif pt < 94:
          cal = 4
        elif pt >= 94:
          cal = 5
        print(cal)
        if filtro == 3:
          sql = "UPDATE matxalum SET cal = %s WHERE id_alumno = %s and id_materia = %s and id_curso = %s"
        else:
          sql = "UPDATE matxalum SET cal2 = %s WHERE id_alumno = %s and id_materia = %s and id_curso = %s"
        val = (cal, alumno[2], a[1], a[10])
        mycursor.execute(sql, val)
        mydb.commit()
        if inicio != 36:
          ws['B{a}'.format(a=str(inicio + 1))] = alumno[1] + ", " + alumno[0]
          ws['C{a}'.format(a=str(inicio + 1))] = totalp
          ws['D{a}'.format(a=str(inicio + 1))] = puntaje[1]
          # Puntaje proyecto
          ws['G{a}'.format(a=str(inicio + 1))] = 15
          ws['H{a}'.format(a=str(inicio + 1))] = 15
          # Puntaje institucional
          ws['K{a}'.format(a=str(inicio + 1))] = 10
          ws['L{a}'.format(a=str(inicio + 1))] = 10
          inicio += 1
        else:
          inicio = 48
          ws['B{a}'.format(a=str(inicio + 1))] = alumno[1] + ", " + alumno[0]
          ws['C{a}'.format(a=str(inicio + 1))] = totalp
          ws['D{a}'.format(a=str(inicio + 1))] = puntaje[1]
          #Puntaje proyecto
          ws['G{a}'.format(a=str(inicio + 1))] = 15
          ws['H{a}'.format(a=str(inicio + 1))] = 15
          #Puntaje institucional
          ws['K{a}'.format(a=str(inicio + 1))] = 10
          ws['L{a}'.format(a=str(inicio + 1))] = 10
          inicio += 1
      # Guarda el archivo
      ws = wb["Hoja1"]
      ws.title = "Planilla de Calificaciones"
      wb.save('./static/resources/planillas_creadas/Planilla de Calificaciones.xlsx')
      ruta = pathlib.Path('./static/resources/planillas_creadas/')
      filename = "Planilla de Calificaciones.xlsx"
      archivo = ruta / filename  # Si existe un docx con su nombre
      print(archivo)
      if archivo.exists():
        print("El arhivo existe")
        return send_file(archivo, as_attachment=True)
      else:
        print("No existe")
  return redirect(url_for('profesores.modproceso', id = id))

#Vista de modificar trabajo, puntajes de alumnos
@profesores_views.route('/modificarproceso2/<int:id>', methods = ['GET', 'POST'])
def modificarproceso2(id):
  print(id)
  datos = session['username']
  #Alumnos
  global user_datos
  mycursor = mydb.cursor()
  #Saca los alumnos relacionados con el trabajo
  sql = "SELECT nmb_a, ape_a, id_txa, traxalum.id_alumno, pun_l, fec_t, id_trabajo FROM traxalum, alumnos WHERE id_trabajo = %s and traxalum.id_alumno = alumnos.id_alumno"
  val = [id]
  mycursor.execute(sql, val)
  datas = mycursor.fetchall()
  datas = sorted(datas, key=lambda datas: datas[1]) #Ordena por orden alfabetico
  print(datas)
  #Saca los datos relacionados con el trabajo
  sql = "SELECT des_t, pun_t, tipo_t, id_materia, id_trabajo FROM trabajos WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  print(datas)
  punt_t = trabajos[0]
  print(punt_t)
  alumnos = []
  if request.method == 'POST':
    #Saca los ids de los alumnos
    idalum = request.form.getlist(('idalum'))
    print(idalum)
    #Saca los alumnos seleccionados por el docente
    for x in range(0, len(idalum)):
      aux = int(idalum[x])
      print(aux)
      mycursor = mydb.cursor()
      sql = "SELECT id_alumno,nmb_a, ape_a FROM alumnos WHERE id_alumno = %s"
      val = [aux]
      mycursor.execute(sql, val)
      aux2 = mycursor.fetchall()
      aux2 = aux2[0]
      #print(aux2)
      alumnos.append(aux2)
    #print(alumnos)
    #Guardamos en una global para la funcion de cargar puntaje
    user_datos = alumnos
    return redirect(url_for('profesores.modificarpuntaje', id = id))
  return render_template('modproceso2.html', datos=datos, datas = datas, trabajos = trabajos[0])

#modificar puntaje por alumno, ya seleccionado por el profesor
@profesores_views.route('/modificarpuntaje/<int:id>', methods = ['GET', 'POST']) #Modificar puntaje obtenido por el alumno
def modificarpuntaje(id):
  datos = session['username']
  #Alumnos
  global user_datos
  # Saca el trabajo
  mycursor = mydb.cursor()
  sql = "SELECT id_trabajo, des_t, pun_t, id_materia, id_curso, id_profesor, etapa FROM trabajos WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  trabajos = trabajos[0]
  print(trabajos)
  # Saca los indicadores
  sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  indicadores = mycursor.fetchall()
  print(indicadores)
  #Saca los alumnos
  global user_datos
  alumnos = user_datos
  print(alumnos)
  #Comprobar que este bien cargado
  create = True
  #Para mandar mensajes
  global band
  # Comprueba que la variable global este funcionando correctamente
  if not alumnos:
    band = 2
    return redirect(url_for('profesores.modificarproceso2', id=id))
  if request.method == 'POST':
    puntajes = []
    pl = 0
    #Comprobar que se hayan cargado todos los indicadores
    for x in range(0, len(alumnos)):
      for y in range(1, len(indicadores) + 1):
        aux = request.form.getlist(('{a}{i}'.format(a=x + 1, i=y)))
        if not aux:
          create = False
    #Solo si esta cargado correctamente
    if create == True:
      #Carga de indicador por alumno, actualizancion de puntaje logrado por indicador y puntaje total logrado
      for x in range(0, len(alumnos)):
        aux1 = alumnos[x]
        # print(aux1)
        for y in range(1, len(indicadores) + 1):
          aux2 = indicadores[y - 1]
          print(aux2)
          if y == 1:
            pl = 0
            # print("Alumno {}".format(alumnos[x]))
          aux = request.form.getlist(('{a}{i}'.format(a=x + 1, i=y)))
          # print('{a}{i}'.format(a=x + 1, i=y))
          print(aux)
          p = int(aux[0])
          # print(p)
          pl += int(aux[0])
          sql = "UPDATE indxalum SET pun_l =  %s WHERE id_alumno = %s and id_trabajo = %s and id_indicador = %s"
          val = (p,aux1[0],id, aux2[0])
          mycursor.execute(sql, val)
          mydb.commit()
          if y == len(indicadores):
            puntajes.append(pl)
      print(puntajes)
      #Actualizando el puntaje total logrado por el alumno
      for x in range(0, len(alumnos)):
        aux1 = alumnos[x]
        aux2 = puntajes[x]
        #Sacando primer puntaje obtenido
        sql = "SELECT id_alumno, pun_l FROM traxalum WHERE id_alumno = %s and id_trabajo = %s"
        val = [aux1[0], id]
        mycursor.execute(sql, val)
        pp = mycursor.fetchall()
        pp = pp[0]
        print(pp)
        # Actualizando el trabajo al alumno con su puntaje logrado
        sql = "UPDATE traxalum SET pun_l =  %s WHERE id_alumno = %s and id_trabajo = %s"
        val = (aux2, aux1[0], id)
        mycursor.execute(sql, val)
        #Actualizando su puntaje acumulado
        ac = aux2 - int(pp[1])
        print(ac)
        if trabajos[6] == 1:  # Primera Etapa..
          if ac > 0:
            print('mas')
            sql = "UPDATE matxalum SET pun_ac = (pun_ac + %s) WHERE id_alumno = %s and id_materia = %s and id_profesor = %s"  # Saco mas puntaje
            val = (ac, aux1[0], trabajos[3], trabajos[5])
            mycursor.execute(sql, val)
            mydb.commit()
          else:
            print('menos')
            ac *= -1 #sino no  resta bien el acumulado
            sql = "UPDATE matxalum SET pun_ac = (pun_ac - %s) WHERE id_alumno = %s and id_materia = %s and id_profesor = %s"  # Saco mas puntaje
            val = (ac, aux1[0], trabajos[3], trabajos[5])
            mycursor.execute(sql, val)
            mydb.commit()
        if trabajos[6] == 2: #Segunda
          if ac > 0:
            print('mas')
            sql = "UPDATE matxalum SET pun_ac2 = (pun_ac2 + %s) WHERE id_alumno = %s and id_materia = %s and id_profesor = %s"  # Saco mas puntaje
            val = (ac, aux1[0], trabajos[3], trabajos[5])
            mycursor.execute(sql, val)
            mydb.commit()
          else:
            print('menos')
            ac *= -1 #sino no  resta bien el acumulado
            sql = "UPDATE matxalum SET pun_ac2 = (pun_ac2 - %s) WHERE id_alumno = %s and id_materia = %s and id_profesor = %s"  # Saco mas puntaje
            val = (ac, aux1[0], trabajos[3], trabajos[5])
            mycursor.execute(sql, val)
            mydb.commit()
            band = 1
        return redirect(url_for('profesores.modificarproceso2', id=id))
    else:
      flash("error", "error")
  return render_template('modproceso3.html', datos=datos, trabajos =trabajos, indicadores = indicadores, alumnos=user_datos)

 #Eliminar el trabajo seleccionado
@profesores_views.route('/eliminartrabajo/<int:id>', methods = ['GET', 'POST'])
def eliminartrabajo(id):
  print(id)
  datos = session['username']
  # Saca el trabajo
  mycursor = mydb.cursor()
  sql = "SELECT id_trabajo, des_t, pun_t, id_materia, id_curso, id_profesor, etapa FROM trabajos WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  trabajos = trabajos[0]
  print(trabajos)
  # Saca los indicadores
  sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
  val = [id]
  mycursor.execute(sql, val)
  indicadores = mycursor.fetchall()
  print(indicadores)
  # Saca los alumnos relacionados con el trabajo
  sql = "SELECT nmb_a, ape_a, id_txa, traxalum.id_alumno, pun_l, fec_t, id_trabajo FROM traxalum, alumnos WHERE id_trabajo = %s and traxalum.id_alumno = alumnos.id_alumno"
  val = [id]
  mycursor.execute(sql, val)
  datas = mycursor.fetchall()
  datas = sorted(datas, key=lambda datas: datas[1])  # Ordena por orden alfabetico
  print(datas)
  if request.method == 'POST':
    #Borra los indicadores por alumno
    mycursor = mydb.cursor()
    sql = "DELETE FROM indxalum WHERE id_trabajo = %s"
    val = [id]
    mycursor.execute(sql, val)
    mydb.commit()
    # Borra el trabajo del sistema
    mycursor = mydb.cursor()
    sql = "DELETE FROM trabajos WHERE id_trabajo = %s"
    val = [id]
    mycursor.execute(sql, val)
    mydb.commit()
    # Borra el trabajo al alumno
    mycursor = mydb.cursor()
    sql = "DELETE FROM traxalum WHERE id_trabajo = %s"
    val = [id]
    mycursor.execute(sql, val)
    mydb.commit()
    # Updatea al acumulado del alumno en la materia
    for x in range(0, len(datas)):
      print(x)
      aux = datas[x]
      print(aux)
      aux2 = aux[4]
      print(aux2)
      aux3 = aux[3]
      print(aux3)
      if trabajos[6] == 1: #Primera Etapa:
        sql = "UPDATE matxalum SET pun_ac = (pun_ac - %s) WHERE id_materia = %s and id_profesor = %s and id_curso = %s and id_alumno = %s"
        val = (aux2, trabajos[3], trabajos[5], trabajos[4], aux3)
        mycursor.execute(sql, val)
        mydb.commit()
      if trabajos[6] == 2:# Segunda Etapa:
        sql = "UPDATE matxalum SET pun_ac2 = (pun_ac2 - %s) WHERE id_materia = %s and id_profesor = %s and id_curso = %s and id_alumno = %s"
        val = (aux2, trabajos[3], trabajos[5], trabajos[4])
        mycursor.execute(sql, val)
        mydb.commit()
    current_app.config['band'] = 3
    return redirect(url_for('profesores.modproceso', id = trabajos[3]))
  return render_template('modproceso4.html', datos=datos, trabajos=trabajos, indicadores=indicadores,
                         datas=datas, id = id)

#Llamar lista
@profesores_views.route('/llamarlista/<int:id>', methods = ['POST', 'GET'])
def asistenciaalumnos(id):
  datos = session['username']
  #print(id)
  #Saca el curso relacionado
  mycursor = mydb.cursor()
  sql = "SELECT des_m, des_c, des_e, cursos.id_curso FROM matxpro, materias, cursos, enfasis WHERE id_profesor = %s and matxpro.id_materia = %s and matxpro.id_materia = materias.id_materia " \
        "and matxpro.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  #Saca los alumnos matriculados
  sql = "SELECT matxalum.id_alumno, matxalum.id_materia, matxalum.id_curso, nmb_a, ape_a FROM matxalum, alumnos WHERE id_profesor = %s and matxalum.id_materia = %s and matxalum.id_alumno = alumnos.id_alumno"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  mat = mycursor.fetchall()
  print(mat)
  # Ordena alfabetaicamente los alumnos del vector...
  mat = sorted(mat, key=lambda mat: mat[4])
  #Estados posibles
  lista = ["P", "A"]
  # Extraemos la fecha
  inf = datetime.now()
  fecha = datetime.strftime(inf, '%Y/%m/%d')
  #Dia
  dia = datetime.today().weekday()
  dia += 1
  print(dia)
  # Comprueba que el profesor tenga clases hoy
  sql = "SELECT id_horario, id_curso, id_materia, id_dia, hora_i, hora_f FROM horarios WHERE id_profesor = %s and id_dia =%s and id_materia = %s and id_curso = %s ORDER BY horarios.hora_i ASC"
  val = [datos[0], dia, id, cursos[3]]
  mycursor.execute(sql, val)
  comp = mycursor.fetchall()
  print(comp)
  # Comprueba que aun no halla llamado lista hoy
  sql = "SELECT DISTINCT fecha FROM asistenciaalum WHERE fecha = %s"
  val = [fecha]
  mycursor.execute(sql, val)
  comp2 = mycursor.fetchall()
  print(comp2)
  #Alertas
  global band
  #Si no tiene clases hoy redirige
  if not comp:
    current_app.config['band'] = 7
    return redirect(url_for('profesores.miscursos'))
  #Si ya llamo clases redirige
  if comp2:
    current_app.config['band'] = 5
    return redirect(url_for('profesores.miscursos'))
  #Comprobar que cargo bien
  create = True
  if request.method == "POST":
    for x in range(0, len(mat)):
      aux = request.form.getlist(('{a}'.format(a=x + 1)))
      if not aux:
        create = False
    if create == True:
      for x in range(0, len(mat)):
        aux = request.form.getlist(('{a}'.format(a=x + 1)))
        aux2 = mat[x]
        #Guarda por alumno si estuvo presente o ausente
        mycursor.execute(
          'INSERT INTO asistenciaalum (id_alumno, id_materia, fecha, asistio, id_curso) VALUES (%s, %s, %s, %s, %s)',
          (aux2[0], id, fecha, aux[0], aux2[2]))
        mydb.commit()
      band = 6
      return redirect(url_for('profesores.miscursos'))
    else:
      flash("","ec")
  return render_template('llamarlista.html', datos=datos, cursos = cursos, alumnos = mat, listas = lista, fecha = fecha, id = id)

#Listado de asistencias llamadas por el docente
@profesores_views.route('/verasistencias/<int:id>', methods = ['POST', 'GET'])
def verasistencias(id):
  datos = session['username']
  mycursor = mydb.cursor()
  #Saca el curso y el nombre de la materia
  sql = "SELECT des_m, des_c, des_e, cursos.id_curso FROM matxpro, materias, cursos, enfasis WHERE id_profesor = %s and matxpro.id_materia = %s and matxpro.id_materia = materias.id_materia " \
        "and matxpro.id_curso = cursos.id_curso and cursos.id_enfasis = enfasis.id_enfasis"
  val = [datos[0], id]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  #Saca los datos de la asistencia
  sql = "SELECT DISTINCT fecha FROM asistenciaalum WHERE id_materia = %s and id_curso = %s  ORDER BY asistenciaalum.fecha DESC LIMIT 7"
  val = [id, cursos[3]]
  mycursor.execute(sql, val)
  asis = mycursor.fetchall()
  print(asis)
  #ID MATERIA
  global user_datos
  user_datos = id
  current_app.config['userdatos'] = id
  global vector
  vector = cursos[3]
  current_app.config['vector'] = cursos[3]
  if request.method == "POST":
    filtro = int(request.form.get("idfiltro"))
    print(filtro)
    if filtro == 1:
      return redirect(url_for('profesores.verasistencias', id = id))
    if filtro == 2:
      sql = "SELECT DISTINCT fecha FROM asistenciaalum WHERE id_materia = %s and id_curso = %s  ORDER BY asistenciaalum.fecha DESC LIMIT 31"
      val = [id, cursos[3]]
      mycursor.execute(sql, val)
      asis = mycursor.fetchall()
      print(asis)
      return render_template('verlista.html', datos=datos, cursos=cursos, listas=asis, filtro=filtro, id=id)
    if filtro == 3:
      sql = "SELECT DISTINCT fecha FROM asistenciaalum WHERE id_materia = %s and id_curso = %s  ORDER BY asistenciaalum.fecha DESC"
      val = [id, cursos[3]]
      mycursor.execute(sql, val)
      asis = mycursor.fetchall()
      print(asis)
      return render_template('verlista.html', datos=datos, cursos=cursos, listas=asis, filtro=filtro, id=id)
  return render_template('verlista.html', datos=datos, cursos = cursos, listas = asis, filtro = 1, id = id)

#Ver la asistencia selecionada
@profesores_views.route('/verificarlista/<string:id>', methods = ['POST', 'GET'])
def verificarlista(id):
  print(id)
  if not current_app.config['userdatos'] or not current_app.config['vector']:
    return redirect(url_for('profesores.miscursos'))
  datos = session['username']
  mycursor = mydb.cursor()
  # Saca los datos de la asistencia
  sql = "SELECT id_asisalum, asistenciaalum.id_alumno, asistio, nmb_a, ape_a FROM asistenciaalum, alumnos WHERE id_materia = %s and asistenciaalum.id_curso = %s and fecha = %s " \
        "and asistenciaalum.id_alumno = alumnos.id_alumno"
  val = [user_datos, vector, id]
  mycursor.execute(sql, val)
  asis = mycursor.fetchall()
  print(asis)
  # Ordena alfabetaicamente los alumnos del vector...
  asis = sorted(asis, key=lambda asis: asis[4])
  return render_template('verificarlista.html', datos=datos, listas = asis, id = id, vector = user_datos)

#Modificar lista seleccionada
@profesores_views.route('/modificarlista/<string:id>', methods = ['POST', 'GET'])
def modlista(id):
  if not current_app.config['userdatos'] or not current_app.config['vector']:
    return redirect(url_for('profesores.miscursos'))
  datos = session['username']
  mycursor = mydb.cursor()
  # Saca los datos de la asistencia
  sql = "SELECT id_asisalum, asistenciaalum.id_alumno, asistio, nmb_a, ape_a FROM asistenciaalum, alumnos WHERE id_materia = %s and asistenciaalum.id_curso = %s and fecha = %s " \
        "and asistenciaalum.id_alumno = alumnos.id_alumno"
  val = [user_datos, vector, id]
  mycursor.execute(sql, val)
  asis = mycursor.fetchall()
  # Ordena alfabetaicamente los alumnos del vector...
  asis = sorted(asis, key=lambda asis: asis[4])
  print(asis)
  alumnos = []
  if request.method == 'POST':
    # Saca los ids de los alumnos
    idalum = request.form.getlist(('idalum'))
    print(idalum)
    # Saca los alumnos seleccionados por el docente
    for x in range(0, len(idalum)):
      aux = int(idalum[x])
      print(aux)
      mycursor = mydb.cursor()
      sql = "SELECT id_alumno,nmb_a, ape_a FROM alumnos WHERE id_alumno = %s"
      val = [aux]
      mycursor.execute(sql, val)
      aux2 = mycursor.fetchall()
      aux2 = aux2[0]
      # print(aux2)
      alumnos.append(aux2)
    print(alumnos)
    # Ordena alfabetaicamente los alumnos del vector...
    alumnos = sorted(alumnos, key=lambda alumnos: alumnos[2])
    current_app.config['vector2'] = alumnos
    return redirect(url_for('profesores.carmodlista', id=id))
  return render_template('modlista.html', datos=datos, listas=asis, id=id, vector = user_datos)

#Modificar alumnos selecionados
@profesores_views.route('/cargarmodificacion/<string:id>', methods = ['POST', 'GET'])
def carmodlista(id):
  datos = session['username']
  print(id)
  #alumnos
  print(current_app.config['vector2'])
  if not current_app.config['userdatos']:
    return redirect(url_for('profesores.miscursos'))
  if not current_app.config['userdatos'] or not current_app.config['vector'] or not current_app.config['vector2']:
    return redirect(url_for('profesores.miscursos'))
  mycursor = mydb.cursor()
  # Estados posibles
  lista = ["P", "A"]
  # Comprobar que cargo bien
  create = True
  #alertas
  global band
  if request.method == "POST":
    for x in range(0, len(current_app.config['vector2'])):
      aux = request.form.getlist(('{a}'.format(a=x + 1)))
      if not aux:
        create = False
    if create == True:
      for x in range(0, len(current_app.config['vector2'])):
        aux = request.form.getlist(('{a}'.format(a=x + 1)))
        aux2 = current_app.config['vector2'][x]
        # Actualiza por alumno si estuvo presente o ausente
        sql = "UPDATE asistenciaalum SET asistio = %s WHERE id_alumno = %s and id_materia = %s and id_curso = %s and fecha = %s"
        val = (aux[0], aux2[0], user_datos, vector, id)
        mycursor.execute(sql, val)
        mydb.commit()
        mydb.commit()
      band = 6
      return redirect(url_for('profesores.modlista', id=id))
    else:
      flash("", "ec")
  return render_template('modlista2.html', datos=datos, alumnos = current_app.config['vector2'], listas = lista, id = id, vector = id)

#El profesor ve su asistencia
@profesores_views.route('/vermiasistencia', methods = ['POST', 'GET'])
def verasistenciaprofe():
  datos = session['username']
  print(datos)
  #Sacamos las veces que marco asistencia
  mycursor = mydb.cursor()
  sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC LIMIT 7"
  val = [datos[0]]
  mycursor.execute(sql, val)
  asistencias = mycursor.fetchall()
  print(asistencias)
  dt = 0  # Dias trabajados
  dp = 0  # Dias no trabajados
  dr = 0  # Dias que no marco salida
  for x in range(0, len(asistencias)):
    aux = asistencias[x]
    if aux[0] or aux[1]:
      dt += 1
    if not aux[0] and not aux[1]:
      dp += 1
    if aux[0] and not aux[1]:
      dr += 1
  print(dt, dp, dr)
  if request.method == "POST":
    dt = 0
    dp = 0
    dr = 0
    filtro = int(request.form.get("idfiltro"))
    print(filtro)
    if filtro == 2:
      sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC LIMIT 31"
      val = [datos[0]]
      mycursor.execute(sql, val)
      asistencias = mycursor.fetchall()
      for x in range(0 , len(asistencias)):
        aux = asistencias[x]
        if aux[0] or aux[1]:
          dt += 1
        if not aux[0] and not aux[1]:
          dp += 1
        if aux[0] and not aux[1]:
          dr += 1
      return render_template('verasistenciaprofe.html', datos=datos, asistencias=asistencias, filtro=filtro, dt = dt, dp = dp, dr = dr)
    if filtro == 3:
      sql = "SELECT hora_e, hora_s, fec_a FROM asistenciaprof WHERE id_profesor = %s ORDER BY asistenciaprof.id_asisprof DESC"
      val = [datos[0]]
      mycursor.execute(sql, val)
      asistencias = mycursor.fetchall()
      for x in range(0 , len(asistencias)):
        aux = asistencias[x]
        if aux[0] or aux[1]:
          dt += 1
        if not aux[0] and not aux[1]:
          dp += 1
        if aux[0] and not aux[1]:
          dr += 1
      return render_template('verasistenciaprofe.html', datos=datos, asistencias=asistencias, filtro=filtro, dt = dt, dp = dp, dr = dr)
    if filtro == 1:
      return redirect(url_for('profesores.verasistenciaprofe'))
  return render_template('verasistenciaprofe.html', datos=datos, asistencias=asistencias, filtro = 1, dt = dt, dp = dp, dr = dr)

#Parte de Proceso de Alumnos, Carga de trabajos
@profesores_views.route('/proceso', methods = ['GET', 'POST']) #Procesos ✔
def proceso():
  datos = session['username']
  #print(datos)
  mycursor = mydb.cursor()
  sql = "SELECT * FROM matxpro WHERE id_profesor = %s"
  val = [datos[0]]
  mycursor.execute(sql, val)
  profe = mycursor.fetchall()
  #print(profe) #ID, Id Materia, Id profe, Carga horaria, Id curso
  profemat = []
  profemati = []
  cantm = 0
  for x in range(0, len(profe)):
    #print(x)
    aux = profe[x]
    sql = "SELECT * FROM materias WHERE id_materia = %s"
    val = [aux[1]]
    mycursor.execute(sql, val)
    aux = mycursor.fetchall()
    aux = aux[0]
    #print(aux)
    #print(x)
    if x == 0:
      profemat.append(aux)
      #print(profemat)
      cantm += 1
      #print(x)
      #print(cantm)
    elif profemat[x-cantm-1] != aux:
      #print(cantm)
      cantm += 1
      profemat.append(aux)
      #print(profemat)
  #print(profemat)
  task = userform.Task(request.form)
  if request.method == 'POST' and task.validate(): #ID prof = datos[0],
    #Informacion del trabajo
    print("Info: ")
    nombre_t = task.nombre.data
    print(nombre_t)
    tipo_t = request.form.get(('tipo_t'))
    print(tipo_t)
    puntaje_t = task.puntaje.data
    puntaje_t = int(puntaje_t)
    print(puntaje_t)
    etapa_t = request.form.get(('etapa_t'))
    print(etapa_t)
    print(current_app.config['idcurso'])
    # Informacion del trabajo
    print("Indicadores: ")
    indicadores = []
    puntaje_i = []
    create = True
    for x in range(1, 11):
      if x == 1:
        i1 = request.form.get(('i1'))
        p1 = request.form.get(('p1'))
        if i1 and p1:
          indicadores.append(i1)
          puntaje_i.append(p1)
      if x == 2:
        i2 = request.form.get(('i2'))
        p2 = request.form.get(('p2'))
        if i2 and p2:
          indicadores.append(i2)
          puntaje_i.append(p2)
      if x == 3:
        i3 = request.form.get(('i3'))
        p3 = request.form.get(('p3'))
        if i3 and p3:
          indicadores.append(i3)
          puntaje_i.append(p3)
      if x == 4:
        i4 = request.form.get(('i4'))
        p4 = request.form.get(('p4'))
        if i4 and p4:
          indicadores.append(i4)
          puntaje_i.append(p4)
      if x == 5:
        i5 = request.form.get(('i5'))
        p5 = request.form.get(('p5'))
        if i5 and p5:
          indicadores.append(i5)
          puntaje_i.append(p5)
      if x == 6:
        i6 = request.form.get(('i6'))
        p6 = request.form.get(('p6'))
        if i6 and p6:
          indicadores.append(i6)
          puntaje_i.append(p6)
      if x == 7:
        i7 = request.form.get(('i7'))
        p7 = request.form.get(('p7'))
        if i7 and p7:
          indicadores.append(i7)
          puntaje_i.append(p7)
      if x == 8:
        i8 = request.form.get(('i8'))
        p8 = request.form.get(('p8'))
        if i8 and p8:
          indicadores.append(i8)
          puntaje_i.append(p8)
      if x == 9:
        i9 = request.form.get(('i9'))
        p9 = request.form.get(('p9'))
        if i9 and p9:
          indicadores.append(i9)
          puntaje_i.append(p9)
      if x == 10:
        i10 = request.form.get(('i10'))
        p10 = request.form.get(('p10'))
        if i10 and p10:
          indicadores.append(i10)
          puntaje_i.append(p10)
    tt = 0
    for x in range(0, len(puntaje_i)):
      aux = int(puntaje_i[x])
      tt = tt + aux
    print(indicadores)
    print(puntaje_i)
    print(tt)
    #Comprobar si el puntaje de los indicadores es mayor que el puntaje del trabajo o si no esta completo
    if tt > puntaje_t:
      create = False
      flash("pun_m", "pun_m")
      #flash("")
    if tt != puntaje_t:
      create = False
      print("El puntaje no coincide")
      flash("pun_m", "pun_m")
    fecha = datetime.now()
    fecha = datetime.strftime(fecha, '%Y/%m/%d')
    if current_app.config['idmaterias'] != 0:
      if current_app.config['idcurso'] !=0:
        if create == True:
          clavet = crearclavet()
          #Guardamos el trabajo
          mycursor = mydb.cursor()
          mycursor.execute(
            'INSERT INTO trabajos (fec_t, des_t, pun_t, id_profesor, id_materia, tipo_t, clave_t, id_curso, etapa) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)',
            (fecha, nombre_t, puntaje_t, datos[0], current_app.config['idmaterias'], tipo_t, clavet, current_app.config['idcurso'], etapa_t))
          mydb.commit()
          # Saca el id del trabajo
          sql = "SELECT id_trabajo, id_curso FROM trabajos WHERE id_materia = %s and id_profesor = %s and fec_t = %s and des_t= %s and pun_t = %s and tipo_t = %s and clave_t = %s"
          val = [current_app.config['idmaterias'], datos[0], fecha, task.nombre.data, task.puntaje.data, tipo_t, clavet]
          mycursor.execute(sql, val)
          idtra = mycursor.fetchall()
          idtra = idtra[0]
          print(idtra[0])
          #Guarda los indicadores
          for x in range(0, len(puntaje_i)):
            aux = indicadores[x]
            aux2 = puntaje_i[x]
            #print(aux)
            #print(aux2)
            mycursor.execute(
              'INSERT INTO indicadores (des_i, pun_i, id_trabajo) VALUES (%s, %s, %s)',
              (aux, aux2, idtra[0]))
            mydb.commit()
            print(idtra[0])
          return redirect(url_for('profesores.cargarpuntaje', id = idtra[0]))
      else:
        # Errorsito
        flash("Error", "cur")
    else:
      # Errorsito
      flash("Error", "mat")
  return render_template('procesotest.html', task = task, datos = datos, materias = profemat, bcurso = current_app.config['bcurso'], cursos = current_app.config['cursos'],
                         alumnos = current_app.config['alumnos'], balumno = current_app.config['balumno'] )

#Crea una clave random para el trabajo
def crearclavet():
  length = 5
  letters = string.ascii_letters + string.digits
  clave = ''.join(random.choice(letters) for i in range(length))
  return clave

@profesores_views.route('/procesoss', methods = ['POST']) #Parte de procesos, aca quitamos la materia ✔
def procesoss():
  if request.method == 'POST':
    id_mat = request.form.get(('idmat'))
    if id_mat:
      current_app.config['bcurso'] = 1
    print(id_mat)
    datos = session['username']
    mycursor = mydb.cursor()
    print(datos[0])
    print(id_mat)
    sql = "SELECT * FROM matxpro WHERE id_profesor = %s and id_materia =%s"
    val = [datos[0], id_mat]
    current_app.config['idmaterias'] = id_mat
    mycursor.execute(sql, val)
    data = mycursor.fetchall()
    print("Cursos con esa materia:")
    print(data)
    #print(cursos[0])
    #print(len(cursos))
    current_app.config['cursos'] = []
    if current_app.config['cursos']:
      current_app.config['balumno'] = 1
    for x in range(0, len(data)):
      print(x)
      aux = data[x]
      mycursor = mydb.cursor()
      sql = "SELECT id_curso, des_c, sec_c,des_e FROM cursos,enfasis WHERE id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
      val = [aux[3]]
      mycursor.execute(sql, val)
      aux = mycursor.fetchall()
      aux = aux[0]
      current_app.config['cursos'].append(aux)
    print("Cursos procesados:")
    print(current_app.config['cursos'])
  return redirect(url_for('profesores.proceso'))

@profesores_views.route('/alumnos', methods = ['POST']) #Parte de procesos, quitamos los cursos y los alumnos relacionados con la materia ✔
def alumnosproceso():
  if request.method == 'POST':
    id_curso = request.form.get(('idcurso'))
    datos = session['username']
    current_app.config['balumno'] = 1
    mycursor = mydb.cursor()
    sql = "SELECT alumnos.id_alumno,nmb_a, ape_a FROM alumnos, " \
          "matxalum WHERE matxalum.id_curso =%s and alumnos.id_alumno = matxalum.id_alumno " \
          "and matxalum.id_materia = %s"
    val = [id_curso, current_app.config['idmaterias']]
    mycursor.execute(sql, val)
    current_app.config['idcurso'] = id_curso
    data = mycursor.fetchall()
    data = sorted(data, key=lambda data: data[2])  # Ordena por orden alfabetico
    print(id_curso)
    print(current_app.config['idmaterias'])
    print("Alumnos con esta materia:")
    print(data)
    current_app.config['alumnos'] = data
  return redirect(url_for('profesores.proceso'))

# Carga de trabajo y indicadores ✔
@profesores_views.route('/cargarpuntaje/<int:id>', methods =  ['GET', 'POST'])
def cargarpuntaje(id):
  datos = session['username']
  idtra = id
  print(idtra)
  # Saca el trabajo
  mycursor = mydb.cursor()
  sql = "SELECT id_trabajo, des_t, pun_t, id_materia, id_curso, id_profesor, etapa FROM trabajos WHERE id_trabajo = %s"
  val = [idtra]
  mycursor.execute(sql, val)
  trabajos = mycursor.fetchall()
  trabajos = trabajos[0]
  print(trabajos)
  # Saca los indicadores
  sql = "SELECT * FROM indicadores WHERE id_trabajo = %s"
  val = [idtra]
  mycursor.execute(sql, val)
  indicadores = mycursor.fetchall()
  print(indicadores)
  # Saca los alumnos
  mycursor = mydb.cursor()
  sql = "SELECT alumnos.id_alumno,nmb_a, ape_a FROM alumnos, " \
        "matxalum WHERE matxalum.id_curso =%s and alumnos.id_alumno = matxalum.id_alumno " \
        "and matxalum.id_materia = %s"
  val = [trabajos[4], trabajos[3]]
  mycursor.execute(sql, val)
  alumnos = mycursor.fetchall()
  alumnos = sorted(alumnos, key=lambda alumnos: alumnos[2])
  print(alumnos)
  # Saca el curso:
  sql = "SELECT id_curso, des_c, sec_c, des_e FROM cursos, enfasis WHERE cursos.id_curso = %s and cursos.id_enfasis = enfasis.id_enfasis"
  val = [trabajos[4]]
  mycursor.execute(sql, val)
  cursos = mycursor.fetchall()
  cursos = cursos[0]
  print(cursos)
  #Comprobar si se cargo correctamente
  create = True
  if request.method == 'POST':
    puntajes = []
    pl = 0
    # Comprobar que se hayan cargado todos los indicadores
    for x in range(0, len(alumnos)):
      for y in range(1, len(indicadores) + 1):
        aux = request.form.getlist(('{a}{i}'.format(a=x + 1, i=y)))
        if not aux:
          create = False
    # Solo si esta cargado correctamente
    if create == True:
      #Carga de indicador por alumno, puntaje logrado por indicador y acumular puntaje total logrado
      for x in range(0, len(alumnos)):
        aux1 = alumnos[x]
        #print(aux1)
        for y in range(1, len(indicadores) + 1):
          aux2 = indicadores[y - 1]
          #print(aux2)
          if y == 1:
            pl = 0
            #print("Alumno {}".format(alumnos[x]))
          aux = request.form.getlist(('{a}{i}'.format(a = x + 1, i = y)))
          #print('{a}{i}'.format(a=x + 1, i=y))
          #print(aux)
          p = int(aux[0])
          #print(p)
          pl +=int(aux[0])
          mycursor.execute(
            'INSERT INTO indxalum (id_indicador, id_trabajo, id_alumno, pun_l) VALUES (%s, %s, %s, %s)',
            (aux2[0], trabajos[0], aux1[0], p))
          mydb.commit()
          if y == len(indicadores):
            puntajes.append(pl)
      print(puntajes)
      fecha = datetime.now()
      fecha = datetime.strftime(fecha, '%Y')
      #print(fecha)
      for x in range(0, len(alumnos)):
        aux1 = alumnos[x]
        aux2 = puntajes[x]
        #Cargando el trabajo al alumno con su puntaje logrado
        mycursor.execute(
          'INSERT INTO traxalum (id_alumno, pun_l, fec_t, id_trabajo) VALUES (%s, %s, %s, %s)',
          (aux1[0], aux2, fecha, trabajos[0]))
        mydb.commit()
        #Sumando a su puntaje acumulado
        if trabajos[6] == 1: #Primera Etapa..
          sql = "UPDATE matxalum SET pun_ac = (pun_ac + %s) WHERE id_alumno = %s and id_materia = %s and id_curso = %s and id_profesor = %s"
          val = (aux2, aux1[0],  trabajos[3], trabajos[4], trabajos[5])
          mycursor.execute(sql, val)
          mydb.commit()
        if trabajos[6] == 2:
          sql = "UPDATE matxalum SET pun_ac2 = (pun_ac2 + %s) WHERE id_alumno = %s and id_materia = %s and id_curso = %s and id_profesor = %s"
          val = (aux2, aux1[0], trabajos[3], trabajos[4], trabajos[5])
          mycursor.execute(sql, val)
          mydb.commit()
      #Mandar mensaje flash..
      global band
      current_app.config['band'] = 1
      return redirect(url_for('profesores.proceso'))
    else:
      flash("error","error")
  return render_template('cargartrabajo.html', datos = datos, trabajo = trabajos, alumnos = alumnos, indicadores= indicadores,
                         cursos = cursos)

#Enviar planillas docente
@profesores_views.route('/enviarplanilla', methods = ['POST', 'GET'])
def enviarplanilla():
  datos = session['username']
  #Si el profesor mando planillas
  mycursor = mydb.cursor()
  sql = "SELECT estado, fecha_m, fecha_r, des_p, cp FROM planillas WHERE id_profesor = %s LIMIT 45"
  val = [datos[0]]
  mycursor.execute(sql, val)
  planillas = mycursor.fetchall()
  print(planillas)
  if planillas:
    print("ya tiene")
    cp = planillas[len(planillas) - 1]
    cp = int(cp[4]) + 1
  else:
    cp = 1
  if request.method == "POST":
    planilla = request.files["planilla"]
    desc = request.form.get("desc")
    if "planilla" not in request.files:
      print("No envio nada")
      pass
    elif planilla.filename == "":
      print("No mando nada")
      pass
    elif planilla and archpermi2(planilla.filename):
      filename = planilla.filename.split('.')
      ext = filename[len(filename) - 1]
      print(ext)
      print(filename)
      filename = "planillapendiente_" + str(datos[0]) + str(cp) + "." + ext
      print(filename)
      print(desc)
      planilla.save(os.path.join(os.path.abspath("./static/resources/control_planillas"), filename))
      flash("", "")  # Ya guarda el archivo
      estado = 0
      inf = datetime.now()
      # Extraemos la fecha
      fecha = datetime.strftime(inf, '%Y/%m/%d')
      mycursor = mydb.cursor()
      if planillas:
        print("ya tiene")
        mycursor.execute(
          'INSERT INTO planillas (filename, estado, id_profesor, fecha_m, des_p, cp) VALUES (%s, %s, %s, %s, %s, %s)',
          (filename, estado, datos[0], fecha, desc, cp))
        mydb.commit()
      else:
        print("no tiene")
        mycursor.execute(
          'INSERT INTO planillas (filename, estado, id_profesor, fecha_m, des_p, cp) VALUES (%s, %s, %s, %s, %s, %s)',
          (filename, estado, datos[0], fecha, desc, 1))
        mydb.commit()
      print("mandado")
      return redirect(url_for('profesores.enviarplanilla'))
    else:
      print("archivo no permitido")
  return render_template('enviarplanillas.html', datos=datos, planillas = planillas)


#Para planillas, arhivos permitidos
ext_c = set(["xls", "xlsx", "xlsm", "xlsb", "xltx"])
def archpermi2(filename):
  filename = filename.split('.')
  if filename[len(filename) - 1] in ext_c:
      return True
  return False
